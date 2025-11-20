import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
from datetime import datetime


from collections import defaultdict
from cloud_sync import SimpleCloudSync

class CustomOrderManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("定制拆单工作室记账工具")
        self.root.geometry("1400x900")
        
        # 数据存储
        self.orders = {}  # 订单数据
        self.manufacturers = {}  # 厂家数据
        try:
            import sys as _sys
        except Exception:
            _sys = None
        
        # 修复文件路径问题 - 增强版路径检测和修复
        current_file = os.path.abspath(__file__)
        current_dir = os.path.dirname(current_file)
        working_dir = os.getcwd()
        is_frozen = hasattr(_sys, 'frozen') and _sys.frozen
        
        print(f"路径检测 - 运行模式: {'打包EXE' if is_frozen else 'Python脚本'}")
        print(f"路径检测 - 当前文件: {current_file}")
        print(f"路径检测 - 工作目录: {working_dir}")
        
        # 候选路径列表（按优先级排序）
        candidate_paths = []
        
        if is_frozen:
            # 打包模式下，优先使用可执行文件所在目录
            exe_dir = os.path.dirname(_sys.executable)
            candidate_paths.append(os.path.join(exe_dir, "data.json"))
            print(f"路径检测 - EXE目录: {exe_dir}")
        
        # 添加当前脚本所在目录
        candidate_paths.append(os.path.join(current_dir, "data.json"))
        
        # 添加工作目录
        candidate_paths.append(os.path.join(working_dir, "data.json"))
        
        # 用户数据目录（备用方案）
        user_data_dir = os.path.join(os.path.expanduser("~"), "面积计算工具")
        user_data_file = os.path.join(user_data_dir, "data.json")
        candidate_paths.append(user_data_file)
        
        # 测试每个候选路径的可写性
        self.data_file = None
        for path in candidate_paths:
            dir_path = os.path.dirname(path)
            
            # 确保目录存在
            try:
                os.makedirs(dir_path, exist_ok=True)
            except:
                continue
            
            # 测试写入权限
            try:
                test_file = os.path.join(dir_path, "test_write.tmp")
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                
                # 检查是否已存在数据文件
                if os.path.exists(path):
                    print(f"路径检测 - 发现现有数据文件: {path}")
                
                self.data_file = path
                print(f"路径检测 - 选择路径: {self.data_file}")
                break
                
            except (OSError, PermissionError):
                continue
        
        # 如果所有路径都失败，使用最后的选择
        if not self.data_file:
            self.data_file = candidate_paths[-1]
            print(f"路径检测 - 使用最后选择: {self.data_file}")
            # 确保目录存在
            try:
                os.makedirs(os.path.dirname(self.data_file), exist_ok=True)
            except:
                pass
        
        self.data_loaded = False  # 数据加载状态标志
        self.bound_order_dir = ""
        
        # 厂家配置相关
        self.app_config_file = os.path.join(os.path.dirname(self.data_file), "app_config.json")
        self.current_manufacturer = None  # 当前登录的厂家
        self.is_admin = False  # 是否为管理员
        self.admin_password = "627813"  # 管理员密码
        
        # 云同步管理器
        self.cloud_sync = SimpleCloudSync(self.data_file)
        
        # 程序关闭处理
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.unsaved_changes = False  # 标记是否有未保存的更改
        
        # 先创建界面
        self.create_widgets()
        
        # 延迟加载数据和刷新界面（尽快在空闲时执行）
        self.root.after_idle(self.delayed_initialization)
        
    def delayed_initialization(self):
        """延迟初始化：加载数据并刷新界面"""
        print("开始加载数据...")
        
        # 更新云同步状态显示
        self.update_sync_status_display()
        
        # 首先检查云同步配置
        if not self.cloud_sync or not self.cloud_sync.github_sync:
            print("未配置云同步，允许用户选择是否配置")
            # 显示云同步配置提醒（允许跳过）
            self.root.after(1000, self.show_cloud_sync_config_optional)
            return
        
        # 已配置云同步，尝试从云端下载数据
        print("🔄 检测到云同步配置，正在尝试从云端下载最新数据...")
        try:
            # 尝试从云端下载数据
            cloud_data = self.cloud_sync.sync_down()
            if cloud_data:
                # 成功下载云端数据，使用云端数据
                self.orders = cloud_data.get("orders", {})
                self.manufacturers = cloud_data.get("manufacturers", {})
                self.bound_order_dir = cloud_data.get("bound_order_dir", "")
                # 保存到本地（不进行云同步）
                self.save_data_local_only()
                print("✅ 已从云端成功下载并加载最新数据")
                print(f"📊 下载数据包含: {len(self.orders)} 个订单, {len(self.manufacturers)} 个厂家")
            else:
                # 云端没有数据，使用本地数据
                print("☁️ 云端没有数据，使用本地数据")
                self.load_data_local_only()
                
        except Exception as e:
            print(f"❌ 从云端下载数据失败: {e}，使用本地数据")
            self.load_data_local_only()
        
        # 现在检查厂家配置（从已加载的数据中获取厂家信息）
        if not self.load_app_config():
            # 首次运行，需要配置厂家（此时已有厂家数据）
            self.root.after(1000, self.show_manufacturer_config_after_sync)
            return
        
        # 已有厂家配置，继续正常流程
        self.continue_normal_startup()
    
    def update_sync_status_display(self):
        """更新云同步状态显示"""
        if self.cloud_sync.github_sync:
            repo = self.cloud_sync.github_sync.repo
            if self.cloud_sync.auto_sync:
                self.sync_status_label.config(text=f"已配置: {repo}", foreground="green")
            else:
                self.sync_status_label.config(text=f"已配置: {repo} (手动)", foreground="blue")
        else:
            self.sync_status_label.config(text="未配置云同步", foreground="gray")
    
    def load_app_config(self):
        """加载应用配置（厂家信息等）"""
        try:
            if os.path.exists(self.app_config_file):
                with open(self.app_config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                self.current_manufacturer = config.get("current_manufacturer")
                self.is_admin = config.get("is_admin", False)
                
                # 更新窗口标题显示当前厂家
                if self.current_manufacturer:
                    self.root.title(f"定制拆单工作室记账工具 - {self.current_manufacturer}")
                elif self.is_admin:
                    self.root.title("定制拆单工作室记账工具 - 管理员")
                
                return True
        except Exception as e:
            print(f"加载应用配置失败: {e}")
            return False
    
    def clear_manufacturer_input(self):
        """清空厂家输入框"""
        self.manufacturer_name_var.set("")
        self.unit_price_var.set("")
        self.permission_var.set("读写")  # 重置权限为默认值
        # 清除列表选择
        if hasattr(self, 'manufacturers_tree') and self.manufacturers_tree.selection():
            self.manufacturers_tree.selection_remove(self.manufacturers_tree.selection())
    
    def save_app_config(self):
        """保存应用配置"""
        try:
            config = {
                "current_manufacturer": self.current_manufacturer,
                "is_admin": self.is_admin
            }
            with open(self.app_config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存应用配置失败: {e}")
            return False
    
    def show_cloud_sync_config_first(self):
        """首先显示云同步配置对话框"""
        result = messagebox.askyesno(
            "配置云同步",
            "欢迎使用！请先配置云同步功能。\n\n"
            "配置后可以实现：\n"
            "• 从GitHub下载最新的厂家数据\n"
            "• 多设备数据同步\n"
            "• 数据云端备份\n\n"
            "是否现在配置云同步？"
        )
        
        if result:
            self.setup_cloud_sync()
        else:
            messagebox.showwarning("警告", "必须先配置云同步才能使用本软件")
            self.root.quit()
    
    def show_cloud_sync_config_optional(self):
        """显示可选的云同步配置对话框 - 允许用户跳过"""
        result = messagebox.askyesno(
            "配置云同步",
            "欢迎使用！是否配置云同步功能？\n\n"
            "配置后可以实现：\n"
            "• 从GitHub下载最新的厂家数据\n"
            "• 多设备数据同步\n"
            "• 数据云端备份\n\n"
            "您也可以选择跳过，稍后在设置中配置。\n\n"
            "是否现在配置云同步？"
        )
        
        if result:
            # 用户选择配置
            self.setup_cloud_sync()
        else:
            # 用户选择跳过 - 继续使用本地数据
            print("用户选择跳过云同步配置，使用本地数据")
            # 直接加载本地数据并继续
            self.load_data_local_only()
            # 继续检查厂家配置
            if not self.load_app_config():
                # 首次运行，需要配置厂家
                self.root.after(1000, self.show_manufacturer_config_after_sync)
            else:
                # 已有厂家配置，继续正常流程
                self.continue_normal_startup()
    
    def show_manufacturer_config_after_sync(self):
        """云同步配置完成后，显示厂家配置对话框"""
        config_window = tk.Toplevel(self.root)
        config_window.title("配置厂家身份")
        config_window.geometry("400x350")
        config_window.transient(self.root)
        config_window.grab_set()
        
        # 居中显示
        config_window.update_idletasks()
        x = (config_window.winfo_screenwidth() - config_window.winfo_width()) // 2
        y = (config_window.winfo_screenheight() - config_window.winfo_height()) // 2
        config_window.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(config_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="云同步配置完成！请选择身份", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 说明文字
        info_label = ttk.Label(main_frame, text="请选择您的身份（从已下载的数据中选择）：", 
                               font=("Arial", 10))
        info_label.pack(pady=5)
        
        # 从已加载的数据中获取厂家列表
        manufacturers = list(self.manufacturers.keys()) if self.manufacturers else []
        
        if not manufacturers:
            # 如果没有厂家数据，显示警告
            warning_label = ttk.Label(main_frame, 
                                    text="警告：未找到厂家数据！\n请检查云同步配置或联系管理员。", 
                                    foreground="red", font=("Arial", 10))
            warning_label.pack(pady=10)
            
            # 只有管理员选项
            admin_frame = ttk.LabelFrame(main_frame, text="管理员登录", padding="10")
            admin_frame.pack(fill=tk.X, pady=10)
            
            ttk.Button(admin_frame, text="管理员登录", 
                      command=lambda: self.show_admin_login(config_window)).pack(pady=5)
            
        else:
            # 厂家选择
            manufacturer_frame = ttk.LabelFrame(main_frame, text="选择厂家", padding="10")
            manufacturer_frame.pack(fill=tk.X, pady=10)
            
            manufacturer_var = tk.StringVar()
            manufacturer_combo = ttk.Combobox(manufacturer_frame, textvariable=manufacturer_var, 
                                          values=manufacturers, state="readonly", width=30)
            manufacturer_combo.pack(fill=tk.X, pady=5)
            if manufacturers:
                manufacturer_combo.current(0)
            
            # 管理员登录按钮
            admin_frame = ttk.LabelFrame(main_frame, text="管理员登录", padding="10")
            admin_frame.pack(fill=tk.X, pady=10)
            
            ttk.Button(admin_frame, text="管理员登录", 
                      command=lambda: self.show_admin_login(config_window)).pack(pady=5)
        
        def save_manufacturer_config():
            """保存厂家配置"""
            selected_manufacturer = manufacturer_var.get() if manufacturers else None
            
            if not selected_manufacturer and not self.is_admin:
                messagebox.showwarning("警告", "请选择厂家或以管理员身份登录")
                return
            
            if selected_manufacturer:
                self.current_manufacturer = selected_manufacturer
                self.is_admin = False
                self.root.title(f"定制拆单工作室记账工具 - {selected_manufacturer}")
            
            # 保存配置
            self.save_app_config()
            
            # 关闭配置窗口
            config_window.destroy()
            
            # 继续正常启动流程
            self.continue_normal_startup()
        
        # 按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        if manufacturers:
            ttk.Button(button_frame, text="确定", command=save_manufacturer_config).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="退出", command=lambda: self.root.quit()).pack(side=tk.LEFT, padx=5)
    
    def show_first_time_config(self):
        """首次运行时显示厂家配置对话框"""
        config_window = tk.Toplevel(self.root)
        config_window.title("首次配置 - 选择厂家")
        config_window.geometry("400x300")
        config_window.transient(self.root)
        config_window.grab_set()
        
        # 居中显示
        config_window.update_idletasks()
        x = (config_window.winfo_screenwidth() - config_window.winfo_width()) // 2
        y = (config_window.winfo_screenheight() - config_window.winfo_height()) // 2
        config_window.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(config_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="欢迎使用！请进行初始配置", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 说明文字
        info_label = ttk.Label(main_frame, text="请选择您的身份：", 
                               font=("Arial", 10))
        info_label.pack(pady=5)
        
        # 加载数据获取厂家列表
        try:
            self.load_data_local_only()
            manufacturers = list(self.manufacturers.keys()) if self.manufacturers else []
        except:
            manufacturers = []
        
        # 厂家选择
        manufacturer_frame = ttk.LabelFrame(main_frame, text="选择厂家", padding="10")
        manufacturer_frame.pack(fill=tk.X, pady=10)
        
        manufacturer_var = tk.StringVar()
        if manufacturers:
            manufacturer_combo = ttk.Combobox(manufacturer_frame, textvariable=manufacturer_var, 
                                            values=manufacturers, state="readonly", width=30)
            manufacturer_combo.pack(fill=tk.X, pady=5)
            if manufacturers:
                manufacturer_combo.current(0)
        else:
            ttk.Label(manufacturer_frame, text="暂无厂家数据，请先添加厂家", foreground="red").pack(pady=5)
        
        # 管理员登录按钮
        admin_frame = ttk.LabelFrame(main_frame, text="管理员登录", padding="10")
        admin_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(admin_frame, text="管理员登录", command=lambda: self.show_admin_login(config_window)).pack(pady=5)
        
        def save_manufacturer_config():
            """保存厂家配置"""
            selected_manufacturer = manufacturer_var.get()
            if not selected_manufacturer and not self.is_admin:
                messagebox.showwarning("警告", "请选择厂家或以管理员身份登录")
                return
            
            if selected_manufacturer:
                self.current_manufacturer = selected_manufacturer
                self.is_admin = False
                self.root.title(f"定制拆单工作室记账工具 - {selected_manufacturer}")
            
            # 保存配置
            self.save_app_config()
            
            # 关闭配置窗口
            config_window.destroy()
            
            # 继续正常启动流程
            self.continue_normal_startup()
        
        # 按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="确定", command=save_manufacturer_config).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=lambda: self.root.quit()).pack(side=tk.LEFT, padx=5)
    
    def show_admin_login(self, parent_window):
        """显示管理员登录对话框"""
        login_window = tk.Toplevel(parent_window)
        login_window.title("管理员登录")
        login_window.geometry("300x200")
        login_window.transient(parent_window)
        login_window.grab_set()
        
        main_frame = ttk.Frame(login_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="管理员登录", font=("Arial", 12, "bold"))
        title_label.pack(pady=10)
        
        # 密码输入
        password_frame = ttk.Frame(main_frame)
        password_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(password_frame, text="管理员密码:").pack(side=tk.LEFT, padx=5)
        password_var = tk.StringVar()
        password_entry = ttk.Entry(password_frame, textvariable=password_var, show="*", width=20)
        password_entry.pack(side=tk.LEFT, padx=5)
        
        def verify_admin():
            password = password_var.get()
            if password == self.admin_password:
                self.is_admin = True
                self.current_manufacturer = None
                self.root.title("定制拆单工作室记账工具 - 管理员")
                login_window.destroy()
                messagebox.showinfo("成功", "管理员登录成功！")
            else:
                messagebox.showerror("错误", "密码错误！")
        
        # 按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="登录", command=verify_admin).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=login_window.destroy).pack(side=tk.LEFT, padx=5)
        
        # 聚焦到密码输入框
        password_entry.focus()
    
    def get_filtered_orders(self):
        """获取根据权限过滤后的订单"""
        if self.is_admin:
            # 管理员可以看到所有订单
            return self.orders
        elif self.current_manufacturer:
            # 厂家只能看到自己的订单
            filtered_orders = {}
            for order_id, order in self.orders.items():
                if order.get('manufacturer') == self.current_manufacturer:
                    filtered_orders[order_id] = order
            return filtered_orders
        else:
            # 未配置厂家，返回空
            return {}
    
    def get_filtered_manufacturers(self):
        """获取根据权限过滤后的厂家"""
        if self.is_admin:
            # 管理员可以看到所有厂家
            return self.manufacturers
        elif self.current_manufacturer:
            # 厂家用户：根据权限过滤
            filtered_manufacturers = {}
            
            # 如果当前厂家存在且有权限
            if self.current_manufacturer in self.manufacturers:
                manufacturer_data = self.manufacturers[self.current_manufacturer]
                permission = manufacturer_data.get("permission", "读写")
                
                # 如果有读权限，可以看到自己的厂家信息
                if "读" in permission or permission == "读写":
                    filtered_manufacturers[self.current_manufacturer] = manufacturer_data
            
            return filtered_manufacturers
        else:
            # 未配置厂家，返回空
            return {}
    
    def continue_after_cloud_sync_config(self):
        """云同步配置完成后继续流程"""
        print("云同步配置完成，继续下载数据和配置厂家...")
        
        # 尝试从云端下载数据
        print("🔄 云同步配置完成，正在下载云端数据...")
        try:
            cloud_data = self.cloud_sync.sync_from_cloud()
            if cloud_data:
                # 成功下载云端数据，使用云端数据
                self.orders = cloud_data.get("orders", {})
                self.manufacturers = cloud_data.get("manufacturers", {})
                self.bound_order_dir = cloud_data.get("bound_order_dir", "")
                # 保存到本地（不进行云同步）
                self.save_data_local_only()
                print("✅ 已从云端成功下载并加载最新数据")
                print(f"📊 下载数据包含: {len(self.orders)} 个订单, {len(self.manufacturers)} 个厂家")
            else:
                # 云端没有数据，使用本地数据
                print("☁️ 云端没有数据，使用本地数据")
                self.load_data_local_only()
                
        except Exception as e:
            print(f"❌ 从云端下载数据失败: {e}，使用本地数据")
            self.load_data_local_only()
        
        # 检查厂家配置
        if not self.load_app_config():
            # 首次运行，需要配置厂家（此时已有厂家数据）
            self.show_manufacturer_config_after_sync()
        else:
            # 已有厂家配置，继续正常流程
            self.continue_normal_startup()
    
    def continue_normal_startup(self):
        """继续正常启动流程"""
        # 重新执行数据加载流程，但跳过厂家配置检查
        print("继续正常启动流程...")
        
        # 数据已经在delayed_initialization中加载，这里不再重复下载
        # 避免启动时的双重下载问题
        print("数据已在启动时加载，跳过重复下载步骤")
        
        # 跳过重复的数据下载逻辑，直接继续后续流程
        self.data_loaded = True  # 标记数据已加载
        print("数据加载完成，刷新界面...")
        self.update_dashboard()
        # 重新显示仪表板内容
        self.show_dashboard_summary()
        print("界面刷新完成")
        # 数据已在delayed_initialization中加载，这里不再重复下载
        # 避免启动时的双重下载问题
        print("数据已在启动时加载，跳过重复下载步骤")
        # 以下代码已禁用，避免重复下载
        # print("检测到云同步配置，尝试从云端下载最新数据...")
        # try:
        #     # 尝试从云端下载数据
        #     cloud_data = self.cloud_sync.sync_from_cloud()
        #     if cloud_data:
        #         # 成功下载云端数据，使用云端数据
        #         self.orders = cloud_data.get("orders", {})
        #         self.manufacturers = cloud_data.get("manufacturers", {})
        #         self.bound_order_dir = cloud_data.get("bound_order_dir", "")
        #         # 保存到本地（不进行云同步）
        #         self.save_data_local_only()
        #         print("已从云端成功下载并加载最新数据")
        #     else:
        #         # 云端没有数据，使用本地数据
        #         print("云端没有数据，使用本地数据")
        #         self.load_data_local_only()
        # except Exception as e:
        #     print(f"从云端下载数据失败: {e}，使用本地数据")
        #     self.load_data_local_only()
        # else:
        #     # 未配置云同步，使用本地数据
        #     print("未配置云同步，使用本地数据")
        #     self.load_data_local_only()
        #     # 显示配置提醒
        #     self.root.after(2000, self.show_sync_config_reminder)  # 2秒后显示提醒
        pass  # 跳过重复下载逻辑
    
    def show_sync_config_reminder(self):
        """显示云同步配置提醒"""
        try:
            result = messagebox.askyesno(
                "配置云同步",
                "您还没有配置云同步功能。\n\n"
                "配置后可以实现：\n"
                "• 数据云端备份，永不丢失\n"
                "• 多设备同步，随时随地访问\n"
                "• 数据版本管理，可回溯历史\n\n"
                "是否现在配置云同步？",
                icon=messagebox.QUESTION
            )
            
            if result:
                # 用户选择配置，打开配置窗口
                self.setup_cloud_sync()
        except:
            pass  # 忽略任何错误，不影响程序正常运行
        
    def create_widgets(self):
        # 创建选项卡
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建各个功能页面
        self.create_dashboard_tab()
        self.create_order_tab()
        
        
    def create_dashboard_tab(self):
        self.dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.dashboard_frame, text="综合管理面板")
        
        # 创建左右分栏布局
        paned_window = ttk.PanedWindow(self.dashboard_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 左侧菜单区域
        left_frame = ttk.Frame(paned_window)
        paned_window.add(left_frame, weight=1)
        
        # 右侧内容区域
        right_frame = ttk.Frame(paned_window)
        paned_window.add(right_frame, weight=3)
        
        # 左侧菜单
        menu_frame = ttk.LabelFrame(left_frame, text="功能菜单", padding=10)
        menu_frame.pack(fill=tk.BOTH, expand=True)
        
        # 云同步设置区域
        sync_frame = ttk.LabelFrame(left_frame, text="云同步", padding=10)
        sync_frame.pack(fill=tk.X, pady=5)
        
        # 云同步控制按钮
        self.sync_status_label = ttk.Label(sync_frame, text="未配置云同步", foreground="gray")
        self.sync_status_label.pack(fill=tk.X, pady=2)
        
        ttk.Button(sync_frame, text="配置GitHub同步", command=self.setup_cloud_sync).pack(fill=tk.X, pady=2)
        ttk.Button(sync_frame, text="立即同步", command=self.manual_sync).pack(fill=tk.X, pady=2)
        # 移除从云端下载按钮，统一使用智能同步逻辑
        
        # 菜单按钮
        dashboard_btn = ttk.Button(menu_frame, text="仪表板", command=self.show_dashboard_summary)
        dashboard_btn.pack(fill=tk.X, pady=5)
        
        unpaid_btn = ttk.Button(menu_frame, text="未结账订单", command=self.show_unpaid_orders)
        unpaid_btn.pack(fill=tk.X, pady=5)
        
        manufacturer_btn = ttk.Button(menu_frame, text="厂家配置", command=self.show_manufacturer_panel)
        manufacturer_btn.pack(fill=tk.X, pady=5)
        
        # 导入导出功能按钮
        import_export_btn = ttk.Button(menu_frame, text="导入导出", command=self.show_import_export_panel)
        import_export_btn.pack(fill=tk.X, pady=5)
        
        # 右侧内容显示区域
        self.content_frame = ttk.Frame(right_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 默认显示仪表板
        self.show_dashboard_summary()
        
    def show_dashboard_summary(self):
        # 清空内容区域
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # 标题
        title_label = ttk.Label(self.content_frame, text="仪表板", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 检查数据是否已加载
        if not hasattr(self, 'data_loaded') or not self.data_loaded:
            # 显示加载提示
            loading_frame = ttk.Frame(self.content_frame)
            loading_frame.pack(fill=tk.BOTH, expand=True)
            
            loading_label = ttk.Label(loading_frame, text="正在加载数据，请稍候...", 
                                    font=("Arial", 14), foreground="#666")
            loading_label.pack(expand=True)
            
            # 添加进度条
            progress = ttk.Progressbar(loading_frame, mode='indeterminate')
            progress.pack(pady=20)
            progress.start()
            
            return
        
        # 权限提示
        if self.current_manufacturer:
            permission_label = ttk.Label(self.content_frame, 
                                       text=f"当前厂家: {self.current_manufacturer}", 
                                       font=("Arial", 10), foreground="blue")
            permission_label.pack(pady=5)
        elif self.is_admin:
            permission_label = ttk.Label(self.content_frame, 
                                       text="管理员模式 - 可查看所有数据", 
                                       font=("Arial", 10), foreground="green")
            permission_label.pack(pady=5)
        
        # 汇总信息卡片
        summary_frame = ttk.LabelFrame(self.content_frame, text="汇总信息", padding=10)
        summary_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 计算汇总数据
        total_unpaid = 0
        unpaid_count = 0
        total_paid = 0
        paid_count = 0
        
        # 获取过滤后的订单数据
        filtered_orders = self.get_filtered_orders()
        
        for order_data in filtered_orders.values():
            # 检查订单是否有支付状态字段
            if "paid" not in order_data:
                continue
            if not order_data["paid"]:
                total_unpaid += order_data.get("total_price", 0)
                unpaid_count += 1
            else:
                total_paid += order_data.get("total_price", 0)
                paid_count += 1
        
        # 创建信息卡片
        cards_frame = ttk.Frame(summary_frame)
        cards_frame.pack(fill=tk.X)
        
        # 未结账卡片
        unpaid_card = tk.Frame(cards_frame, relief=tk.RAISED, bd=2, bg="#ffebee")
        unpaid_card.pack(side=tk.LEFT, padx=10, pady=5, fill=tk.X, expand=True)
        
        tk.Label(unpaid_card, text="未结账", font=("Arial", 12, "bold"), bg="#ffebee").pack(pady=5)
        tk.Label(unpaid_card, text=f"{unpaid_count} 个订单", font=("Arial", 10), bg="#ffebee").pack()
        tk.Label(unpaid_card, text=f"¥ {total_unpaid:.2f}", font=("Arial", 14, "bold"), bg="#ffebee").pack(pady=5)
        
        # 已结账卡片
        paid_card = tk.Frame(cards_frame, relief=tk.RAISED, bd=2, bg="#e8f5e9")
        paid_card.pack(side=tk.LEFT, padx=10, pady=5, fill=tk.X, expand=True)
        
        tk.Label(paid_card, text="已结账", font=("Arial", 12, "bold"), bg="#e8f5e9").pack(pady=5)
        tk.Label(paid_card, text=f"{paid_count} 个订单", font=("Arial", 10), bg="#e8f5e9").pack()
        tk.Label(paid_card, text=f"¥ {total_paid:.2f}", font=("Arial", 14, "bold"), bg="#e8f5e9").pack(pady=5)
        
        # 总计卡片
        total_card = tk.Frame(cards_frame, relief=tk.RAISED, bd=2, bg="#fff3e0")
        total_card.pack(side=tk.LEFT, padx=10, pady=5, fill=tk.X, expand=True)
        
        tk.Label(total_card, text="总计", font=("Arial", 12, "bold"), bg="#fff3e0").pack(pady=5)
        tk.Label(total_card, text=f"{unpaid_count + paid_count} 个订单", font=("Arial", 10), bg="#fff3e0").pack()
        tk.Label(total_card, text=f"¥ {total_unpaid + total_paid:.2f}", font=("Arial", 14, "bold"), bg="#fff3e0").pack(pady=5)
        
        # 盈利日历显示
        calendar_frame = ttk.LabelFrame(self.content_frame, text="📅 盈利日历", padding=15)
        calendar_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 设置日历框架样式
        style = ttk.Style()
        style.configure("Calendar.TLabelframe", background="#F8F9FA")
        style.configure("Calendar.TLabelframe.Label", font=("Arial", 14, "bold"), foreground="#1976D2")
        calendar_frame.configure(style="Calendar.TLabelframe")
        
        control_bar = ttk.Frame(calendar_frame)
        control_bar.pack(fill=tk.X, padx=8, pady=(0, 8))
        years = set()
        from datetime import datetime
        for order in filtered_orders.values():
            try:
                ds = order.get('date', '')
                if not ds:
                    continue
                d = datetime.strptime(ds, '%Y-%m-%d %H:%M:%S') if ' ' in ds else datetime.strptime(ds, '%Y-%m-%d')
                years.add(d.year)
            except Exception:
                continue
        if not years:
            years = {datetime.now().year}
        year_list = sorted(list(years))
        if not hasattr(self, 'calendar_year_var'):
            self.calendar_year_var = tk.StringVar(value=str(year_list[-1]))
        if not hasattr(self, 'calendar_month_var'):
            self.calendar_month_var = tk.StringVar(value=str(datetime.now().month))
        ttk.Label(control_bar, text="年份:").pack(side=tk.LEFT, padx=(0,5))
        year_combo = ttk.Combobox(control_bar, textvariable=self.calendar_year_var, values=[str(y) for y in year_list], state="readonly", width=8)
        year_combo.pack(side=tk.LEFT)
        ttk.Label(control_bar, text="月份:").pack(side=tk.LEFT, padx=(15,5))
        month_combo = ttk.Combobox(control_bar, textvariable=self.calendar_month_var, values=[str(i) for i in range(1,13)], state="readonly", width=6)
        month_combo.pack(side=tk.LEFT)
        def on_calendar_select(e=None):
            self.update_profit_calendar()
        year_combo.bind("<<ComboboxSelected>>", on_calendar_select)
        month_combo.bind("<<ComboboxSelected>>", on_calendar_select)
        
        # 创建日历网格容器
        calendar_container = tk.Frame(calendar_frame, bg="#F5F5F5", relief=tk.SUNKEN, bd=2)
        calendar_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # 存储日历按钮的字典
        self.calendar_buttons = {}
        
        # 创建日历网格 (7列 x 6行)
        for week in range(6):
            for day in range(7):
                btn = tk.Button(calendar_container, text="", width=8, height=3, 
                               font=('Arial', 10, 'bold'), relief=tk.RAISED, bd=2,
                               highlightthickness=1, highlightcolor="#4A90E2",
                               activebackground="#E8F4FD", cursor="hand2")
                btn.grid(row=week, column=day, padx=3, pady=3, sticky="nsew")
                self.calendar_buttons[(week, day)] = btn
        
        # 配置网格权重
        for i in range(7):
            calendar_container.columnconfigure(i, weight=1)
        for i in range(6):
            calendar_container.rowconfigure(i, weight=1)
        
        # 初始化日历显示
        self.update_profit_calendar()
        
    def show_unpaid_orders(self):
        # 清空内容区域
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # 标题
        title_label = ttk.Label(self.content_frame, text="未结账订单", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 排序控制框架
        sort_frame = ttk.LabelFrame(self.content_frame, text="排序和筛选设置")
        sort_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(sort_frame, text="时间排序:").pack(side=tk.LEFT, padx=5, pady=5)
        unpaid_sort_var = tk.StringVar(value="最新在前")
        unpaid_sort_combo = ttk.Combobox(sort_frame, textvariable=unpaid_sort_var,
                                        values=["最新在前", "最旧在前"],
                                        state="readonly", width=15)
        unpaid_sort_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 厂家筛选
        ttk.Label(sort_frame, text="厂家筛选:").pack(side=tk.LEFT, padx=(20, 5), pady=5)
        unpaid_manufacturer_var = tk.StringVar(value="全部厂家")
        
        # 获取所有厂家列表
        manufacturers = set(["全部厂家"])
        for order_data in self.orders.values():
            # 检查订单是否有支付状态字段
            if "paid" not in order_data:
                continue
            if not order_data["paid"]:
                manufacturers.add(order_data["manufacturer"])
        manufacturer_list = sorted(list(manufacturers))
        
        unpaid_manufacturer_combo = ttk.Combobox(sort_frame, textvariable=unpaid_manufacturer_var,
                                                values=manufacturer_list,
                                                state="readonly", width=15)
        unpaid_manufacturer_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 未结账订单列表
        unpaid_frame = ttk.LabelFrame(self.content_frame, text="未结账订单列表")
        unpaid_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 创建表格
        self.unpaid_tree = ttk.Treeview(unpaid_frame, columns=("订单号", "厂家", "面积", "总价", "日期"), show="headings")
        self.unpaid_tree.heading("订单号", text="订单号")
        self.unpaid_tree.heading("厂家", text="厂家")
        self.unpaid_tree.heading("面积", text="面积(㎡)")
        self.unpaid_tree.heading("总价", text="总价(元)")
        self.unpaid_tree.heading("日期", text="日期")
        
        self.unpaid_tree.column("订单号", width=150)
        self.unpaid_tree.column("厂家", width=150)
        self.unpaid_tree.column("面积", width=100)
        self.unpaid_tree.column("总价", width=100)
        self.unpaid_tree.column("日期", width=150)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(unpaid_frame, orient=tk.VERTICAL, command=self.unpaid_tree.yview)
        self.unpaid_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.unpaid_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        def update_unpaid_list():
            """更新未结账订单列表"""
            # 清空现有项目
            for item in self.unpaid_tree.get_children():
                self.unpaid_tree.delete(item)
            
            # 收集未结账订单
            unpaid_orders = []
            selected_manufacturer = unpaid_manufacturer_var.get()
            
            for order_data in self.orders.values():
                # 检查订单是否有支付状态字段
                if "paid" not in order_data:
                    continue
                if not order_data["paid"]:
                    # 厂家筛选
                    if selected_manufacturer == "全部厂家" or order_data.get("manufacturer") == selected_manufacturer:
                        unpaid_orders.append(order_data)
            
            # 按时间排序
            sort_order = unpaid_sort_var.get()
            try:
                if sort_order == "最新在前":
                    unpaid_orders.sort(key=lambda x: datetime.strptime(x.get("date", "2000-01-01 00:00:00"), "%Y-%m-%d %H:%M:%S"), reverse=True)
                else:
                    unpaid_orders.sort(key=lambda x: datetime.strptime(x.get("date", "2000-01-01 00:00:00"), "%Y-%m-%d %H:%M:%S"), reverse=False)
            except (ValueError, KeyError):
                # 如果日期格式有问题，按订单名排序
                unpaid_orders.sort(key=lambda x: x.get("name", ""))
            
            # 添加到列表
            for order_data in unpaid_orders:
                self.unpaid_tree.insert("", tk.END, values=(
                    order_data.get("name", ""),
                    order_data.get("manufacturer", ""),
                    f"{order_data.get('total_area', 0):.2f}",
                    f"{order_data.get('total_price', 0):.2f}",
                    order_data.get("date", "")
                ))
        
        # 绑定排序和筛选变化事件
        unpaid_sort_combo.bind("<<ComboboxSelected>>", lambda e: update_unpaid_list())
        unpaid_manufacturer_combo.bind("<<ComboboxSelected>>", lambda e: update_unpaid_list())
        
        # 初始填充未结账订单数据
        update_unpaid_list()
        
        # 双击事件绑定 - 查看订单详情
        self.unpaid_tree.bind("<Double-1>", self.view_order_from_dashboard)
        
        # 结账按钮
        pay_button_frame = ttk.Frame(self.content_frame)
        pay_button_frame.pack(pady=10)
        
        self.pay_button = ttk.Button(pay_button_frame, text="结账", command=self.mark_as_paid)
        self.pay_button.pack(side=tk.LEFT, padx=5)
        
        refresh_button = ttk.Button(pay_button_frame, text="刷新", command=self.show_unpaid_orders)
        refresh_button.pack(side=tk.LEFT, padx=5)
        
    def view_order_from_dashboard(self, event):
        """从综合管理面板查看订单详情"""
        selected = self.unpaid_tree.selection()
        if not selected:
            return
            
        item = selected[0]
        order_name = self.unpaid_tree.item(item, "values")[0]
        
        # 在订单数据中查找
        for name, order_data in self.orders.items():
            if order_data["name"] == order_name:
                self.show_order_detail_popup(order_data)
                break
                
    def show_order_detail_popup(self, order):
        """显示订单详情弹窗"""
        # 隐藏主窗口
        self.root.withdraw()
        
        # 创建详细信息窗口
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"订单详情 - {order['name']}")
        detail_window.geometry("800x600")
        
        # 设置窗口关闭事件，恢复显示主窗口
        def on_popup_close():
            self.root.deiconify()  # 显示主窗口
            detail_window.destroy()
        
        detail_window.protocol("WM_DELETE_WINDOW", on_popup_close)
        
        # 订单基本信息
        info_frame = ttk.LabelFrame(detail_window, text="订单信息")
        info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(info_frame, text=f"订单号: {order['name']}").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(info_frame, text=f"创建日期: {order['date']}").grid(row=0, column=1, sticky="w", padx=5, pady=2)
        ttk.Label(info_frame, text=f"状态: {'已结账' if order['paid'] else '未结账'}").grid(row=0, column=2, sticky="w", padx=5, pady=2)
        
        # 厂家和价格信息
        price_frame = ttk.LabelFrame(detail_window, text="价格信息")
        price_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(price_frame, text=f"厂家: {order['manufacturer']}").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(price_frame, text=f"单价: {order['unit_price']:.2f} 元/㎡").grid(row=0, column=1, sticky="w", padx=5, pady=2)
        ttk.Label(price_frame, text=f"总面积: {order['total_area']:.2f} ㎡").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(price_frame, text=f"总价: {order['total_price']:.2f} 元").grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        # 房间和柜体信息
        rooms_frame = ttk.LabelFrame(detail_window, text="房间和柜体详情")
        rooms_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 创建左右分栏
        rooms_paned = ttk.PanedWindow(rooms_frame, orient=tk.HORIZONTAL)
        rooms_paned.pack(fill=tk.BOTH, expand=True)
        
        # 左侧房间列表
        rooms_left_frame = ttk.Frame(rooms_paned)
        rooms_paned.add(rooms_left_frame, weight=1)
        
        # 右侧柜体详情
        rooms_right_frame = ttk.Frame(rooms_paned)
        rooms_paned.add(rooms_right_frame, weight=2)
        
        # 房间列表
        rooms_listbox = tk.Listbox(rooms_left_frame, width=20)
        rooms_listbox.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0), pady=5)
        
        # 添加滚动条
        rooms_scrollbar = ttk.Scrollbar(rooms_left_frame, orient=tk.VERTICAL, command=rooms_listbox.yview)
        rooms_listbox.configure(yscrollcommand=rooms_scrollbar.set)
        rooms_scrollbar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5), pady=5)
        
        # 柜体详细信息
        cabinets_tree = ttk.Treeview(rooms_right_frame, columns=("名称", "宽度", "高度", "面积"), show="headings")
        cabinets_tree.heading("名称", text="柜体名称")
        cabinets_tree.heading("宽度", text="宽度(mm)")
        cabinets_tree.heading("高度", text="高度(mm)")
        cabinets_tree.heading("面积", text="面积(㎡)")
        
        cabinets_tree.column("名称", width=150)
        cabinets_tree.column("宽度", width=100)
        cabinets_tree.column("高度", width=100)
        cabinets_tree.column("面积", width=100)
        
        # 添加滚动条
        cabinets_scrollbar = ttk.Scrollbar(rooms_right_frame, orient=tk.VERTICAL, command=cabinets_tree.yview)
        cabinets_tree.configure(yscrollcommand=cabinets_scrollbar.set)
        cabinets_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        cabinets_tree.pack(fill=tk.BOTH, expand=True)
        
        def update_cabinets_list(room_name):
            # 清空现有项目
            for item in cabinets_tree.get_children():
                cabinets_tree.delete(item)
                
            # 添加柜体
            if room_name in order["rooms"]:
                for cabinet_data in order["rooms"][room_name]["cabinets"].values():
                    cabinets_tree.insert("", tk.END, values=(
                        cabinet_data["name"],
                        cabinet_data["width"],
                        cabinet_data["height"],
                        f"{cabinet_data['area']:.4f}"
                    ))
        
        def on_room_select(event):
            selection = rooms_listbox.curselection()
            if selection:
                room_name = rooms_listbox.get(selection[0])
                update_cabinets_list(room_name)
        
        rooms_listbox.bind("<<ListboxSelect>>", on_room_select)
        
        # 填充房间列表
        for room_name in order["rooms"]:
            rooms_listbox.insert(tk.END, room_name)
            
        # 底部按钮
        button_frame = ttk.Frame(detail_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def mark_paid():
            if not order["paid"]:
                order["paid"] = True
                self.update_dashboard()
                self.save_data()
                messagebox.showinfo("成功", f"订单 {order['name']} 已结账")
                on_popup_close()
            else:
                messagebox.showinfo("信息", f"订单 {order['name']} 已经结账")
        
        ttk.Button(button_frame, text="结账", command=mark_paid).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="关闭", command=on_popup_close).pack(side=tk.LEFT, padx=5)
        
    def show_manufacturer_panel(self):
        # 清空内容区域
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # 标题
        title_label = ttk.Label(self.content_frame, text="厂家配置设置", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 厂家配置管理区域
        manufacturer_mgmt_frame = ttk.LabelFrame(self.content_frame, text="厂家配置管理")
        manufacturer_mgmt_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(manufacturer_mgmt_frame, text="厂家名称:").grid(row=0, column=0, padx=5, pady=5)
        self.manufacturer_name_var = tk.StringVar()
        ttk.Entry(manufacturer_mgmt_frame, textvariable=self.manufacturer_name_var, width=25).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(manufacturer_mgmt_frame, text="单价(元/㎡):").grid(row=0, column=2, padx=5, pady=5)
        self.unit_price_var = tk.StringVar()
        ttk.Entry(manufacturer_mgmt_frame, textvariable=self.unit_price_var, width=12).grid(row=0, column=3, padx=5, pady=5)
        
        # 权限管理
        ttk.Label(manufacturer_mgmt_frame, text="权限:").grid(row=0, column=4, padx=5, pady=5)
        self.permission_var = tk.StringVar(value="读写")
        permission_combo = ttk.Combobox(manufacturer_mgmt_frame, textvariable=self.permission_var, 
                                       values=["读", "写", "读写"], width=8, state="readonly")
        permission_combo.grid(row=0, column=5, padx=5, pady=5)
        
        # 创建按钮框架
        button_frame = ttk.Frame(manufacturer_mgmt_frame)
        button_frame.grid(row=0, column=6, padx=10, pady=5)
        
        ttk.Button(button_frame, text="➕ 添加厂家", command=self.add_manufacturer).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="✏️ 更新选中", command=self.update_selected_manufacturer).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="🗑️ 删除选中", command=self.delete_manufacturer).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="🔄 清空输入", command=self.clear_manufacturer_input).pack(side=tk.LEFT, padx=2)
        
        # 添加操作说明
        help_frame = ttk.Frame(self.content_frame)
        help_frame.pack(fill=tk.X, padx=10, pady=5)
        
        help_text = "💡 操作说明：点击表格中的厂家可自动填充到输入框，然后可以修改单价；管理员拥有完整权限，厂家用户只能查看数据。"
        help_label = ttk.Label(help_frame, text=help_text, foreground="blue", wraplength=600)
        help_label.pack(pady=5)
        
        # 当前身份显示和修改区域
        identity_frame = ttk.LabelFrame(self.content_frame, text="当前身份管理")
        identity_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 当前身份显示
        identity_display_frame = ttk.Frame(identity_frame)
        identity_display_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.current_identity_label = ttk.Label(identity_display_frame, text="", font=("Arial", 10, "bold"))
        self.current_identity_label.pack(side=tk.LEFT, padx=5)
        
        self.update_identity_display()  # 更新显示
        
        # 修改身份按钮
        ttk.Button(identity_display_frame, text="🔄 修改当前身份", 
                  command=self.change_current_identity).pack(side=tk.RIGHT, padx=5)
        
        # 厂家列表
        manufacturers_frame = ttk.LabelFrame(self.content_frame, text="厂家列表")
        manufacturers_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.manufacturers_tree = ttk.Treeview(manufacturers_frame, columns=("厂家", "单价", "权限"), show="headings")
        self.manufacturers_tree.heading("厂家", text="厂家")
        self.manufacturers_tree.heading("单价", text="单价(元/㎡)")
        self.manufacturers_tree.heading("权限", text="权限")
        
        self.manufacturers_tree.column("厂家", width=180)
        self.manufacturers_tree.column("单价", width=120)
        self.manufacturers_tree.column("权限", width=80)
        
        # 添加滚动条
        manufacturers_scrollbar = ttk.Scrollbar(manufacturers_frame, orient=tk.VERTICAL, command=self.manufacturers_tree.yview)
        self.manufacturers_tree.configure(yscroll=manufacturers_scrollbar.set)
        manufacturers_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.manufacturers_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 绑定选择事件和双击事件
        self.manufacturers_tree.bind("<<TreeviewSelect>>", self.on_manufacturer_select)
        self.manufacturers_tree.bind("<Double-1>", self.on_manufacturer_double_click)
        
        # 填充厂家数据
        for item in self.manufacturers_tree.get_children():
            self.manufacturers_tree.delete(item)
            
        for name, data in self.manufacturers.items():
            permission = data.get("permission", "读写")  # 获取权限，默认为读写
            self.manufacturers_tree.insert("", tk.END, values=(name, data["unit_price"], permission))
        
        # 添加列表操作说明
        list_help_frame = ttk.Frame(self.content_frame)
        list_help_frame.pack(fill=tk.X, padx=10, pady=5)
        
        list_help_text = "📋 列表操作：点击厂家名称可自动填充到上方输入框，然后可以修改单价等信息；支持双击快速选择。"
        list_help_label = ttk.Label(list_help_frame, text=list_help_text, foreground="green", wraplength=600)
        list_help_label.pack(pady=5)
        
    def create_order_tab(self):
        self.order_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.order_frame, text="订单管理")
        
        # 筛选和排序框架
        filter_frame = ttk.LabelFrame(self.order_frame, text="筛选和排序条件")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 创建主要内容区域
        main_content_frame = ttk.Frame(filter_frame)
        main_content_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 左侧筛选条件
        left_filter_frame = ttk.Frame(main_content_frame)
        left_filter_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 第一行：状态和厂家筛选
        row1_frame = ttk.Frame(left_filter_frame)
        row1_frame.pack(fill=tk.X, pady=2)
        
        # 状态筛选
        ttk.Label(row1_frame, text="状态:").pack(side=tk.LEFT, padx=5, pady=5)
        self.status_filter_var = tk.StringVar(value="全部")
        status_filter_combo = ttk.Combobox(row1_frame, textvariable=self.status_filter_var,
                                          values=["全部", "未结账", "已结账"],
                                          state="readonly", width=15)
        status_filter_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 厂家筛选
        ttk.Label(row1_frame, text="厂家:").pack(side=tk.LEFT, padx=(20, 5), pady=5)
        self.manufacturer_filter_var = tk.StringVar(value="全部")
        self.manufacturer_filter_combo = ttk.Combobox(row1_frame, textvariable=self.manufacturer_filter_var, width=20, state="readonly")
        self.manufacturer_filter_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 第二行：月度筛选和时间排序
        row2_frame = ttk.Frame(left_filter_frame)
        row2_frame.pack(fill=tk.X, pady=2)
        
        # 月度筛选
        ttk.Label(row2_frame, text="月度:").pack(side=tk.LEFT, padx=5, pady=5)
        self.month_filter_var = tk.StringVar(value="全部")
        self.month_filter_combo = ttk.Combobox(row2_frame, textvariable=self.month_filter_var, width=15, state="readonly")
        self.month_filter_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 时间排序控制
        ttk.Label(row2_frame, text="时间排序:").pack(side=tk.LEFT, padx=(20, 5), pady=5)
        self.time_sort_var = tk.StringVar(value="最新在前")
        time_sort_combo = ttk.Combobox(row2_frame, textvariable=self.time_sort_var, 
                                      values=["最新在前", "最旧在前"], 
                                      state="readonly", width=15)
        time_sort_combo.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 右侧汇总数据区域
        summary_frame = ttk.LabelFrame(main_content_frame, text="汇总数据", padding=10)
        summary_frame.pack(side=tk.RIGHT, padx=(20, 0))
        
        # 汇总数据显示
        self.summary_labels = {}
        summary_data = [
            ("总订单数:", "total_orders"),
            ("未结账订单:", "unpaid_orders"),
            ("总面积:", "total_area"),
            ("总金额:", "total_amount")
        ]
        
        for i, (label_text, key) in enumerate(summary_data):
            row = i // 2
            col = i % 2
            
            label_frame = ttk.Frame(summary_frame)
            label_frame.grid(row=row, column=col, padx=10, pady=2, sticky="w")
            
            ttk.Label(label_frame, text=label_text, font=("Arial", 9)).pack(side=tk.LEFT)
            value_label = ttk.Label(label_frame, text="0", font=("Arial", 9, "bold"), foreground="blue")
            value_label.pack(side=tk.LEFT, padx=(5, 0))
            self.summary_labels[key] = value_label
        
        # 第三行：搜索功能
        search_frame = ttk.Frame(filter_frame)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 搜索输入
        ttk.Label(search_frame, text="🔍 搜索:").pack(side=tk.LEFT, padx=5, pady=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)  # 实时搜索
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 搜索范围选择
        self.search_scope_var = tk.StringVar(value="全部")
        search_scope_combo = ttk.Combobox(search_frame, textvariable=self.search_scope_var, 
                                         values=["全部", "订单号", "客户名称", "厂家名称"], 
                                         state="readonly", width=12)
        search_scope_combo.pack(side=tk.LEFT, padx=5, pady=5)
        search_scope_combo.bind("<<ComboboxSelected>>", lambda e: self.on_search_change())
        
        # 筛选和重置按钮
        button_frame = ttk.Frame(search_frame)
        button_frame.pack(side=tk.LEFT, padx=(20, 5), pady=5)
        
        ttk.Button(button_frame, text="筛选", command=self.filter_orders).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="重置", command=self.reset_filter).pack(side=tk.LEFT, padx=2)
        
        # 绑定筛选条件变化事件
        status_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_orders())
        self.manufacturer_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_orders())
        self.month_filter_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_orders())
        time_sort_combo.bind("<<ComboboxSelected>>", lambda e: self.filter_orders())
        
        # 选择文件夹作为订单
        folder_frame = ttk.LabelFrame(self.order_frame, text="批量创建订单（一次性多选文件夹）")
        folder_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 添加当前厂家信息显示
        current_manufacturer_frame = ttk.Frame(folder_frame)
        current_manufacturer_frame.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        self.current_manufacturer_label = ttk.Label(current_manufacturer_frame, text="", font=("Arial", 9))
        self.current_manufacturer_label.pack(side=tk.LEFT, padx=5)
        self.update_current_manufacturer_display()  # 更新显示
        
        ttk.Button(folder_frame, text="一次性选择多个文件夹", command=self.select_folder).pack(side=tk.LEFT, padx=5, pady=5)
        self.folder_path_var = tk.StringVar()
        # 创建显示变量
        self.folder_display_var = tk.StringVar()
        self.folder_display_var.set("尚未选择文件夹")
        self.folder_path_entry = ttk.Entry(folder_frame, textvariable=self.folder_display_var, width=50, state=tk.DISABLED)
        self.folder_path_entry.pack(side=tk.LEFT, padx=5, pady=5)
        
        ttk.Button(folder_frame, text="批量创建订单", command=self.create_order).pack(side=tk.LEFT, padx=5, pady=5)
        
        # 搜索结果统计
        self.search_result_frame = ttk.Frame(self.order_frame)
        self.search_result_frame.pack(fill=tk.X, padx=10, pady=(5, 0))
        
        self.search_result_label = ttk.Label(self.search_result_frame, text="", foreground="blue")
        self.search_result_label.pack(side=tk.LEFT, padx=5)
        
        # 订单列表
        orders_frame = ttk.LabelFrame(self.order_frame, text="订单列表")
        orders_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.orders_tree = ttk.Treeview(orders_frame, columns=("订单号", "厂家", "面积", "总价", "状态", "日期"), show="headings")
        self.orders_tree.heading("订单号", text="订单号")
        self.orders_tree.heading("厂家", text="厂家")
        self.orders_tree.heading("面积", text="面积(㎡)")
        self.orders_tree.heading("总价", text="总价(元)")
        self.orders_tree.heading("状态", text="状态")
        self.orders_tree.heading("日期", text="日期")
        
        self.orders_tree.column("订单号", width=150)
        self.orders_tree.column("厂家", width=150)
        self.orders_tree.column("面积", width=100)
        self.orders_tree.column("总价", width=100)
        self.orders_tree.column("状态", width=100)
        self.orders_tree.column("日期", width=150)
        
        # 添加滚动条
        orders_scrollbar = ttk.Scrollbar(orders_frame, orient=tk.VERTICAL, command=self.orders_tree.yview)
        self.orders_tree.configure(yscroll=orders_scrollbar.set)
        orders_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.orders_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 添加管理按钮
        button_frame = ttk.Frame(self.order_frame)
        button_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.pay_selected_button = ttk.Button(button_frame, text="选中订单结账", command=self.pay_selected_orders)
        self.pay_selected_button.pack(side=tk.LEFT, padx=5)
        
        self.delete_selected_button = ttk.Button(button_frame, text="删除选中订单", command=self.delete_selected_orders)
        self.delete_selected_button.pack(side=tk.LEFT, padx=5)
        
        # 双击订单查看详细信息
        self.orders_tree.bind("<Double-1>", self.view_order_details)
        
        # 更新订单列表和筛选器（数据加载完成后再刷新，避免重复构建）
        if self.data_loaded:
            self.update_orders_list()
            self.update_manufacturer_filter()
            self.update_month_filter()
        
    def select_folder(self):
        # 创建多文件夹选择对话框
        folder_dialog = tk.Toplevel(self.root)
        folder_dialog.title("批量选择文件夹")
        folder_dialog.geometry("700x900")
        folder_dialog.resizable(True, True)
        folder_dialog.transient(self.root)
        folder_dialog.grab_set()
        
        # 创建主框架
        main_frame = ttk.Frame(folder_dialog, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="批量选择文件夹创建订单", font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # 说明
        info_label = ttk.Label(main_frame, text="请选择订单目录进行绑定，系统将自动识别该目录下的所有子文件夹作为订单")
        info_label.pack(pady=(0, 15))
        
        # 选择父目录按钮
        select_frame = ttk.Frame(main_frame)
        select_frame.pack(fill=tk.X, pady=(0, 10))
        
        selected_parent_dir = tk.StringVar()
        detected_folders = []
        
        def select_parent_directory():
            parent_dir = filedialog.askdirectory(title="绑定订单目录")
            if parent_dir:
                selected_parent_dir.set(parent_dir)
                self.bound_order_dir = parent_dir
                scan_subdirectories(parent_dir)
        
        def scan_subdirectories(parent_dir):
            """扫描父目录下的所有子文件夹并排序"""
            try:
                detected_folders.clear()
                folders_listbox.delete(0, tk.END)
                
                # 获取所有子文件夹及其创建时间
                folder_info = []
                for item in os.listdir(parent_dir):
                    item_path = os.path.join(parent_dir, item)
                    if os.path.isdir(item_path):
                        try:
                            # 获取文件夹创建时间
                            create_time = os.path.getctime(item_path)
                            folder_info.append((item_path, create_time))
                        except OSError:
                            # 如果无法获取创建时间，使用当前时间
                            folder_info.append((item_path, 0))
                
                # 按选择的排序方式排序
                sort_order = sort_var.get()
                if sort_order == "最新先创建":
                    folder_info.sort(key=lambda x: x[1], reverse=True)  # 按创建时间降序
                elif sort_order == "最旧先创建":
                    folder_info.sort(key=lambda x: x[1], reverse=False)  # 按创建时间升序
                else:  # 按名称排序
                    folder_info.sort(key=lambda x: os.path.basename(x[0]))
                
                # 添加到列表和数组
                for folder_path, create_time in folder_info:
                    detected_folders.append(folder_path)
                    folder_name = os.path.basename(folder_path)
                    
                    # 根据排序方式显示不同信息
                    if sort_order in ["最新先创建", "最旧先创建"] and create_time > 0:
                        # 显示创建时间
                        from datetime import datetime
                        create_time_str = datetime.fromtimestamp(create_time).strftime("%Y-%m-%d %H:%M")
                        display_text = f"{folder_name} [{create_time_str}] ({folder_path})"
                    else:
                        # 只显示名称和路径
                        display_text = f"{folder_name} ({folder_path})"
                    
                    folders_listbox.insert(tk.END, display_text)
                    # 默认选中所有文件夹
                    folders_listbox.selection_set(folders_listbox.size() - 1)
                
                if detected_folders:
                    status_label.config(text=f"检测到 {len(detected_folders)} 个文件夹（已按{sort_order}排序）", foreground="green")
                else:
                    status_label.config(text="未检测到子文件夹", foreground="orange")
                    
            except Exception as e:
                status_label.config(text=f"扫描失败: {str(e)}", foreground="red")
        
        ttk.Button(select_frame, text="绑定订单目录", command=select_parent_directory).pack(side=tk.LEFT, padx=5)
        
        # 当前选择的目录显示
        current_dir_label = ttk.Label(select_frame, textvariable=selected_parent_dir, foreground="blue")
        current_dir_label.pack(side=tk.LEFT, padx=10)
        
        # 排序控制框架
        sort_frame = ttk.LabelFrame(main_frame, text="排序方式")
        sort_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(sort_frame, text="排序方式:").pack(side=tk.LEFT, padx=5)
        sort_var = tk.StringVar(value="最新先创建")
        sort_combo = ttk.Combobox(sort_frame, textvariable=sort_var,
                                 values=["最新先创建", "最旧先创建", "按名称排序"],
                                 state="readonly", width=15)
        sort_combo.pack(side=tk.LEFT, padx=5)
        
        # 绑定排序方式变化事件
        def on_sort_change(event=None):
            if selected_parent_dir.get():
                scan_subdirectories(selected_parent_dir.get())
        
        sort_combo.bind("<<ComboboxSelected>>", on_sort_change)
        
        # 状态显示
        status_label = ttk.Label(main_frame, text="请先绑定订单目录", foreground="gray")
        status_label.pack(pady=5)
        
        # 检测到的文件夹列表
        list_frame = ttk.LabelFrame(main_frame, text="检测到的文件夹（可多选）", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # 创建列表框和滚动条
        list_container = ttk.Frame(list_frame)
        list_container.pack(fill=tk.BOTH, expand=True)
        
        folders_listbox = tk.Listbox(list_container, selectmode=tk.EXTENDED, height=15)
        list_scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=folders_listbox.yview)
        folders_listbox.configure(yscrollcommand=list_scrollbar.set)
        
        folders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 批量选择按钮
        selection_frame = ttk.Frame(list_frame)
        selection_frame.pack(fill=tk.X, pady=(10, 0))
        
        def select_all():
            folders_listbox.selection_set(0, tk.END)
        
        def deselect_all():
            folders_listbox.selection_clear(0, tk.END)
        
        def invert_selection():
            current_selection = set(folders_listbox.curselection())
            folders_listbox.selection_clear(0, tk.END)
            for i in range(folders_listbox.size()):
                if i not in current_selection:
                    folders_listbox.selection_set(i)
        
        ttk.Button(selection_frame, text="全选", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(selection_frame, text="全不选", command=deselect_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(selection_frame, text="反选", command=invert_selection).pack(side=tk.LEFT, padx=5)
        
        # 底部按钮
        final_button_frame = ttk.Frame(main_frame)
        final_button_frame.pack(pady=10)
        
        result = {"folders": []}
        
        def confirm_selection():
            selected_indices = folders_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("警告", "请至少选择一个文件夹")
                return
            
            # 获取选中的文件夹路径
            selected_folders = [detected_folders[i] for i in selected_indices]
            result["folders"] = selected_folders
            folder_dialog.destroy()
        
        def cancel_selection():
            result["folders"] = []
            folder_dialog.destroy()
        
        # 显示选择数量
        def update_selection_count(event=None):
            count = len(folders_listbox.curselection())
            confirm_btn.config(text=f"确定选择 ({count} 个文件夹)")
        
        folders_listbox.bind('<<ListboxSelect>>', update_selection_count)
        
        confirm_btn = ttk.Button(final_button_frame, text="确定选择 (0 个文件夹)", command=confirm_selection)
        confirm_btn.pack(side=tk.LEFT, padx=10)
        ttk.Button(final_button_frame, text="取消", command=cancel_selection).pack(side=tk.LEFT, padx=10)
        
        # 快捷操作提示
        tip_frame = ttk.Frame(main_frame)
        tip_frame.pack(fill=tk.X, pady=(10, 0))
        
        tip_text = "提示：\n• 按住Ctrl点击可多选\n• 按住Shift点击可连续选择\n• 支持全选/反选操作"
        tip_label = ttk.Label(tip_frame, text=tip_text, font=("Arial", 9), foreground="gray", justify=tk.LEFT)
        tip_label.pack(anchor="w")
        
        # 等待用户选择
        if self.bound_order_dir and os.path.isdir(self.bound_order_dir):
            selected_parent_dir.set(self.bound_order_dir)
            scan_subdirectories(self.bound_order_dir)
            status_label.config(text=f"已绑定目录: {self.bound_order_dir}", foreground="green")
            current_dir_label.config(textvariable=selected_parent_dir)

        folder_dialog.wait_window()
        
        # 处理选择结果
        if result["folders"]:
            # 将多个文件夹路径存储到folder_path_var中，用分号分隔
            self.folder_path_var.set(";".join(result["folders"]))
            # 在文件夹路径输入框中显示选择的数量
            if len(result["folders"]) == 1:
                display_text = os.path.basename(result["folders"][0])
            else:
                count = len(result['folders'])
                display_text = f"已选择 {count} 个文件夹"
            
            # 创建一个临时显示变量
            if not hasattr(self, 'folder_display_var'):
                self.folder_display_var = tk.StringVar()
            self.folder_display_var.set(display_text)
            
            # 更新显示
            if hasattr(self.folder_path_entry, 'config'):
                self.folder_path_entry.config(textvariable=self.folder_display_var)
            
    def create_order(self):
        # 检查写权限
        if not self.check_write_permission("创建订单"):
            return
            
        folder_paths_str = self.folder_path_var.get()
        if not folder_paths_str:
            messagebox.showerror("错误", "请选择有效的文件夹路径")
            return
        
        # 分割多个文件夹路径
        folder_paths = [path.strip() for path in folder_paths_str.split(";") if path.strip()]
        
        if not folder_paths:
            messagebox.showerror("错误", "请选择有效的文件夹路径")
            return
        
        # 验证所有文件夹路径是否存在
        invalid_paths = []
        for folder_path in folder_paths:
            if not os.path.exists(folder_path):
                invalid_paths.append(folder_path)
        
        if invalid_paths:
            messagebox.showerror("错误", f"以下文件夹路径不存在：\n{chr(10).join(invalid_paths)}")
            return
        
        # 检查是否有重复的订单名
        existing_orders = []
        new_orders_info = []
        
        for folder_path in folder_paths:
            order_name = os.path.basename(folder_path)
            if order_name in self.orders:
                existing_orders.append(order_name)
            else:
                new_orders_info.append((order_name, folder_path))
        
        # 如果有重复订单，询问用户是否继续
        if existing_orders:
            if not messagebox.askyesno("重复订单", 
                                     f"以下订单已存在，将跳过创建：\n{chr(10).join(existing_orders)}\n\n是否继续创建其他订单？"):
                return
        
        if not new_orders_info:
            messagebox.showinfo("信息", "所有订单都已存在，无需创建")
            return
        
        # 批量创建订单
        created_orders = []
        failed_orders = []
        
        for order_name, folder_path in new_orders_info:
            try:
                # 获取文件夹的创建时间
                folder_create_time = os.path.getctime(folder_path)
                folder_date = datetime.fromtimestamp(folder_create_time).strftime("%Y-%m-%d %H:%M:%S")
                
                # 创建订单结构，厂家默认为当前厂家
                current_manufacturer = self.current_manufacturer if self.current_manufacturer else ""
                unit_price = 0
                if current_manufacturer and current_manufacturer in self.manufacturers:
                    unit_price = self.manufacturers[current_manufacturer]["unit_price"]
                
                self.orders[order_name] = {
                    "name": order_name,
                    "path": folder_path,
                    "rooms": {},
                    "total_area": 0,
                    "total_price": 0,
                    "manufacturer": current_manufacturer,
                    "unit_price": unit_price,
                    "paid": False,
                    "date": folder_date
                }
                
                # 解析文件夹结构
                self.parse_folder_structure(order_name, folder_path)
                created_orders.append(order_name)
                
            except Exception as e:
                failed_orders.append(f"{order_name}: {str(e)}")
                # 如果创建失败，从订单字典中移除
                if order_name in self.orders:
                    del self.orders[order_name]
        
        # 清空文件夹选择
        self.folder_path_var.set("")
        
        # 更新界面
        self.update_orders_list()
        self.update_dashboard()
        self.save_data()
        
        # 显示创建结果
        success_count = len(created_orders)
        result_message = f"成功创建 {success_count} 个订单"
        
        if created_orders:
            result_message += f"\n\n创建成功的订单：\n{chr(10).join(created_orders)}"
        
        if failed_orders:
            result_message += f"\n\n创建失败的订单：\n{chr(10).join(failed_orders)}"
        
        if existing_orders:
            result_message += f"\n\n跳过的重复订单：\n{chr(10).join(existing_orders)}"
        
        messagebox.showinfo("批量创建订单结果", result_message)
        
    def parse_folder_structure(self, order_name, folder_path):
        """Parse folder structure, create rooms and cabinets"""
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            if os.path.isdir(item_path):
                # 子文件夹作为房间
                self.orders[order_name]["rooms"][item] = {
                    "name": item,
                    "cabinets": {}
                }
                
    def update_dashboard(self):
        """更新仪表板"""
        # 更新厂家列表（应用权限过滤）
        if hasattr(self, 'manufacturers_tree') and self.manufacturers_tree.winfo_exists():
            for item in self.manufacturers_tree.get_children():
                self.manufacturers_tree.delete(item)
                
            # 获取权限过滤后的厂家数据
            filtered_manufacturers = self.get_filtered_manufacturers()
            for name, data in filtered_manufacturers.items():
                self.manufacturers_tree.insert("", tk.END, values=(name, data["unit_price"]))
        
        # 更新当前厂家显示
        self.update_current_manufacturer_display()
                
        # 更新订单列表和筛选器
        if hasattr(self, 'orders_tree'):
            self.update_orders_list()
            self.update_manufacturer_filter()
            self.update_month_filter()
            
    def update_manufacturer_filter(self):
        """Update manufacturer filter options (apply permission filtering)"""
        if hasattr(self, 'manufacturer_filter_var') and hasattr(self, 'manufacturer_filter_combo'):
            # 获取权限过滤后的厂家数据
            filtered_manufacturers = self.get_filtered_manufacturers()
            manufacturers = ["全部"] + list(filtered_manufacturers.keys())
            self.manufacturer_filter_combo['values'] = manufacturers
            if self.manufacturer_filter_var.get() not in manufacturers:
                self.manufacturer_filter_var.set("全部")
        
    def on_search_change(self, *args):
        """搜索内容变化时的实时搜索"""
        self.filter_orders()
    
    def search_orders(self, orders_data, search_text, search_scope):
        """搜索订单数据"""
        if not search_text.strip():
            return orders_data
        
        search_text = search_text.strip().lower()
        filtered_orders = {}
        
        for order_id, order_data in orders_data.items():
            match = False
            
            if search_scope == "全部":
                # 搜索所有字段
                if (search_text in order_id.lower() or 
                    search_text in order_data.get("customer_name", "").lower() or
                    search_text in order_data.get("manufacturer", "").lower()):
                    match = True
            elif search_scope == "订单号":
                if search_text in order_id.lower():
                    match = True
            elif search_scope == "客户名称":
                if search_text in order_data.get("customer_name", "").lower():
                    match = True
            elif search_scope == "厂家名称":
                if search_text in order_data.get("manufacturer", "").lower():
                    match = True
            
            if match:
                filtered_orders[order_id] = order_data
        
        return filtered_orders
    
    def filter_orders(self):
        """Filter orders based on criteria (includes search functionality)"""
        self.update_orders_list()
        
    def reset_filter(self):
        """重置筛选条件和排序设置"""
        if hasattr(self, 'status_filter_var'):
            self.status_filter_var.set("全部")
        if hasattr(self, 'manufacturer_filter_var'):
            self.manufacturer_filter_var.set("全部")
        if hasattr(self, 'month_filter_var'):
            self.month_filter_var.set("全部")
        if hasattr(self, 'time_sort_var'):
            self.time_sort_var.set("最新在前")
        if hasattr(self, 'search_var'):
            self.search_var.set("")  # 清空搜索
        if hasattr(self, 'search_scope_var'):
            self.search_scope_var.set("全部")  # 重置搜索范围
        self.update_orders_list()
        
    def update_month_filter(self):
        """更新月度筛选器选项"""
        if hasattr(self, 'month_filter_var') and hasattr(self, 'month_filter_combo'):
            # 收集所有订单的月份
            months = set(["全部"])
            for order_data in self.orders.values():
                try:
                    # 检查订单是否有日期字段
                    if "date" not in order_data:
                        continue
                    order_date = datetime.strptime(order_data["date"], "%Y-%m-%d %H:%M:%S")
                    month_str = order_date.strftime("%Y-%m")
                    months.add(month_str)
                except (ValueError, KeyError):
                    continue
            
            # 按时间排序月份选项
            month_list = ["全部"] + sorted([m for m in months if m != "全部"], reverse=True)
            self.month_filter_combo['values'] = month_list
            if self.month_filter_var.get() not in month_list:
                self.month_filter_var.set("全部")
    
    def update_summary_data(self, filtered_orders):
        """更新汇总数据显示"""
        if not hasattr(self, 'summary_labels'):
            return
            
        total_orders = len(filtered_orders)
        unpaid_orders = sum(1 for _, order_data in filtered_orders if "paid" in order_data and not order_data["paid"])
        total_area = sum(order_data.get("total_area", 0) for _, order_data in filtered_orders)
        total_amount = sum(order_data.get("total_price", 0) for _, order_data in filtered_orders)
        
        # 更新显示
        self.summary_labels["total_orders"].config(text=str(total_orders))
        self.summary_labels["unpaid_orders"].config(text=str(unpaid_orders))
        self.summary_labels["total_area"].config(text=f"{total_area:.2f}㎡")
        self.summary_labels["total_amount"].config(text=f"¥{total_amount:.2f}")
    
    def update_orders_list(self):
        """Update order list (supports status, manufacturer, monthly filtering and time sorting)"""
        # 清空现有项目
        if not hasattr(self, 'orders_tree'):
            return
            
        for item in self.orders_tree.get_children():
            self.orders_tree.delete(item)
            
        # 获取筛选条件
        selected_status = self.status_filter_var.get() if hasattr(self, 'status_filter_var') else "全部"
        selected_manufacturer = self.manufacturer_filter_var.get() if hasattr(self, 'manufacturer_filter_var') else "全部"
        selected_month = self.month_filter_var.get() if hasattr(self, 'month_filter_var') else "全部"
        
        # 获取排序方式
        sort_order = self.time_sort_var.get() if hasattr(self, 'time_sort_var') else "最新在前"
        
        # 获取权限过滤后的订单数据
        permission_filtered_orders = self.get_filtered_orders()
        
        # 应用搜索过滤
        search_text = self.search_var.get() if hasattr(self, 'search_var') else ""
        search_scope = self.search_scope_var.get() if hasattr(self, 'search_scope_var') else "全部"
        if search_text.strip():
            permission_filtered_orders = self.search_orders(permission_filtered_orders, search_text, search_scope)
        
        # 收集符合筛选条件的订单
        filtered_orders = []
        for order_name, order_data in permission_filtered_orders.items():
            # 应用状态筛选
            if "paid" not in order_data:
                continue
            if selected_status == "未结账" and order_data["paid"]:
                continue
            elif selected_status == "已结账" and not order_data["paid"]:
                continue
            
            # 应用厂家筛选
            if selected_manufacturer != "全部" and order_data.get("manufacturer") != selected_manufacturer:
                continue
            
            # 应用月度筛选
            if selected_month != "全部":
                try:
                    order_date = datetime.strptime(order_data["date"], "%Y-%m-%d %H:%M:%S")
                    order_month = order_date.strftime("%Y-%m")
                    if order_month != selected_month:
                        continue
                except ValueError:
                    continue
            
            filtered_orders.append((order_name, order_data))
        
        # 按时间排序
        try:
            if sort_order == "最新在前":
                # 按日期降序排序（最新的在前）
                filtered_orders.sort(key=lambda x: datetime.strptime(x[1]["date"], "%Y-%m-%d %H:%M:%S"), reverse=True)
            else:
                # 按日期升序排序（最旧的在前）
                filtered_orders.sort(key=lambda x: datetime.strptime(x[1]["date"], "%Y-%m-%d %H:%M:%S"), reverse=False)
        except ValueError:
            # 如果日期格式有问题，按订单名排序
            filtered_orders.sort(key=lambda x: x[0])
        
        # 添加排序后的订单到列表
        for order_name, order_data in filtered_orders:
            status = "已结账" if order_data.get("paid", False) else "未结账"
            self.orders_tree.insert("", tk.END, values=(
                order_name,
                order_data.get("manufacturer", ""),
                f"{order_data.get('total_area', 0):.2f}",
                f"{order_data.get('total_price', 0):.2f}",
                status,
                order_data.get("date", "")
            ))
        
        # 更新搜索结果统计
        if hasattr(self, 'search_result_label'):
            search_text = self.search_var.get() if hasattr(self, 'search_var') and self.search_var.get().strip() else ""
            if search_text:
                self.search_result_label.config(
                    text=f"🔍 搜索结果: {len(filtered_orders)} 个订单 (搜索: '{search_text}')"
                )
            else:
                self.search_result_label.config(
                    text=f"📋 显示订单: {len(filtered_orders)} 个"
                )
        
        # 更新汇总数据
        self.update_summary_data(filtered_orders)
            
    def mark_as_paid(self):
        """标记为已结账"""
        # 检查写权限
        if not self.check_write_permission("结账操作"):
            return
            
        selected = self.unpaid_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请选择要结账的订单")
            return
            
        for item in selected:
            order_name = self.unpaid_tree.item(item, "values")[0]
            if order_name in self.orders:
                self.orders[order_name]["paid"] = True
                
        self.update_orders_list()
        self.update_dashboard()
        self.save_data()
        messagebox.showinfo("成功", "订单已标记为已结账")
        
    def pay_selected_orders(self):
        """选中订单结账"""
        # 检查写权限
        if not self.check_write_permission("结账操作"):
            return
            
        selected_items = self.orders_tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请选择要结账的订单")
            return
            
        #  确认结账
        if not messagebox.askyesno("确认", f"确定要为选中的 {len(selected_items)} 个订单结账吗？"):
            return
            
        # 结账选中的订单
        for item in selected_items:
            order_name = self.orders_tree.item(item, "values")[0]
            if order_name in self.orders:
                self.orders[order_name]["paid"] = True
                
        self.update_orders_list()
        self.update_dashboard()
        self.save_data()
        messagebox.showinfo("成功", f"已为 {len(selected_items)} 个订单结账")
    
    def delete_selected_orders(self):
        """删除选中的订单"""
        # 检查写权限
        if not self.check_write_permission("删除订单"):
            return
            
        selected_items = self.orders_tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请选择要删除的订单")
            return
        
        # 获取选中的订单名称
        order_names = []
        for item in selected_items:
            order_name = self.orders_tree.item(item, "values")[0]
            order_names.append(order_name)
        
        # 确认删除
        if not messagebox.askyesno("确认", f"确定要删除选中的 {len(selected_items)} 个订单吗？\n\n订单列表：\n{chr(10).join(order_names)}"):
            return
        
        # 删除订单
        deleted_count = 0
        for order_name in order_names:
            if order_name in self.orders:
                del self.orders[order_name]
                deleted_count += 1
        
        # 更新界面
        self.update_orders_list()
        self.update_dashboard()
        self.save_data()
        messagebox.showinfo("成功", f"已成功删除 {deleted_count} 个订单")
        
    def view_order_details(self, event):
        """查看订单详细信息"""
        selected = self.orders_tree.selection()
        if not selected:
            return
            
        item = selected[0]
        order_name = self.orders_tree.item(item, "values")[0]
        if order_name not in self.orders:
            return
            
        order = self.orders[order_name]
        
        # 隐藏主窗口
        self.root.withdraw()
        
        # 创建详细信息窗口
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"订单详情 - {order_name}")
        detail_window.geometry("900x700")
        
        # 设置窗口关闭事件，恢复显示主窗口
        def on_detail_window_close():
            self.root.deiconify()  # 显示主窗口
            detail_window.destroy()
        
        detail_window.protocol("WM_DELETE_WINDOW", on_detail_window_close)
        
        # 创建选项卡
        detail_notebook = ttk.Notebook(detail_window)
        detail_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 订单信息选项卡
        info_frame = ttk.Frame(detail_notebook)
        detail_notebook.add(info_frame, text="订单信息")
        
        # 订单基本信息
        basic_info_frame = ttk.LabelFrame(info_frame, text="基本信息")
        basic_info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(basic_info_frame, text=f"订单号: {order['name']}").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(basic_info_frame, text=f"创建日期: {order['date']}").grid(row=0, column=1, sticky="w", padx=5, pady=2)
        
        status_var = tk.BooleanVar(value=order["paid"])
        status_check = ttk.Checkbutton(basic_info_frame, text="已结账", variable=status_var)
        status_check.grid(row=0, column=2, sticky="w", padx=5, pady=2)
        
        # 厂家和价格信息
        price_frame = ttk.LabelFrame(info_frame, text="价格信息")
        price_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(price_frame, text="选择厂家:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        manufacturer_var = tk.StringVar(value=order["manufacturer"])
        manufacturer_combo = ttk.Combobox(price_frame, textvariable=manufacturer_var, values=list(self.manufacturers.keys()), width=20)
        manufacturer_combo.grid(row=0, column=1, sticky="w", padx=5, pady=2)
        
        def update_manufacturer(*args):
            # 检查写权限 - 修改厂家需要写权限
            if not self.check_write_permission("修改订单厂家"):
                # 权限不足，恢复原来的厂家
                original_manufacturer = order.get("manufacturer", "")
                if original_manufacturer in self.manufacturers:
                    manufacturer_var.set(original_manufacturer)
                return
                
            manufacturer = manufacturer_var.get()
            if manufacturer in self.manufacturers:
                unit_price = self.manufacturers[manufacturer]["unit_price"]
                order["manufacturer"] = manufacturer
                order["unit_price"] = unit_price
                # 重新计算总价
                order["total_price"] = order["total_area"] * unit_price
                self.update_orders_list()
                self.update_dashboard()
                self.save_data()
                update_totals()
        
        manufacturer_var.trace("w", update_manufacturer)
        
        # 面积和总价显示
        totals_frame = ttk.Frame(price_frame)
        totals_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=5, pady=5)
        
        area_label = ttk.Label(totals_frame, text=f"总面积: {order['total_area']:.2f} ㎡")
        area_label.pack(side=tk.LEFT, padx=5)
        
        price_label = ttk.Label(totals_frame, text=f"总价: {order['total_price']:.2f} 元")
        price_label.pack(side=tk.LEFT, padx=5)
        
        def update_totals():
            area_label.config(text=f"总面积: {order['total_area']:.2f} ㎡")
            price_label.config(text=f"总价: {order['total_price']:.2f} 元")
        
        # 保存按钮
        button_frame = ttk.Frame(info_frame)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def save_order_info():
            # 检查写权限
            if not self.check_write_permission("保存订单信息"):
                return
                
            # 保存结账状态
            order["paid"] = status_var.get()
            self.update_orders_list()
            self.update_dashboard()
            self.save_data_local_only()
            messagebox.showinfo("成功", "订单信息已保存")
        
        ttk.Button(button_frame, text="保存信息", command=save_order_info).pack(side=tk.LEFT, padx=5)
        
        def manual_sync():
            """手动同步数据到云端"""
            try:
                from datetime import datetime
                data = {
                    "orders": self.orders,
                    "manufacturers": self.manufacturers,
                    "bound_order_dir": self.bound_order_dir,
                    "timestamp": datetime.now().isoformat(),
                    "version": "1.0"
                }
                self.cloud_sync.save_and_sync(data, sync_to_cloud=True)
                messagebox.showinfo("同步成功", "数据已同步到云端")
            except Exception as e:
                messagebox.showerror("同步失败", f"数据同步失败: {str(e)}")
        
        ttk.Button(button_frame, text="同步到云端", command=manual_sync).pack(side=tk.LEFT, padx=5)
        
        # 房间和柜体选项卡
        rooms_frame = ttk.Frame(detail_notebook)
        detail_notebook.add(rooms_frame, text="房间和柜体")
        
        # 创建左右分栏
        rooms_paned = ttk.PanedWindow(rooms_frame, orient=tk.HORIZONTAL)
        rooms_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 左侧房间列表
        rooms_left_frame = ttk.Frame(rooms_paned)
        rooms_paned.add(rooms_left_frame, weight=1)
        
        # 右侧柜体详情
        rooms_right_frame = ttk.Frame(rooms_paned)
        rooms_paned.add(rooms_right_frame, weight=2)
        
        # 房间列表
        rooms_list_frame = ttk.LabelFrame(rooms_left_frame, text="房间列表")
        rooms_list_frame.pack(fill=tk.BOTH, expand=True)
        
        rooms_listbox = tk.Listbox(rooms_list_frame, width=30)
        rooms_listbox.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0), pady=5)
        
        # 添加滚动条
        rooms_scrollbar = ttk.Scrollbar(rooms_list_frame, orient=tk.VERTICAL, command=rooms_listbox.yview)
        rooms_listbox.configure(yscrollcommand=rooms_scrollbar.set)
        rooms_scrollbar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5), pady=5)
        
        # 房间管理按钮
        room_buttons_frame = ttk.Frame(rooms_left_frame)
        room_buttons_frame.pack(fill=tk.X, pady=5)
        
        def add_room():
            # 检查写权限
            if not self.check_write_permission():
                messagebox.showerror("权限不足", "您没有写权限，无法添加房间")
                return
                
            # 创建添加房间对话框
            add_room_window = tk.Toplevel(detail_window)
            add_room_window.title("添加房间")
            add_room_window.geometry("400x200")
            add_room_window.resizable(False, False)
            
            ttk.Label(add_room_window, text="房间名称:").pack(pady=10)
            room_name_var = tk.StringVar()
            room_name_entry = ttk.Entry(add_room_window, textvariable=room_name_var, width=30)
            room_name_entry.pack(pady=5)
            room_name_entry.focus()
            
            def save_room():
                room_name = room_name_var.get().strip()
                if not room_name:
                    messagebox.showerror("错误", "请输入房间名称")
                    return
                
                if room_name in order["rooms"]:
                    messagebox.showerror("错误", f"房间 '{room_name}' 已存在")
                    return
                
                # 添加房间
                order["rooms"][room_name] = {"name": room_name, "cabinets": {}}
                
                # 更新界面
                self.save_data_local_only()
                refresh_rooms_list()
                add_room_window.destroy()
                messagebox.showinfo("成功", f"房间 '{room_name}' 已添加")
            
            button_frame = ttk.Frame(add_room_window)
            button_frame.pack(pady=20)
            
            ttk.Button(button_frame, text="确定", command=save_room).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="取消", command=add_room_window.destroy).pack(side=tk.LEFT, padx=5)
            
            # 绑定回车键
            add_room_window.bind('<Return>', lambda e: save_room())
        
        def delete_room():
            # 检查写权限
            if not self.check_write_permission():
                messagebox.showerror("权限不足", "您没有写权限，无法删除房间")
                return
                
            selected_room_idx = rooms_listbox.curselection()
            if not selected_room_idx:
                messagebox.showwarning("警告", "请选择要删除的房间")
                return
            
            room_name = rooms_listbox.get(selected_room_idx[0])
            
            # 检查房间中是否有柜体
            cabinet_count = len(order["rooms"][room_name]["cabinets"]) if room_name in order["rooms"] else 0
            
            confirm_msg = f"确定要删除房间 '{room_name}' 吗？"
            if cabinet_count > 0:
                confirm_msg += f"\n\n注意：该房间包含 {cabinet_count} 个柜体，删除后将一同删除。"
            
            if not messagebox.askyesno("确认删除", confirm_msg):
                return
            
            # 删除房间
            if room_name in order["rooms"]:
                del order["rooms"][room_name]
            
            # 重新计算总面积和总价
            total_area = 0
            for room_data in order["rooms"].values():
                for cabinet_data in room_data["cabinets"].values():
                    total_area += cabinet_data["area"]
            
            order["total_area"] = total_area
            order["total_price"] = total_area * order["unit_price"]
            
            # 更新界面
            self.update_orders_list()
            self.update_dashboard()
            self.save_data_local_only()
            update_totals()
            refresh_rooms_list()
            
            # 清空柜体列表
            for item in cabinets_tree.get_children():
                cabinets_tree.delete(item)
            
            messagebox.showinfo("成功", f"房间 '{room_name}' 已删除")
        
        ttk.Button(room_buttons_frame, text="添加房间", command=add_room).pack(side=tk.LEFT, padx=2)
        ttk.Button(room_buttons_frame, text="删除房间", command=delete_room).pack(side=tk.LEFT, padx=2)
        
        # 柜体详细信息
        cabinets_detail_frame = ttk.LabelFrame(rooms_right_frame, text="柜体详情")
        cabinets_detail_frame.pack(fill=tk.BOTH, expand=True)
        
        cabinets_tree = ttk.Treeview(cabinets_detail_frame, columns=("名称", "宽度", "高度", "面积"), show="headings")
        cabinets_tree.heading("名称", text="柜体名称")
        cabinets_tree.heading("宽度", text="宽度(mm)")
        cabinets_tree.heading("高度", text="高度(mm)")
        cabinets_tree.heading("面积", text="面积(㎡)")
        
        cabinets_tree.column("名称", width=150)
        cabinets_tree.column("宽度", width=100)
        cabinets_tree.column("高度", width=100)
        cabinets_tree.column("面积", width=100)
        
        # 添加滚动条
        cabinets_scrollbar = ttk.Scrollbar(cabinets_detail_frame, orient=tk.VERTICAL, command=cabinets_tree.yview)
        cabinets_tree.configure(yscrollcommand=cabinets_scrollbar.set)
        cabinets_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        cabinets_tree.pack(fill=tk.BOTH, expand=True)
        
        # 添加柜体输入框和按钮
        add_cabinet_frame = ttk.LabelFrame(cabinets_detail_frame, text="添加柜体", padding=5)
        add_cabinet_frame.pack(fill=tk.X, pady=5)
        
        # 第一行：宽度和高度输入
        input_row1 = ttk.Frame(add_cabinet_frame)
        input_row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(input_row1, text="宽度(mm):").pack(side=tk.LEFT, padx=2)
        width_var = tk.StringVar()
        width_entry = ttk.Entry(input_row1, textvariable=width_var, width=12)
        width_entry.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(input_row1, text="高度(mm):").pack(side=tk.LEFT, padx=2)
        height_var = tk.StringVar()
        height_entry = ttk.Entry(input_row1, textvariable=height_var, width=12)
        height_entry.pack(side=tk.LEFT, padx=2)
        
        # 第二行：操作按钮
        button_row = ttk.Frame(add_cabinet_frame)
        button_row.pack(fill=tk.X, pady=5)
        
        def add_cabinet():
            # 检查写权限
            if not self.check_write_permission():
                messagebox.showerror("权限不足", "您没有写权限，无法添加柜体")
                return
                
            selected_room_idx = rooms_listbox.curselection()
            if not selected_room_idx:
                messagebox.showwarning("警告", "请选择一个房间")
                return
                
            room_name = rooms_listbox.get(selected_room_idx[0])
            
            # 获取当前房间中最大的柜体编号
            cabinet_number = 1
            if room_name in order["rooms"]:
                existing_cabinets = order["rooms"][room_name]["cabinets"]
                if existing_cabinets:
                    # 提取现有柜体名称中的数字部分
                    numbers = []
                    for cabinet_name in existing_cabinets.keys():
                        if cabinet_name.startswith("柜体"):
                            try:
                                num = int(cabinet_name[2:])  # 提取"柜体"后的数字
                                numbers.append(num)
                            except ValueError:
                                pass
                    if numbers:
                        cabinet_number = max(numbers) + 1
            
            cabinet_name = f"柜体{cabinet_number}"
            width = width_var.get()
            height = height_var.get()
            
            if not width or not height:
                messagebox.showerror("错误", "请填写完整信息")
                return
                
            try:
                width = float(width)
                height = float(height)
            except ValueError:
                messagebox.showerror("错误", "宽度和高度必须是数字")
                return
                
            # 计算面积（平方米）- 从毫米转换为平方米
            area = (width * height) / 1000000
            
            # 添加到订单数据
            if room_name not in order["rooms"]:
                order["rooms"][room_name] = {"name": room_name, "cabinets": {}}
                
            order["rooms"][room_name]["cabinets"][cabinet_name] = {
                "name": cabinet_name,
                "width": width,
                "height": height,
                "area": area
            }
            
            # 重新计算总面积
            total_area = 0
            for room_data in order["rooms"].values():
                for cabinet_data in room_data["cabinets"].values():
                    total_area += cabinet_data["area"]
                    
            order["total_area"] = total_area
            # 重新计算总价
            order["total_price"] = total_area * order["unit_price"]
            
            # 更新界面
            self.update_orders_list()
            self.update_dashboard()
            self.save_data_local_only()
            update_totals()
            update_cabinets_list(room_name)
            
            # 清空输入（但保留自动命名）
            width_var.set("")
            height_var.set("")
            messagebox.showinfo("成功", f"柜体 '{cabinet_name}' 已添加")
        
        def save_order_data():
            """保存订单数据的确定按钮功能"""
            # 保存所有数据
            self.update_orders_list()
            self.update_dashboard()
            self.save_data_local_only()
            update_totals()
            messagebox.showinfo("成功", "订单信息已保存")
            # 使用统一的关闭函数
            on_detail_window_close()
        
        # 定义键盘导航功能
        def on_width_tab(event):
            """Width input Tab key event: jump to height input box"""
            height_entry.focus_set()
            return "break"
        
        def on_height_tab(event):
            """Height input Tab key event: auto-click add cabinet button"""
            add_cabinet()
            return "break"
        
        def on_width_enter(event):
            """Width input Enter key event: jump to height input box"""
            height_entry.focus_set()
            return "break"
        
        def on_height_enter(event):
            """Height input Enter key event: add cabinet"""
            add_cabinet()
            return "break"
        
        def on_room_enter(event):
            """Room list Enter key event: select next room, save if none"""
            current_selection = rooms_listbox.curselection()
            if current_selection:
                current_index = current_selection[0]
                next_index = current_index + 1
                if next_index < rooms_listbox.size():
                    # 选择下一个房间
                    rooms_listbox.selection_clear(0, tk.END)
                    rooms_listbox.selection_set(next_index)
                    rooms_listbox.activate(next_index)
                    room_name = rooms_listbox.get(next_index)
                    update_cabinets_list(room_name)
                    # 设置焦点到宽度输入框
                    width_entry.focus_set()
                else:
                    # 没有下一个房间，自动保存
                    save_order_data()
            return "break"
        
        # 绑定键盘事件
        width_entry.bind('<Tab>', on_width_tab)
        height_entry.bind('<Tab>', on_height_tab)
        width_entry.bind('<Return>', on_width_enter)
        height_entry.bind('<Return>', on_height_enter)
        rooms_listbox.bind('<Return>', on_room_enter)
        
        # 按钮布局
        ttk.Button(button_row, text="添加柜体", command=add_cabinet).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_row, text="确定保存", command=save_order_data).pack(side=tk.LEFT, padx=5)
        
        # 提示信息
        tip_label = ttk.Label(add_cabinet_frame, 
                             text="提示：Tab键跳转输入框，Enter键选择下一个房间或保存订单", 
                             font=("Arial", 8), foreground="gray")
        tip_label.pack(pady=2)
        
        # 修改柜体功能
        def edit_cabinet():
            # 检查写权限
            if not self.check_write_permission():
                messagebox.showerror("权限不足", "您没有写权限，无法修改柜体")
                return
                
            selected_cabinet = cabinets_tree.selection()
            if not selected_cabinet:
                messagebox.showwarning("警告", "请选择要修改的柜体")
                return
                
            cabinet_values = cabinets_tree.item(selected_cabinet[0], "values")
            cabinet_name = cabinet_values[0]
            
            # 创建修改窗口
            edit_window = tk.Toplevel(detail_window)
            edit_window.title(f"修改柜体 - {cabinet_name}")
            edit_window.geometry("450x300")
            
            selected_room_idx = rooms_listbox.curselection()
            if not selected_room_idx:
                return
                
            room_name = rooms_listbox.get(selected_room_idx[0])
            
            # 获取柜体数据
            cabinet_data = order["rooms"][room_name]["cabinets"][cabinet_name]
            
            ttk.Label(edit_window, text="柜体名称:").pack(pady=5)
            name_entry = ttk.Entry(edit_window, width=30)
            name_entry.pack(pady=5)
            name_entry.insert(0, cabinet_name)
            name_entry.config(state="readonly")
            
            ttk.Label(edit_window, text="宽度(mm):").pack(pady=5)
            width_entry = ttk.Entry(edit_window, width=30)
            width_entry.pack(pady=5)
            width_entry.insert(0, cabinet_data["width"])
            
            ttk.Label(edit_window, text="高度(mm):").pack(pady=5)
            height_entry = ttk.Entry(edit_window, width=30)
            height_entry.pack(pady=5)
            height_entry.insert(0, cabinet_data["height"])
            
            def save_changes():
                try:
                    new_width = float(width_entry.get())
                    new_height = float(height_entry.get())
                except ValueError:
                    messagebox.showerror("错误", "宽度和高度必须是数字")
                    return
                    
                # 计算新面积 - 从毫米转换为平方米
                new_area = (new_width * new_height) / 1000000
                
                # 更新数据
                cabinet_data["width"] = new_width
                cabinet_data["height"] = new_height
                cabinet_data["area"] = new_area
                
                # 重新计算总面积
                total_area = 0
                for room_data in order["rooms"].values():
                    for cabinet_info in room_data["cabinets"].values():
                        total_area += cabinet_info["area"]
                        
                order["total_area"] = total_area
                # 重新计算总价
                order["total_price"] = total_area * order["unit_price"]
                
                # 更新界面
                self.update_orders_list()
                self.update_dashboard()
                self.save_data_local_only()
                update_totals()
                update_cabinets_list(room_name)
                
                edit_window.destroy()
                messagebox.showinfo("成功", "柜体信息已更新")
                
            def delete_cabinet():
                # 检查写权限
                if not self.check_write_permission():
                    messagebox.showerror("权限不足", "您没有写权限，无法删除柜体")
                    return
                    
                if messagebox.askyesno("确认", f"确定要删除柜体 {cabinet_name} 吗？"):
                    del order["rooms"][room_name]["cabinets"][cabinet_name]
                    
                    # 重新计算总面积
                    total_area = 0
                    for room_data in order["rooms"].values():
                        for cabinet_info in room_data["cabinets"].values():
                            total_area += cabinet_info["area"]
                            
                    order["total_area"] = total_area
                    # 重新计算总价
                    order["total_price"] = total_area * order["unit_price"]
                    
                    # 更新界面
                    self.update_orders_list()
                    self.update_dashboard()
                    self.save_data_local_only()
                    update_totals()
                    update_cabinets_list(room_name)
                    
                    edit_window.destroy()
                    messagebox.showinfo("成功", "柜体已删除")
            
            button_frame = ttk.Frame(edit_window)
            button_frame.pack(pady=10)
            
            ttk.Button(button_frame, text="保存", command=save_changes).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="删除", command=delete_cabinet).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="取消", command=edit_window.destroy).pack(side=tk.LEFT, padx=5)
        
        # 添加修改和删除按钮
        edit_button_frame = ttk.Frame(cabinets_detail_frame)
        edit_button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(edit_button_frame, text="修改柜体", command=edit_cabinet).pack(side=tk.LEFT, padx=5)
        
        def delete_cabinet_only():
            """单独的删除柜体功能"""
            # 检查写权限
            if not self.check_write_permission():
                messagebox.showerror("权限不足", "您没有写权限，无法删除柜体")
                return
                
            selected_cabinet = cabinets_tree.selection()
            if not selected_cabinet:
                messagebox.showwarning("警告", "请选择要删除的柜体")
                return
            
            cabinet_values = cabinets_tree.item(selected_cabinet[0], "values")
            cabinet_name = cabinet_values[0]
            
            selected_room_idx = rooms_listbox.curselection()
            if not selected_room_idx:
                messagebox.showwarning("警告", "请选择房间")
                return
            
            room_name = rooms_listbox.get(selected_room_idx[0])
            
            if not messagebox.askyesno("确认", f"确定要删除柜体 '{cabinet_name}' 吗？"):
                return
            
            # 删除柜体
            if room_name in order["rooms"] and cabinet_name in order["rooms"][room_name]["cabinets"]:
                del order["rooms"][room_name]["cabinets"][cabinet_name]
            
            # 重新计算总面积
            total_area = 0
            for room_data in order["rooms"].values():
                for cabinet_data in room_data["cabinets"].values():
                    total_area += cabinet_data["area"]
            
            order["total_area"] = total_area
            order["total_price"] = total_area * order["unit_price"]
            
            # 更新界面
            self.update_orders_list()
            self.update_dashboard()
            self.save_data()
            update_totals()
            update_cabinets_list(room_name)
            
            messagebox.showinfo("成功", f"柜体 '{cabinet_name}' 已删除")
        
        ttk.Button(edit_button_frame, text="删除柜体", command=delete_cabinet_only).pack(side=tk.LEFT, padx=5)
        
        def update_cabinets_list(room_name):
            # 清空现有项目
            for item in cabinets_tree.get_children():
                cabinets_tree.delete(item)
                
            # 添加柜体
            if room_name in order["rooms"]:
                for cabinet_data in order["rooms"][room_name]["cabinets"].values():
                    cabinets_tree.insert("", tk.END, values=(
                        cabinet_data["name"],
                        cabinet_data["width"],
                        cabinet_data["height"],
                        f"{cabinet_data['area']:.4f}"
                    ))
        
        def refresh_rooms_list():
            """刷新房间列表"""
            # 清空房间列表
            rooms_listbox.delete(0, tk.END)
            
            # 重新填充房间列表
            for room_name in order["rooms"]:
                rooms_listbox.insert(tk.END, room_name)
        
        def on_room_select(event):
            selection = rooms_listbox.curselection()
            if selection:
                room_name = rooms_listbox.get(selection[0])
                update_cabinets_list(room_name)
        
        rooms_listbox.bind("<<ListboxSelect>>", on_room_select)
        
        # 填充房间列表并设置默认选择
        refresh_rooms_list()
        
        # 默认选择第一个房间并设置焦点
        if rooms_listbox.size() > 0:
            rooms_listbox.selection_set(0)
            rooms_listbox.activate(0)
            first_room_name = rooms_listbox.get(0)
            update_cabinets_list(first_room_name)
            # 设置焦点到宽度输入框
            detail_window.after(100, lambda: width_entry.focus_set())
            
    def on_manufacturer_select(self, event):
        """厂家列表选择事件处理"""
        selection = self.manufacturers_tree.selection()
        if selection:
            item = selection[0]
            values = self.manufacturers_tree.item(item, "values")
            if values:
                manufacturer_name = values[0]
                unit_price = values[1].replace("元/㎡", "")
                permission = values[2] if len(values) > 2 else "读写"  # 获取权限信息
                
                # 自动填充到输入框
                self.manufacturer_name_var.set(manufacturer_name)
                self.unit_price_var.set(unit_price)
                self.permission_var.set(permission)  # 设置权限
    
    def on_manufacturer_double_click(self, event):
        """厂家列表双击事件处理"""
        selection = self.manufacturers_tree.selection()
        if selection:
            item = selection[0]
            values = self.manufacturers_tree.item(item, "values")
            if values:
                manufacturer_name = values[0]
                messagebox.showinfo("厂家信息", f"厂家名称: {manufacturer_name}\n单价: {values[1]}")
    
    def add_manufacturer(self):
        """添加新厂家"""
        name = self.manufacturer_name_var.get().strip()
        unit_price = self.unit_price_var.get().strip()
        
        if not name or not unit_price:
            messagebox.showerror("错误", "请填写完整信息")
            return
            
        try:
            unit_price = float(unit_price)
            if unit_price <= 0:
                messagebox.showerror("错误", "单价必须大于0")
                return
        except ValueError:
            messagebox.showerror("错误", "单价必须是有效的数字")
            return
        
        # 检查厂家是否已存在
        if name in self.manufacturers:
            messagebox.showerror("错误", f"厂家 '{name}' 已存在，请使用更新功能")
            return
            
        # 添加新厂家（包含默认权限）
        self.manufacturers[name] = {
            "name": name,
            "unit_price": unit_price,
            "permission": "读写"  # 默认权限：读写
        }
        
        # 清空输入框
        self.clear_manufacturer_input()
        
        # 刷新界面
        self.update_dashboard()
        self.save_data()
        messagebox.showinfo("成功", f"厂家 '{name}' 添加成功")
    
    def update_selected_manufacturer(self):
        """更新选中的厂家"""
        selection = self.manufacturers_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要更新的厂家")
            return
            
        # 获取选中的厂家名称和当前权限
        item = selection[0]
        values = self.manufacturers_tree.item(item, "values")
        old_name = values[0]
        current_permission = values[2] if len(values) > 2 else "读写"
        new_permission = self.permission_var.get()
        
        # 如果权限发生变更，需要管理员验证
        if current_permission != new_permission:
            if not self.verify_admin_password():
                messagebox.showwarning("权限变更", "权限变更需要管理员密码验证")
                return
        
        # 获取新的厂家信息
        new_name = self.manufacturer_name_var.get().strip()
        unit_price = self.unit_price_var.get().strip()
        
        if not new_name or not unit_price:
            messagebox.showerror("错误", "请填写完整信息")
            return
            
        try:
            unit_price = float(unit_price)
            if unit_price <= 0:
                messagebox.showerror("错误", "单价必须大于0")
                return
        except ValueError:
            messagebox.showerror("错误", "单价必须是有效的数字")
            return
        
        # 如果修改了厂家名称，需要检查新名称是否已存在
        if new_name != old_name and new_name in self.manufacturers:
            messagebox.showerror("错误", f"厂家名称 '{new_name}' 已存在")
            return
        
        # 删除旧的厂家记录（如果名称改变了）
        if new_name != old_name:
            del self.manufacturers[old_name]
            
        # 更新厂家信息（包含权限）
        permission = self.permission_var.get()
        self.manufacturers[new_name] = {
            "name": new_name,
            "unit_price": unit_price,
            "permission": permission
        }
        
        messagebox.showinfo("成功", f"厂家 '{new_name}' 更新成功")
        self.clear_manufacturer_input()
        self.update_dashboard()
        self.save_data()
    
    def update_identity_display(self):
        """更新当前身份显示"""
        if self.is_admin:
            self.current_identity_label.config(text="👨‍💼 当前身份: 管理员（可查看所有数据）", foreground="green")
        elif self.current_manufacturer:
            self.current_identity_label.config(text=f"🏭 当前身份: {self.current_manufacturer}（仅查看本厂家数据）", foreground="blue")
        else:
            self.current_identity_label.config(text="❓ 当前身份: 未配置（需要选择厂家或登录管理员）", foreground="red")
        
        # 同时更新订单界面的厂家显示
        self.update_current_manufacturer_display()
    
    def update_current_manufacturer_display(self):
        """更新订单界面的当前厂家显示"""

        if hasattr(self, 'current_manufacturer_label'):
            if self.is_admin:
                self.current_manufacturer_label.config(
                    text="📋 管理员模式：新订单将默认不指定厂家", 
                    foreground="green"
                )
            elif self.current_manufacturer:
                unit_price = self.manufacturers.get(self.current_manufacturer, {}).get("unit_price", 0)
                self.current_manufacturer_label.config(
                    text=f"🏭 当前厂家：{self.current_manufacturer}（单价：{unit_price}元/㎡）", 
                    foreground="blue"
                )
            else:
                self.current_manufacturer_label.config(
                    text="⚠️ 未设置厂家：新订单将默认不指定厂家", 
                    foreground="orange"
                )
    
    def check_write_permission(self, operation="操作"):
        """检查当前用户是否有写权限
        :param operation: Operation description, used for error messages
        :return: True if has write permission, False if not
        """
        if self.is_admin:
            return True  # 管理员总是有写权限
        
        if not self.current_manufacturer:
            messagebox.showerror("权限错误", f"未配置厂家身份，无法进行{operation}")
            return False
        
        if self.current_manufacturer not in self.manufacturers:
            messagebox.showerror("权限错误", f"当前厂家 '{self.current_manufacturer}' 不存在")
            return False
        
        manufacturer_data = self.manufacturers[self.current_manufacturer]
        permission = manufacturer_data.get("permission", "读写")
        
        if permission not in ["写", "读写"]:
            messagebox.showerror("权限错误", 
                               f"当前厂家 '{self.current_manufacturer}' 权限为 '{permission}'，无法进行{operation}\n\n"
                               f"需要 '写' 或 '读写' 权限")
            return False
        
        return True
    
    def check_read_permission(self, operation="查看"):
        """检查当前用户是否有读权限
        :param operation: Operation description, used for error messages
        :return: True if has read permission, False if not
        """
        if self.is_admin:
            return True  # 管理员总是有读权限
        
        if not self.current_manufacturer:
            messagebox.showerror("权限错误", f"未配置厂家身份，无法{operation}")
            return False
        
        if self.current_manufacturer not in self.manufacturers:
            messagebox.showerror("权限错误", f"当前厂家 '{self.current_manufacturer}' 不存在")
            return False
        
        manufacturer_data = self.manufacturers[self.current_manufacturer]
        permission = manufacturer_data.get("permission", "读写")
        
        # 读权限：只有读权限或读写权限都可以查看数据
        if permission not in ["读", "读写"]:
            messagebox.showerror("权限错误", 
                               f"当前厂家 '{self.current_manufacturer}' 权限为 '{permission}'，无法{operation}\n\n"
                               f"需要 '读' 或 '读写' 权限")
            return False
        
        return True
    
    def check_import_permission(self):
        """Check import permissions - write permission required for import"""
        return self.check_write_permission("导入数据")
    
    def check_export_permission(self):
        """检查导出权限 - 读权限即可导出"""
        return self.check_read_permission("导出数据")
    
    def check_sync_permission(self, sync_type="同步"):
        """检查同步权限
        :param sync_type: 同步类型描述
        :return: True if has permission, False if not
        """
        if self.is_admin:
            return True
        
        if not self.current_manufacturer:
            messagebox.showerror("权限错误", f"未配置厂家身份，无法进行{sync_type}")
            return False
        
        # 检查厂家是否存在
        if self.current_manufacturer not in self.manufacturers:
            messagebox.showerror("权限错误", f"厂家 '{self.current_manufacturer}' 不存在于配置中")
            return False
            
        manufacturer_data = self.manufacturers[self.current_manufacturer]
        permission = manufacturer_data.get("permission", "读写")
        
        # 同步需要读写权限
        if permission != "读写":
            messagebox.showerror("权限错误", 
                               f"当前厂家 '{self.current_manufacturer}' 权限为 '{permission}'，无法进行{sync_type}\n\n"
                               f"需要 '读写' 权限")
            return False
        
        return True
    
    def change_current_identity(self):
        """Modify current identity (requires admin password verification)"""
        # 验证管理员权限
        if not self.verify_admin_password():
            return
        
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择身份")
        select_window.geometry("400x500")
        select_window.transient(self.root)
        select_window.grab_set()
        
        # 窗口居中
        select_window.update_idletasks()
        x = (select_window.winfo_screenwidth() - select_window.winfo_width()) // 2
        y = (select_window.winfo_screenheight() - select_window.winfo_height()) // 2
        select_window.geometry(f"+{x}+{y}")
        
        # 主框架
        main_frame = ttk.Frame(select_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        ttk.Label(main_frame, text="选择新的身份", font=("Arial", 14, "bold")).pack(pady=10)
        
        # 当前身份显示
        current_text = "管理员" if self.is_admin else (self.current_manufacturer or "未设置")
        ttk.Label(main_frame, text=f"当前身份: {current_text}", foreground="blue").pack(pady=5)
        
        # 厂家列表框架
        list_frame = ttk.LabelFrame(main_frame, text="选择厂家身份")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 厂家列表
        manufacturers_listbox = tk.Listbox(list_frame, height=15)
        manufacturers_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=manufacturers_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        manufacturers_listbox.configure(yscrollcommand=scrollbar.set)
        
        # 填充厂家数据
        for name in sorted(self.manufacturers.keys()):
            manufacturers_listbox.insert(tk.END, name)
        
        # 管理员选项
        admin_frame = ttk.Frame(main_frame)
        admin_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(admin_frame, text="👨‍💼 切换为管理员身份", 
                  command=lambda: self.set_admin_identity(select_window)).pack(pady=5)
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        def on_confirm():
            selection = manufacturers_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个厂家身份")
                return
            
            manufacturer_name = manufacturers_listbox.get(selection[0])
            self.set_manufacturer_identity(manufacturer_name, select_window)
        
        ttk.Button(button_frame, text="✅ 确认选择", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="❌ 取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
        
        # 设置默认选中当前厂家
        if self.current_manufacturer and self.current_manufacturer in self.manufacturers:
            index = list(sorted(self.manufacturers.keys())).index(self.current_manufacturer)
            manufacturers_listbox.selection_set(index)
    
    def set_admin_identity(self, window):
        """设置为管理员身份"""
        self.is_admin = True
        self.current_manufacturer = None
        self.save_app_config()
        self.update_identity_display()
        self.update_dashboard()  # 刷新界面以应用新权限
        window.destroy()
        messagebox.showinfo("成功", "已切换为管理员身份")
    
    def set_manufacturer_identity(self, manufacturer_name, window):
        """设置为厂家身份"""
        self.is_admin = False
        self.current_manufacturer = manufacturer_name
        self.save_app_config()
        self.update_identity_display()
        self.update_dashboard()  # 刷新界面以应用新权限
        window.destroy()
        messagebox.showinfo("成功", f"已切换为厂家身份: {manufacturer_name}")
    
    def verify_admin_password(self):
        """验证管理员密码"""
        password = tk.simpledialog.askstring("管理员验证", "请输入管理员密码:", show="*")
        if password is None:  # 用户取消
            return False
        
        if password != self.admin_password:
            messagebox.showerror("错误", "管理员密码错误")
            return False
        
        return True
        
    def delete_manufacturer(self):
        """Delete manufacturer (with data integrity check)"""
        selected = self.manufacturers_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请选择要删除的厂家")
            return
            
        item = selected[0]
        manufacturer_name = self.manufacturers_tree.item(item, "values")[0]
        
        # 检查是否有使用该厂家的订单
        related_orders = []
        for order_id, order_data in self.orders.items():
            if order_data.get("manufacturer") == manufacturer_name:
                related_orders.append(f"{order_id}: {order_data.get('customer_name', '未知客户')}")
        
        if related_orders:
            # 显示相关订单信息
            order_list = "\n".join(related_orders[:5])  # 只显示前5个
            if len(related_orders) > 5:
                order_list += f"\n... 还有 {len(related_orders) - 5} 个订单"
            
            if not messagebox.askyesno("⚠️  重要警告", 
                f"厂家 '{manufacturer_name}' 还有 {len(related_orders)} 个相关订单：\n\n{order_list}\n\n"
                f"删除厂家会导致这些订单无法正确显示厂家信息。\n\n确定要删除吗？"):
                return
        
        # 确认删除
        if messagebox.askyesno("确认删除", f"确定要删除厂家 '{manufacturer_name}' 吗？\n\n此操作不可撤销。"):
            if manufacturer_name in self.manufacturers:
                del self.manufacturers[manufacturer_name]
                self.clear_manufacturer_input()  # 清空输入框
                self.update_dashboard()
                self.save_data()
                messagebox.showinfo("成功", f"厂家 '{manufacturer_name}' 已删除")
                
                # 如果有相关订单，提醒用户更新这些订单
                if related_orders:
                    messagebox.showinfo("提示", 
                        f"已删除 {len(related_orders)} 个相关订单中的厂家信息。\n"
                        f"建议您在订单管理中更新这些订单的厂家信息。")
        
    def show_import_export_panel(self):
        # 清空内容区域
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # 标题
        title_label = ttk.Label(self.content_frame, text="导入导出管理", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 主容器
        main_container = ttk.Frame(self.content_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 导入选项框架
        import_options_frame = ttk.LabelFrame(main_container, text="导入选项", padding=10)
        import_options_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 导入按钮
        import_buttons_frame = ttk.Frame(import_options_frame)
        import_buttons_frame.pack(fill=tk.X)
        
        ttk.Button(import_buttons_frame, text="从Excel导入订单", command=self.import_orders_from_excel, width=25).pack(side=tk.LEFT, padx=5)
        ttk.Button(import_buttons_frame, text="从JSON导入订单", command=self.import_orders_from_json, width=25).pack(side=tk.LEFT, padx=5)
        
        # 导入说明
        import_info_label = ttk.Label(import_options_frame, 
                                    text="支持Excel和JSON格式订单数据导入，可一键配置厂家并检测重复订单", 
                                    font=("Arial", 9), foreground="gray")
        import_info_label.pack(pady=(10, 0))
        
        # 导出选项框架
        export_options_frame = ttk.LabelFrame(main_container, text="导出选项", padding=10)
        export_options_frame.pack(fill=tk.BOTH, expand=True)
        
        
        # 创建导出选项
        options = [
            ("单个订单", self.export_single_order),
            ("多个订单", self.export_multiple_orders),
            ("部分订单", self.export_partial_orders),
            ("所有订单", self.export_all_orders),
            ("未结账订单", self.export_unpaid_orders),
            ("定制厂家所有订单", self.export_manufacturer_all_orders),
            ("定制厂家未结账订单", self.export_manufacturer_unpaid_orders),
            ("月度汇总数据", self.export_monthly_summary),
            ("季度汇总数据", self.export_quarterly_summary),
            ("年度汇总数据", self.export_yearly_summary)
        ]
        
        # 创建按钮网格
        buttons_frame = ttk.Frame(export_options_frame)
        buttons_frame.pack(fill=tk.BOTH, expand=True)
        
        row, col = 0, 0
        for text, command in options:
            btn = ttk.Button(buttons_frame, text=text, command=command, width=25)
            btn.grid(row=row, column=col, padx=5, pady=5, sticky="ew")
            col += 1
            if col > 2:  # 每行3个按钮
                col = 0
                row += 1
                
        # 配置网格列权重
        for i in range(3):
            buttons_frame.columnconfigure(i, weight=1)
            
        # 导出说明
        info_label = ttk.Label(export_options_frame, 
                              text="选择导出选项以导出相应数据到Excel、PDF或JSON格式，支持部分订单导出", 
                              font=("Arial", 10))
        info_label.pack(side=tk.BOTTOM, pady=10)
        
    def export_single_order(self):
        """导出单个订单"""
        # 检查导出权限 - 需要读权限
        if not self.check_export_permission():
            return
            
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择订单")
        select_window.geometry("450x350")
        
        ttk.Label(select_window, text="请选择要导出的订单:", font=("Arial", 12)).pack(pady=10)
        
        # 搜索框架
        search_frame = ttk.Frame(select_window)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 订单列表
        orders_listbox = tk.Listbox(select_window)
        orders_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 存储所有订单名称
        all_orders = list(self.orders.keys())
        
        # 填充订单列表
        def update_orders_list(search_text=""):
            orders_listbox.delete(0, tk.END)
            for order_name in all_orders:
                if search_text.lower() in order_name.lower():
                    orders_listbox.insert(tk.END, order_name)
        
        # 初始填充
        update_orders_list()
        
        # 绑定搜索事件
        def on_search_change(*args):
            update_orders_list(search_var.get())
        
        search_var.trace('w', on_search_change)
            
        def do_export_excel():
            selection = orders_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个订单")
                return
                
            order_name = orders_listbox.get(selection[0])
            if order_name in self.orders:
                self.export_order_to_excel(self.orders[order_name])
                select_window.destroy()
        
        def do_export_pdf():
            selection = orders_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个订单")
                return
                
            order_name = orders_listbox.get(selection[0])
            if order_name in self.orders:
                self.export_order_to_pdf(self.orders[order_name])
                select_window.destroy()
                
        def do_export_json():
            selection = orders_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个订单")
                return
                
            order_name = orders_listbox.get(selection[0])
            if order_name in self.orders:
                self.export_order_to_json(self.orders[order_name])
                select_window.destroy()
                
        button_frame = ttk.Frame(select_window)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="导出为Excel", command=do_export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为PDF", command=do_export_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为JSON", command=do_export_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
        
    def export_multiple_orders(self):
        """导出多个订单"""
        # 检查导出权限 - 需要读权限
        if not self.check_export_permission():
            return
            
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择多个订单")
        select_window.geometry("450x450")
        
        ttk.Label(select_window, text="请选择要导出的订单 (可多选):", font=("Arial", 12)).pack(pady=10)
        
        # 搜索框架
        search_frame = ttk.Frame(select_window)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 订单列表 (支持多选)
        orders_listbox = tk.Listbox(select_window, selectmode=tk.EXTENDED)
        orders_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 存储所有订单名称
        all_orders = list(self.orders.keys())
        
        # 填充订单列表
        def update_orders_list(search_text=""):
            orders_listbox.delete(0, tk.END)
            for order_name in all_orders:
                if search_text.lower() in order_name.lower():
                    orders_listbox.insert(tk.END, order_name)
        
        # 初始填充
        update_orders_list()
        
        # 绑定搜索事件
        def on_search_change(*args):
            update_orders_list(search_var.get())
        
        search_var.trace('w', on_search_change)
            
        def do_export_excel():
            selections = orders_listbox.curselection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
                
            selected_orders = []
            for idx in selections:
                order_name = orders_listbox.get(idx)
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_excel(selected_orders, "多个订单")
                select_window.destroy()
        
        def do_export_pdf():
            selections = orders_listbox.curselection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
                
            selected_orders = []
            for idx in selections:
                order_name = orders_listbox.get(idx)
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_pdf(selected_orders, "多个订单")
                select_window.destroy()
                
        def do_export_json():
            selections = orders_listbox.curselection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
                
            selected_orders = []
            for idx in selections:
                order_name = orders_listbox.get(idx)
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_json(selected_orders, "多个订单")
                select_window.destroy()
                
        button_frame = ttk.Frame(select_window)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="导出为Excel", command=do_export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为PDF", command=do_export_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为JSON", command=do_export_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
        
    def export_all_orders(self):
        """导出所有订单"""
        # 检查导出权限 - 需要读权限
        if not self.check_export_permission():
            return
            
        if not self.orders:
            messagebox.showwarning("警告", "没有订单数据可导出")
            return
        
        # 创建选择导出格式的对话框
        format_window = tk.Toplevel(self.root)
        format_window.title("选择导出格式")
        format_window.geometry("450x200")
        format_window.resizable(False, False)
        
        ttk.Label(format_window, text="请选择导出格式:", font=("Arial", 12)).pack(pady=20)
        
        def export_excel():
            orders_list = list(self.orders.values())
            self.export_orders_to_excel(orders_list, "所有订单")
            format_window.destroy()
            messagebox.showinfo("成功", "所有订单已导出")
        
        def export_pdf():
            orders_list = list(self.orders.values())
            self.export_orders_to_pdf(orders_list, "所有订单")
            format_window.destroy()
            
        def export_json():
            orders_list = list(self.orders.values())
            self.export_orders_to_json(orders_list, "所有订单")
            format_window.destroy()
        
        button_frame = ttk.Frame(format_window)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="Excel格式", command=export_excel).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="PDF格式", command=export_pdf).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="JSON格式", command=export_json).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消", command=format_window.destroy).pack(side=tk.LEFT, padx=10)
        
    def export_unpaid_orders(self):
        """导出未结账订单"""
        unpaid_orders = [order for order in self.orders.values() if not order["paid"]]
        
        if not unpaid_orders:
            messagebox.showinfo("信息", "没有未结账订单")
            return
        
        # 创建选择导出格式的对话框
        format_window = tk.Toplevel(self.root)
        format_window.title("选择导出格式")
        format_window.geometry("450x200")
        format_window.resizable(False, False)
        
        ttk.Label(format_window, text="请选择导出格式:", font=("Arial", 12)).pack(pady=20)
        
        def export_excel():
            self.export_orders_to_excel(unpaid_orders, "未结账订单")
            format_window.destroy()
            messagebox.showinfo("成功", f"已导出 {len(unpaid_orders)} 个未结账订单")
        
        def export_pdf():
            self.export_orders_to_pdf(unpaid_orders, "未结账订单")
            format_window.destroy()
            
        def export_json():
            self.export_orders_to_json(unpaid_orders, "未结账订单")
            format_window.destroy()
        
        button_frame = ttk.Frame(format_window)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="Excel格式", command=export_excel).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="PDF格式", command=export_pdf).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="JSON格式", command=export_json).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消", command=format_window.destroy).pack(side=tk.LEFT, padx=10)
        
    def export_manufacturer_all_orders(self):
        """导出定制厂家所有订单"""
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择厂家")
        select_window.geometry("450x450")
        
        ttk.Label(select_window, text="请选择厂家:", font=("Arial", 12)).pack(pady=10)
        
        # 搜索框架
        search_frame = ttk.Frame(select_window)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 厂家列表
        manufacturer_listbox = tk.Listbox(select_window)
        manufacturer_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 存储所有厂家名称
        all_manufacturers = list(self.manufacturers.keys())
        
        # 填充厂家列表
        def update_manufacturers_list(search_text=""):
            manufacturer_listbox.delete(0, tk.END)
            for manufacturer_name in all_manufacturers:
                if search_text.lower() in manufacturer_name.lower():
                    manufacturer_listbox.insert(tk.END, manufacturer_name)
        
        # 初始填充
        update_manufacturers_list()
        
        # 绑定搜索事件
        def on_search_change(*args):
            update_manufacturers_list(search_var.get())
        
        search_var.trace('w', on_search_change)
            
        def do_export():
            selection = manufacturer_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个厂家")
                return
                
            manufacturer_name = manufacturer_listbox.get(selection[0])
            manufacturer_orders = [order for order in self.orders.values() 
                                 if order["manufacturer"] == manufacturer_name]
            
            if not manufacturer_orders:
                messagebox.showinfo("信息", f"厂家 {manufacturer_name} 没有订单")
                select_window.destroy()
                return
                
            self.export_orders_to_excel(manufacturer_orders, f"{manufacturer_name}_所有订单")
            select_window.destroy()
            messagebox.showinfo("成功", f"已导出厂家 {manufacturer_name} 的 {len(manufacturer_orders)} 个订单")
                
        button_frame = ttk.Frame(select_window)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="导出为Excel", command=do_export).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
        
    def export_manufacturer_unpaid_orders(self):
        """导出定制厂家未结账订单"""
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择厂家")
        select_window.geometry("450x450")
        
        ttk.Label(select_window, text="请选择厂家:", font=("Arial", 12)).pack(pady=10)
        
        # 搜索框架
        search_frame = ttk.Frame(select_window)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 厂家列表
        manufacturer_listbox = tk.Listbox(select_window)
        manufacturer_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 存储所有厂家名称
        all_manufacturers = list(self.manufacturers.keys())
        
        # 填充厂家列表
        def update_manufacturers_list(search_text=""):
            manufacturer_listbox.delete(0, tk.END)
            for manufacturer_name in all_manufacturers:
                if search_text.lower() in manufacturer_name.lower():
                    manufacturer_listbox.insert(tk.END, manufacturer_name)
        
        # 初始填充
        update_manufacturers_list()
        
        # 绑定搜索事件
        def on_search_change(*args):
            update_manufacturers_list(search_var.get())
        
        search_var.trace('w', on_search_change)
            
        def do_export():
            selection = manufacturer_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个厂家")
                return
                
            manufacturer_name = manufacturer_listbox.get(selection[0])
            manufacturer_unpaid_orders = [order for order in self.orders.values() 
                                        if order["manufacturer"] == manufacturer_name and not order["paid"]]
            
            if not manufacturer_unpaid_orders:
                messagebox.showinfo("信息", f"厂家 {manufacturer_name} 没有未结账订单")
                select_window.destroy()
                return
                
            self.export_orders_to_excel(manufacturer_unpaid_orders, f"{manufacturer_name}_未结账订单")
            select_window.destroy()
            messagebox.showinfo("成功", f"已导出厂家 {manufacturer_name} 的 {len(manufacturer_unpaid_orders)} 个未结账订单")
                
        button_frame = ttk.Frame(select_window)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="导出为Excel", command=do_export).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
        
    def export_monthly_summary(self):
        """导出月度汇总数据"""
        self.show_period_selection_dialog("月度", "%Y-%m")
        
    def export_quarterly_summary(self):
        """导出季度汇总数据"""
        self.show_period_selection_dialog("季度", "%Y-Q")
        
    def export_yearly_summary(self):
        """导出年度汇总数据"""
        self.show_period_selection_dialog("年度", "%Y")
        
    def show_period_selection_dialog(self, period_name, date_format):
        """显示周期选择对话框"""
        from datetime import datetime, timedelta
        import calendar
        
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title(f"选择{period_name}")
        select_window.geometry("400x300")
        select_window.resizable(False, False)
        
        # 居中显示
        select_window.transient(self.root)
        select_window.grab_set()
        
        # 标题
        title_frame = ttk.Frame(select_window)
        title_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(title_frame, text=f"请选择要导出的{period_name}", 
                 font=("Arial", 12, "bold")).pack()
        
        # 主要内容区域
        content_frame = ttk.LabelFrame(select_window, text=f"{period_name}选择", padding=15)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        current_year = datetime.now().year
        selected_period = tk.StringVar()
        
        if period_name == "月度":
            # 月度选择
            year_frame = ttk.Frame(content_frame)
            year_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(year_frame, text="年份:").pack(side=tk.LEFT)
            year_var = tk.StringVar(value=str(current_year))
            year_combo = ttk.Combobox(year_frame, textvariable=year_var, width=10, state="readonly")
            year_combo['values'] = [str(year) for year in range(current_year-5, current_year+2)]
            year_combo.pack(side=tk.LEFT, padx=5)
            
            month_frame = ttk.Frame(content_frame)
            month_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(month_frame, text="月份:").pack(side=tk.LEFT)
            month_var = tk.StringVar(value=str(datetime.now().month))
            month_combo = ttk.Combobox(month_frame, textvariable=month_var, width=10, state="readonly")
            month_combo['values'] = [f"{i:02d}" for i in range(1, 13)]
            month_combo.pack(side=tk.LEFT, padx=5)
            
            def get_selected_period():
                return f"{year_var.get()}-{month_var.get().zfill(2)}"
                
        elif period_name == "季度":
            # 季度选择
            year_frame = ttk.Frame(content_frame)
            year_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(year_frame, text="年份:").pack(side=tk.LEFT)
            year_var = tk.StringVar(value=str(current_year))
            year_combo = ttk.Combobox(year_frame, textvariable=year_var, width=10, state="readonly")
            year_combo['values'] = [str(year) for year in range(current_year-5, current_year+2)]
            year_combo.pack(side=tk.LEFT, padx=5)
            
            quarter_frame = ttk.Frame(content_frame)
            quarter_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(quarter_frame, text="季度:").pack(side=tk.LEFT)
            quarter_var = tk.StringVar(value="1")
            quarter_combo = ttk.Combobox(quarter_frame, textvariable=quarter_var, width=10, state="readonly")
            quarter_combo['values'] = [
                "1 (第一季度: 1-3月)",
                "2 (第二季度: 4-6月)", 
                "3 (第三季度: 7-9月)",
                "4 (第四季度: 10-12月)"
            ]
            quarter_combo.pack(side=tk.LEFT, padx=5)
            
            def get_selected_period():
                quarter_num = quarter_var.get().split()[0]
                return f"{year_var.get()}-Q{quarter_num}"
                
        else:  # 年度
            # 年度选择
            year_frame = ttk.Frame(content_frame)
            year_frame.pack(fill=tk.X, pady=5)
            
            ttk.Label(year_frame, text="年份:").pack(side=tk.LEFT)
            year_var = tk.StringVar(value=str(current_year))
            year_combo = ttk.Combobox(year_frame, textvariable=year_var, width=10, state="readonly")
            year_combo['values'] = [str(year) for year in range(current_year-10, current_year+2)]
            year_combo.pack(side=tk.LEFT, padx=5)
            
            def get_selected_period():
                return year_var.get()
        
        # 显示可用数据提示
        info_frame = ttk.Frame(content_frame)
        info_frame.pack(fill=tk.X, pady=10)
        
        # 统计可用数据
        available_periods = set()
        for order in self.orders.values():
            try:
                order_date = datetime.strptime(order['date'], "%Y-%m-%d %H:%M:%S")
                if date_format == "%Y-Q":
                    quarter = (order_date.month - 1) // 3 + 1
                    period_key = f"{order_date.year}-Q{quarter}"
                else:
                    period_key = order_date.strftime(date_format)
                available_periods.add(period_key)
            except ValueError:
                continue
        
        info_text = f"当前系统中可用的{period_name}数据：\n{', '.join(sorted(available_periods)) if available_periods else '无数据'}"
        info_label = ttk.Label(info_frame, text=info_text, font=("Arial", 9), foreground="gray")
        info_label.pack()
        
        # 按钮区域
        button_frame = ttk.Frame(select_window)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def do_export():
            selected = get_selected_period()
            select_window.destroy()
            self.export_specific_period(period_name, date_format, selected)
            
        def cancel_export():
            select_window.destroy()
            
        ttk.Button(button_frame, text="导出", command=do_export).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="取消", command=cancel_export).pack(side=tk.RIGHT)
        
        # 设置窗口居中
        select_window.update_idletasks()
        x = (select_window.winfo_screenwidth() // 2) - (select_window.winfo_width() // 2)
        y = (select_window.winfo_screenheight() // 2) - (select_window.winfo_height() // 2)
        select_window.geometry(f"+{x}+{y}")
    
    def export_specific_period(self, period_name, date_format, selected_period):
        """导出指定周期的数据"""
        from datetime import datetime
        
        # 筛选符合条件的订单
        filtered_orders = []
        for order in self.orders.values():
            try:
                order_date = datetime.strptime(order['date'], "%Y-%m-%d %H:%M:%S")
                
                # 根据格式生成周期标识
                if date_format == "%Y-Q":
                    quarter = (order_date.month - 1) // 3 + 1
                    period_key = f"{order_date.year}-Q{quarter}"
                else:
                    period_key = order_date.strftime(date_format)
                    
                if period_key == selected_period:
                    filtered_orders.append(order)
                    
            except ValueError:
                continue
        
        if not filtered_orders:
            messagebox.showinfo("提示", f"所选{period_name} {selected_period} 中没有找到订单数据")
            return
            
        # 使用原有的导出方法，但只处理筛选后的订单
        self.export_period_summary_with_data(f"{selected_period} {period_name}", date_format, 
                                           {selected_period: self.calculate_period_stats(filtered_orders)}, 
                                           filtered_orders)
    
    def calculate_period_stats(self, orders):
        """计算周期统计数据"""
        stats = {
            'count': len(orders),
            'total_area': sum(order['total_area'] for order in orders),
            'total_price': sum(order['total_price'] for order in orders),
            'paid_count': sum(1 for order in orders if order['paid']),
            'unpaid_count': sum(1 for order in orders if not order['paid'])
        }
        return stats

    def export_period_summary_with_data(self, period_name, date_format, period_data, filtered_orders):
        """Export periodic summary data with given data, including visual dashboard"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, LineChart, Reference
            from openpyxl.chart.label import DataLabelList
            from datetime import datetime
            import calendar
            
            # 创建工作簿
            wb = Workbook()
            
            # 第一个工作表：可视化仪表盘
            summary_ws = wb.active
            summary_ws.title = f"{period_name}仪表盘"
            
            # 添加主标题
            summary_ws['A1'] = f'{period_name}订单汇总可视化仪表盘'
            summary_ws.merge_cells('A1:H1')
            summary_ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
            summary_ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            summary_ws.row_dimensions[1].height = 35
            
            # 添加副标题
            summary_ws['A2'] = f'数据统计时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M")}'
            summary_ws.merge_cells('A2:H2')
            summary_ws['A2'].font = Font(size=11, color="666666", italic=True)
            summary_ws['A2'].alignment = Alignment(horizontal='center')
            summary_ws.row_dimensions[2].height = 20
            
            # 计算汇总统计数据
            total_area = sum(order['total_area'] for order in filtered_orders)
            total_price = sum(order['total_price'] for order in filtered_orders)
            total_paid = sum(1 for order in filtered_orders if order['paid'])
            total_unpaid = len(filtered_orders) - total_paid
            
            # 关键指标概览 (A4:H7)
            summary_ws['A4'] = f'{period_name}关键指标概览'
            summary_ws['A4'].font = Font(bold=True, size=14, color="1F4E79")
            summary_ws.merge_cells('A4:H4')
            summary_ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
            summary_ws.row_dimensions[4].height = 25
            
            # 指标卡片 - 优化布局和颜色
            total_orders = len(filtered_orders)  # 计算订单总数
            metrics = [
                ('订单总数', total_orders, 'A5', 'B5', "4472C4"),
                ('总面积(㎡)', f'{total_area:.2f}', 'C5', 'D5', "70AD47"),
                ('总金额(元)', f'{total_price:.2f}', 'E5', 'F5', "E74C3C"),
                ('已结账', total_paid, 'G5', 'H5', "28A745"),
                ('未结账', total_unpaid, 'A6', 'B6', "DC3545"),
                ('平均单价', f'{total_price/total_area:.2f}' if total_area > 0 else '0', 'C6', 'D6', "FFC107"),
                (f'{period_name}数量', len(period_data), 'E6', 'F6', "6F42C1")
            ]
            
            for metric_name, value, cell1, cell2, color in metrics:
                # 标签单元格
                summary_ws[cell1] = metric_name
                summary_ws[cell1].font = Font(bold=True, size=11, color="FFFFFF")
                summary_ws[cell1].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                summary_ws[cell1].alignment = Alignment(horizontal='center', vertical='center')
                summary_ws[cell1].border = Border(
                    left=Side(style='thin', color='FFFFFF'),
                    right=Side(style='thin', color='FFFFFF'),
                    top=Side(style='thin', color='FFFFFF'),
                    bottom=Side(style='thin', color='FFFFFF')
                )
                
                # 数值单元格
                summary_ws[cell2] = value
                summary_ws[cell2].font = Font(bold=True, size=14, color=color)
                summary_ws[cell2].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                summary_ws[cell2].alignment = Alignment(horizontal='center', vertical='center')
                summary_ws[cell2].border = Border(
                    left=Side(style='thin', color=color),
                    right=Side(style='thin', color=color),
                    top=Side(style='thin', color=color),
                    bottom=Side(style='thin', color=color)
                )
            
            # 设置指标卡片行高
            summary_ws.row_dimensions[5].height = 30
            summary_ws.row_dimensions[6].height = 30
            
            # 结账状态饼图数据区域 (A8:C11)
            if total_paid + total_unpaid > 0:
                summary_ws['A8'] = '结账状态统计'
                summary_ws['A8'].font = Font(bold=True, size=12, color="1F4E79")
                summary_ws.row_dimensions[8].height = 25
                
                # 结账状态表头
                summary_ws['A9'] = '状态'
                summary_ws['B9'] = '数量'
                summary_ws['C9'] = '占比'
                
                for col in ['A9', 'B9', 'C9']:
                    summary_ws[col].font = Font(bold=True, size=10, color="FFFFFF")
                    summary_ws[col].fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
                    summary_ws[col].alignment = Alignment(horizontal='center', vertical='center')
                    summary_ws[col].border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # 结账状态数据
                summary_ws['A10'] = '已结账'
                summary_ws['B10'] = total_paid
                summary_ws['C10'] = f'{total_paid/(total_paid+total_unpaid)*100:.1f}%' if (total_paid+total_unpaid) > 0 else '0%'
                
                summary_ws['A11'] = '未结账'
                summary_ws['B11'] = total_unpaid
                summary_ws['C11'] = f'{total_unpaid/(total_paid+total_unpaid)*100:.1f}%' if (total_paid+total_unpaid) > 0 else '0%'
                
                # 美化数据行
                for row in [10, 11]:
                    for col in ['A', 'B', 'C']:
                        cell = summary_ws[f'{col}{row}']
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        if row == 10:  # 已结账行
                            cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
                        else:  # 未结账行
                            cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                
                # 创建结账状态饼图 - 位置：右上角
                pie_chart = PieChart()
                pie_chart.title = f"{period_name}结账状态分布"
                pie_chart.width = 10
                pie_chart.height = 8
                
                data = Reference(summary_ws, min_col=2, min_row=10, max_row=11)
                labels = Reference(summary_ws, min_col=1, min_row=10, max_row=11)
                pie_chart.add_data(data)
                pie_chart.set_categories(labels)
                
                # 设置饼图样式
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showVal = True
                pie_chart.dataLabels.showCatName = True
                
                # 将饼图放在右上角，与数据表保持距离
                summary_ws.add_chart(pie_chart, "J8")
            
            # 周期趋势数据区域 (A20开始)
            if period_data:
                summary_ws['A20'] = f'{period_name}趋势分析'
                summary_ws['A20'].font = Font(bold=True, size=14, color="1F4E79")
                summary_ws.row_dimensions[20].height = 25
                
                # 表头 - 优化样式
                trend_headers = ['周期', '订单数量', '总面积(㎡)', '总金额(元)', '已结账', '未结账']
                for col, header in enumerate(trend_headers, 1):
                    cell = summary_ws.cell(row=21, column=col, value=header)
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin', color='FFFFFF'),
                        right=Side(style='thin', color='FFFFFF'),
                        top=Side(style='thin', color='FFFFFF'),
                        bottom=Side(style='thin', color='FFFFFF')
                    )
                
                summary_ws.row_dimensions[21].height = 25
                
                # 填充周期数据 - 优化样式
                row = 22
                chart_start_row = row
                row_colors = ["F8F9FA", "E9ECEF"]  # 交替行颜色
                
                for idx, period_key in enumerate(sorted(period_data.keys())):
                    data = period_data[period_key]
                    trend_row_data = [
                        period_key,
                        data['count'],
                        round(data['total_area'], 2),
                        round(data['total_price'], 2),
                        data['paid_count'],
                        data['unpaid_count']
                    ]
                    
                    bg_color = row_colors[idx % 2]
                    
                    for col, value in enumerate(trend_row_data, 1):
                        cell = summary_ws.cell(row=row, column=col, value=value)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                        cell.border = Border(
                            left=Side(style='thin', color='CCCCCC'),
                            right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'),
                            bottom=Side(style='thin', color='CCCCCC')
                        )
                        
                        if col == 1:  # 周期列居中
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = Font(bold=True, size=10)
                        else:  # 数值列右对齐
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.font = Font(size=10)
                    
                    summary_ws.row_dimensions[row].height = 22
                    row += 1
                
                chart_end_row = row - 1
                
                # 创建趋势图表 - 重新设计布局避免重叠
                if len(period_data) > 1:
                    # 金额趋势柱状图 - 位置：趋势表格右侧
                    trend_chart = BarChart()
                    trend_chart.title = f"{period_name}金额趋势分析"
                    trend_chart.width = 15
                    trend_chart.height = 10
                    
                    chart_data = Reference(summary_ws, min_col=4, min_row=21, max_row=chart_end_row)
                    chart_categories = Reference(summary_ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
                    
                    trend_chart.add_data(chart_data, titles_from_data=True)
                    trend_chart.set_categories(chart_categories)
                    
                    trend_chart.x_axis.title = "周期"
                    trend_chart.y_axis.title = "金额(元)"
                    
                    # 将柱状图放在趋势表格右侧
                    summary_ws.add_chart(trend_chart, "H20")
                    
                    # 订单数量趋势线图 - 位置：表格下方，保持足够间距
                    line_chart = LineChart()
                    line_chart.title = f"{period_name}订单数量趋势"
                    line_chart.width = 15
                    line_chart.height = 10
                    
                    line_data = Reference(summary_ws, min_col=2, min_row=21, max_row=chart_end_row)
                    line_categories = Reference(summary_ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
                    
                    line_chart.add_data(line_data, titles_from_data=True)
                    line_chart.set_categories(line_categories)
                    
                    line_chart.x_axis.title = "周期"
                    line_chart.y_axis.title = "订单数量"
                    
                    # 将线图放在表格下方，确保有至少20行的间距避免重叠
                    line_chart_start_row = max(chart_end_row + 20, 45)
                    summary_ws.add_chart(line_chart, f"A{line_chart_start_row}")
            
            # 调整列宽 - 优化图表布局，为图表预留更多空间
            column_widths = {
                'A': 18, 'B': 14, 'C': 14, 'D': 14, 
                'E': 14, 'F': 14, 'G': 14, 'H': 3,
                'I': 3, 'J': 3, 'K': 3, 'L': 3, 'M': 3, 'N': 3, 'O': 3, 'P': 3  # 为图表预留足够的列空间
            }
            for col, width in column_widths.items():
                summary_ws.column_dimensions[col].width = width
            
            # 第二个工作表：符合条件的所有订单详细信息
            detail_ws = wb.create_sheet(title="订单明细")
            
            # 添加主标题
            detail_ws['A1'] = f'{period_name}订单明细列表'
            detail_ws.merge_cells('A1:F1')
            detail_ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            detail_ws['A1'].fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            detail_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            detail_ws.row_dimensions[1].height = 35
            
            # 添加副标题
            detail_ws['A2'] = f'统计时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M")}'
            detail_ws.merge_cells('A2:F2')
            detail_ws['A2'].font = Font(size=10, color="666666", italic=True)
            detail_ws['A2'].alignment = Alignment(horizontal='center')
            detail_ws.row_dimensions[2].height = 18
            
            # 添加表头 - 优化样式
            detail_headers = ['订单名称', '订单面积(㎡)', '订单单价(元/㎡)', '订单总价(元)', '创建时间', '结账状态']
            for col, header in enumerate(detail_headers, 1):
                cell = detail_ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='FFFFFF'),
                    right=Side(style='thin', color='FFFFFF'),
                    top=Side(style='thin', color='FFFFFF'),
                    bottom=Side(style='thin', color='FFFFFF')
                )
            
            detail_ws.row_dimensions[4].height = 28
            
            # 收集符合条件的所有订单
            filtered_orders = []
            for order in self.orders.values():
                try:
                    order_date = datetime.strptime(order['date'], "%Y-%m-%d %H:%M:%S")
                    
                    # 根据格式生成周期标识
                    if date_format == "%Y-Q":
                        quarter = (order_date.month - 1) // 3 + 1
                        period_key = f"{order_date.year}-Q{quarter}"
                    else:
                        period_key = order_date.strftime(date_format)
                    
                    # 如果该订单属于签计的周期范围，则添加到筛选列表
                    if period_key in period_data:
                        filtered_orders.append({
                            'name': order['name'],
                            'total_area': order['total_area'],
                            'unit_price': order['unit_price'],
                            'total_price': order['total_price'],
                            'date': order['date'],
                            'paid': order['paid'],
                            'order_date': order_date  # 用于排序
                        })
                        
                except ValueError:
                    # 如果日期解析失败，跳过该订单
                    continue
            
            # 按创建时间排序（最新的在前）
            filtered_orders.sort(key=lambda x: x['order_date'], reverse=True)
            
            # 填充订单详细信息
            detail_row = 5
            for order in filtered_orders:
                detail_data = [
                    order['name'],
                    round(order['total_area'], 2),
                    order['unit_price'],
                    round(order['total_price'], 2),
                    order['date'],
                    '已结账' if order['paid'] else '未结账'
                ]
                
                for col, value in enumerate(detail_data, 1):
                    cell = detail_ws.cell(row=detail_row, column=col, value=value)
                    
                    # 设置边框
                    cell.border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )
                    
                    # 根据结账状态设置行颜色
                    if order['paid']:
                        cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFEAEA", end_color="FFEAEA", fill_type="solid")
                    
                    # 设置对齐方式和字体
                    if col in [2, 3, 4]:  # 面积、单价、总价列
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.font = Font(size=10)
                    elif col == 5:  # 日期列居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=9)
                    elif col == 6:  # 状态列居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=10, bold=True)
                    else:  # 订单名称列
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.font = Font(size=10)
                
                detail_ws.row_dimensions[detail_row].height = 22
                detail_row += 1
            
            # 添加总计行
            if filtered_orders:
                total_row = detail_row + 1
                detail_ws.cell(row=total_row, column=1, value='总计').font = Font(bold=True)
                detail_ws.cell(row=total_row, column=1).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                
                total_area_sum = sum(order['total_area'] for order in filtered_orders)
                total_price_sum = sum(order['total_price'] for order in filtered_orders)
                
                detail_ws.cell(row=total_row, column=2, value=round(total_area_sum, 2)).font = Font(bold=True)
                detail_ws.cell(row=total_row, column=2).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                detail_ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right')
                
                detail_ws.cell(row=total_row, column=3, value='-').font = Font(bold=True)
                detail_ws.cell(row=total_row, column=3).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                detail_ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='center')
                
                detail_ws.cell(row=total_row, column=4, value=round(total_price_sum, 2)).font = Font(bold=True)
                detail_ws.cell(row=total_row, column=4).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                detail_ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
                
                detail_ws.cell(row=total_row, column=5, value=f'共{len(filtered_orders)}个订单').font = Font(bold=True)
                detail_ws.cell(row=total_row, column=5).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                detail_ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='center')
                
                paid_count = sum(1 for order in filtered_orders if order['paid'])
                unpaid_count = len(filtered_orders) - paid_count
                detail_ws.cell(row=total_row, column=6, value=f'已结:{paid_count} 未结:{unpaid_count}').font = Font(bold=True)
                detail_ws.cell(row=total_row, column=6).fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
                detail_ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='center')
            
            # 调整详细数据工作表列宽
            detail_column_widths = {'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 12}
            for col, width in detail_column_widths.items():
                detail_ws.column_dimensions[col].width = width
                
            # 保存文件
            filename = f"{period_name}可视化汇总数据.xlsx"
            wb.save(filename)
            messagebox.showinfo("成功", f"{period_name}可视化汇总数据已导出到 {filename}\n\n包含以下内容：\n1. {period_name}仪表盘 - 关键指标、结账状态饼图、趋势分析图\n2. 订单明细 - 符合条件的所有订单详细信息")
            
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库以使用导出功能")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            
    def export_orders_to_excel(self, orders, export_type):
        """导出订单列表到Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            
            # 添加汇总工作表
            summary_ws = wb.active
            summary_ws.title = "订单汇总"
            
            # 添加标题
            summary_ws['A1'] = f'{export_type}汇总'
            summary_ws.merge_cells('A1:G1')
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            # 添加表头
            summary_headers = ['订单号', '厂家', '总面积(㎡)', '总价(元)', '状态', '创建日期', '柜体数量']
            for col, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                
            # 填充订单汇总数据
            total_area = 0
            total_price = 0
            row = 4
            
            # 先定义文件名，用于超链接
            filename = f"{export_type}导出.xlsx"
            
            # 为每个订单创建详细工作表（需要先创建工作表才能创建链接）
            order_worksheets = {}
            for order in orders:
                # 创建工作表名称（限制长度并确保唯一性）
                sheet_name = f"订单_{order['name'][:20]}"
                # 如果名称重复，添加序号
                original_name = sheet_name
                counter = 1
                while sheet_name in order_worksheets.values():
                    sheet_name = f"{original_name}_{counter}"
                    counter += 1
                
                ws = wb.create_sheet(title=sheet_name)
                order_worksheets[order['name']] = sheet_name
                
                # 添加订单基本信息
                ws['A1'] = '订单号'
                ws['B1'] = order['name']
                ws['A2'] = '创建日期'
                ws['B2'] = order['date']
                ws['A3'] = '厂家'
                ws['B3'] = order['manufacturer']
                ws['A4'] = '单价(元/㎡)'
                ws['B4'] = order['unit_price']
                ws['A5'] = '总面积(㎡)'
                ws['B5'] = order['total_area']
                ws['A6'] = '总价(元)'
                ws['B6'] = order['total_price']
                ws['A7'] = '状态'
                ws['B7'] = '已结账' if order['paid'] else '未结账'
                
                # 添加空行
                ws.append([])
                
                # 添加柜体信息表头
                ws.append(['房间', '柜体名称', '宽度(mm)', '高度(mm)', '面积(㎡)'])
                
                # 添加柜体数据
                for room_name, room_data in order['rooms'].items():
                    for cabinet_name, cabinet_data in room_data['cabinets'].items():
                        ws.append([
                            room_name,
                            cabinet_name,
                            cabinet_data['width'],
                            cabinet_data['height'],
                            round(cabinet_data['area'], 4)
                        ])
                        
                # 调整列宽
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws.column_dimensions[col].width = 15
            
            # 现在填充汇总表数据，包含超链接
            for order in orders:
                # 计算柜体数量
                cabinet_count = 0
                for room_data in order['rooms'].values():
                    cabinet_count += len(room_data['cabinets'])
                
                # 创建订单号单元格并添加超链接
                order_cell = summary_ws.cell(row=row, column=1, value=order['name'])
                # 创建到对应工作表的超链接（使用标准外部工作簿引用格式）
                sheet_name = order_worksheets[order[name]]
                workbook_name = f"{export_type}.xlsx"
                order_cell.hyperlink = f"{workbook_name}#{sheet_name}!A1"
                # 设置超链接样式
                order_cell.font = Font(color="0000FF", underline="single")
                
                summary_ws.cell(row=row, column=2, value=order['manufacturer'])
                summary_ws.cell(row=row, column=3, value=round(order['total_area'], 2))
                summary_ws.cell(row=row, column=4, value=round(order['total_price'], 2))
                summary_ws.cell(row=row, column=5, value='已结账' if order['paid'] else '未结账')
                summary_ws.cell(row=row, column=6, value=order['date'])
                summary_ws.cell(row=row, column=7, value=cabinet_count)
                
                total_area += order['total_area']
                total_price += order['total_price']
                row += 1
                
            # 添加总计行
            summary_ws.cell(row=row, column=2, value='总计')
            summary_ws.cell(row=row, column=2).font = Font(bold=True)
            summary_ws.cell(row=row, column=3, value=round(total_area, 2))
            summary_ws.cell(row=row, column=3).font = Font(bold=True)
            summary_ws.cell(row=row, column=4, value=round(total_price, 2))
            summary_ws.cell(row=row, column=4).font = Font(bold=True)
                    
            # 调整汇总表列宽
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                summary_ws.column_dimensions[col].width = 15
                
            # 保存文件
            wb.save(filename)
            messagebox.showinfo("成功", f"{export_type}已导出到 {filename}\n\n提示：点击汇总表中的订单号可直接跳转到对应的详细工作表！")
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库以使用导出功能")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            
    def export_order_to_excel(self, order):
        """导出订单到Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            # 先定义文件名
            filename = f"订单_{order['name']}.xlsx"
            
            # 创建工作簿
            wb = Workbook()
            
            # 创建汇总表
            summary_ws = wb.active
            summary_ws.title = "订单汇总"
            
            # 汇总表内容
            summary_ws['A1'] = '单个订单导出'
            summary_ws.merge_cells('A1:G1')
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = '订单号'
            summary_ws['B3'] = '厂家'
            summary_ws['C3'] = '总面积(㎡)'
            summary_ws['D3'] = '总价(元)'
            summary_ws['E3'] = '状态'
            summary_ws['F3'] = '创建日期'
            summary_ws['G3'] = '柜体数量'
            
            # 设置表头样式
            for col in range(1, 8):
                summary_ws.cell(row=3, column=col).font = Font(bold=True)
            
            # 创建详细工作表
            detail_ws = wb.create_sheet(title=f"订单_{order['name'][:20]}")
            
            # 添加订单基本信息
            detail_ws['A1'] = '订单号'
            detail_ws['B1'] = order['name']
            detail_ws['A2'] = '创建日期'
            detail_ws['B2'] = order['date']
            detail_ws['A3'] = '厂家'
            detail_ws['B3'] = order['manufacturer']
            detail_ws['A4'] = '单价(元/㎡)'
            detail_ws['B4'] = order['unit_price']
            detail_ws['A5'] = '总面积(㎡)'
            detail_ws['B5'] = order['total_area']
            detail_ws['A6'] = '总价(元)'
            detail_ws['B6'] = order['total_price']
            detail_ws['A7'] = '状态'
            detail_ws['B7'] = '已结账' if order['paid'] else '未结账'
            
            # 添加返回汇总的链接
            return_cell = detail_ws['A9']
            return_cell.value = '← 返回汇总'
            # 使用标准外部工作簿引用格式
            workbook_name = f"{order[name]}.xlsx"
            return_cell.hyperlink = f"{workbook_name}#订单汇总!A1"
            return_cell.font = Font(color="0000FF", underline="single")
            
            # 添加空行
            detail_ws.append([])
            
            # 添加柜体信息表头
            detail_ws.append(['房间', '柜体名称', '宽度(mm)', '高度(mm)', '面积(㎡)'])
            
            # 添加柜体数据
            for room_name, room_data in order['rooms'].items():
                for cabinet_name, cabinet_data in room_data['cabinets'].items():
                    detail_ws.append([
                        room_name,
                        cabinet_name,
                        cabinet_data['width'],
                        cabinet_data['height'],
                        cabinet_data['area']
                    ])
            
            # 填充汇总表数据并添加超链接
            cabinet_count = sum(len(room_data['cabinets']) for room_data in order['rooms'].values())
            
            # 创建订单号单元格并添加超链接
            order_cell = summary_ws.cell(row=4, column=1, value=order['name'])
            # 使用标准外部工作簿引用格式
            workbook_name = f"{order['name']}.xlsx"
            sheet_name = f"订单_{order['name'][:20]}"
            order_cell.hyperlink = f"{workbook_name}#{sheet_name}!A1"
            order_cell.font = Font(color="0000FF", underline="single")
            
            summary_ws.cell(row=4, column=2, value=order['manufacturer'])
            summary_ws.cell(row=4, column=3, value=round(order['total_area'], 2))
            summary_ws.cell(row=4, column=4, value=round(order['total_price'], 2))
            summary_ws.cell(row=4, column=5, value='已结账' if order['paid'] else '未结账')
            summary_ws.cell(row=4, column=6, value=order['date'])
            summary_ws.cell(row=4, column=7, value=cabinet_count)
            
            # 调整列宽
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                summary_ws.column_dimensions[col].width = 15
                detail_ws.column_dimensions[col].width = 15
            
            # 保存文件
            wb.save(filename)
            messagebox.showinfo("成功", f"订单已导出到 {filename}\n\n提示：\n• 点击汇总表中的订单号可跳转到详细页面\n• 点击详细页面中的'返回汇总'可返回汇总表")
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库以使用导出功能")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def export_order_to_pdf(self, order):
        """导出订单到PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # 注册中文字体
            try:
                font_path = "C:/Windows/Fonts/simhei.ttf"
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('SimHei', font_path))
                    chinese_font = 'SimHei'
                else:
                    chinese_font = 'Helvetica'
            except:
                chinese_font = 'Helvetica'
            
            filename = f"订单_{order['name']}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            
            story = []
            styles = getSampleStyleSheet()
            
            # 标题
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], 
                                       fontName=chinese_font, fontSize=18, alignment=1)
            title = Paragraph(f"定制拆单工作室 - 订单详情", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # 订单基本信息
            basic_info_data = [
                ['订单号', order['name']],
                ['创建日期', order['date']],
                ['厂家', order['manufacturer']],
                ['单价(元/㎡)', f"{order['unit_price']:.2f}"],
                ['总面积(㎡)', f"{order['total_area']:.2f}"],
                ['总价(元)', f"{order['total_price']:.2f}"],
                ['状态', '已结账' if order['paid'] else '未结账']
            ]
            
            basic_table = Table(basic_info_data, colWidths=[2*inch, 4*inch])
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, -1), chinese_font),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(basic_table)
            story.append(Spacer(1, 30))
            
            # 柜体详情
            cabinet_header = Paragraph("柜体详情", 
                                     ParagraphStyle('Header', parent=styles['Normal'],
                                                  fontName=chinese_font, fontSize=14))
            story.append(cabinet_header)
            story.append(Spacer(1, 10))
            
            cabinet_data = [['房间', '柜体名称', '宽度(mm)', '高度(mm)', '面积(㎡)']]
            
            for room_name, room_data in order['rooms'].items():
                for cabinet_name, cabinet_info in room_data['cabinets'].items():
                    cabinet_data.append([
                        room_name, cabinet_name,
                        str(cabinet_info['width']), str(cabinet_info['height']),
                        f"{cabinet_info['area']:.4f}"
                    ])
            
            if len(cabinet_data) > 1:
                cabinet_table = Table(cabinet_data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
                cabinet_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, -1), chinese_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(cabinet_table)
            else:
                story.append(Paragraph("暂无柜体数据", 
                                     ParagraphStyle('NoData', parent=styles['Normal'], fontName=chinese_font)))
            
            doc.build(story)
            messagebox.showinfo("成功", f"订单已导出到 {filename}")
            
        except ImportError:
            messagebox.showerror("错误", "请安装reportlab库以使用PDF导出功能\n\n安装命令： pip install reportlab")
        except Exception as e:
            messagebox.showerror("错误", f"导出PDF失败: {str(e)}")
    
    def export_orders_to_pdf(self, orders, export_type):
        """导出多个订单到PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            try:
                font_path = "C:/Windows/Fonts/simhei.ttf"
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('SimHei', font_path))
                    chinese_font = 'SimHei'
                else:
                    chinese_font = 'Helvetica'
            except:
                chinese_font = 'Helvetica'
            
            filename = f"{export_type}导出.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            
            story = []
            styles = getSampleStyleSheet()
            
            # 标题
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                       fontName=chinese_font, fontSize=18, alignment=1)
            title = Paragraph(f"定制拆单工作室 - {export_type}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # 汇总信息
            total_area = sum(order['total_area'] for order in orders)
            total_price = sum(order['total_price'] for order in orders)
            paid_count = sum(1 for order in orders if order['paid'])
            unpaid_count = len(orders) - paid_count
            
            summary_data = [
                ['订单总数', f'{len(orders)} 个'],
                ['总面积', f'{total_area:.2f} ㎡'],
                ['总金额', f'{total_price:.2f} 元'],
                ['已结账', f'{paid_count} 个'],
                ['未结账', f'{unpaid_count} 个']
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('FONTNAME', (0, 0), (-1, -1), chinese_font),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # 订单列表
            orders_data = [["订单号", "厂家", "面积(㎡)", "总价(元)", "状态", "日期"]]
            
            for order in orders:
                orders_data.append([
                    order['name'][:15] + '...' if len(order['name']) > 15 else order['name'],
                    order['manufacturer'][:10] + '...' if len(order['manufacturer']) > 10 else order['manufacturer'],
                    f"{order['total_area']:.2f}",
                    f"{order['total_price']:.2f}",
                    '已结账' if order['paid'] else '未结账',
                    order['date'][:10]
                ])
            
            orders_table = Table(orders_data, colWidths=[1.5*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch, 1*inch])
            orders_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, -1), chinese_font),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(orders_table)
            
            doc.build(story)
            messagebox.showinfo("成功", f"{export_type}已导出到 {filename}")
            
        except ImportError:
            messagebox.showerror("错误", "请安装reportlab库以使用PDF导出功能\n\n安装命令： pip install reportlab")
        except Exception as e:
            messagebox.showerror("错误", f"导出PDF失败: {str(e)}")
            
    def export_order_to_json(self, order):
        """导出单个订单到JSON文件"""
        try:
            import json
            from datetime import datetime
            
            # 创建导出数据
            export_data = {
                "export_info": {
                    "export_type": "单个订单导出",
                    "export_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_orders": 1
                },
                "order": order
            }
            
            # 生成文件名（处理非法字符）
            safe_order_name = order['name']
            for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
                safe_order_name = safe_order_name.replace(char, '_')
            filename = f"订单_{safe_order_name}.json"
            
            # 写入JSON文件
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("成功", f"订单已导出到 {filename}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出JSON失败: {str(e)}")
            
    def export_orders_to_json(self, orders, export_type):
        """导出多个订单到JSON文件"""
        try:
            import json
            from datetime import datetime
            
            # 计算汇总数据
            total_area = sum(order['total_area'] for order in orders)
            total_price = sum(order['total_price'] for order in orders)
            paid_count = sum(1 for order in orders if order['paid'])
            unpaid_count = len(orders) - paid_count
            
            # 统计厂家分布
            manufacturer_stats = {}
            for order in orders:
                manufacturer = order.get('manufacturer', '未设置')
                if manufacturer not in manufacturer_stats:
                    manufacturer_stats[manufacturer] = {
                        'count': 0,
                        'total_area': 0,
                        'total_price': 0
                    }
                manufacturer_stats[manufacturer]['count'] += 1
                manufacturer_stats[manufacturer]['total_area'] += order['total_area']
                manufacturer_stats[manufacturer]['total_price'] += order['total_price']
            
            # 创建导出数据
            export_data = {
                "export_info": {
                    "export_type": export_type,
                    "export_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_orders": len(orders),
                    "summary": {
                        "total_area": round(total_area, 2),
                        "total_price": round(total_price, 2),
                        "paid_orders": paid_count,
                        "unpaid_orders": unpaid_count,
                        "manufacturer_distribution": {
                            manu: {
                                'count': stats['count'],
                                'total_area': round(stats['total_area'], 2),
                                'total_price': round(stats['total_price'], 2)
                            } for manu, stats in manufacturer_stats.items()
                        }
                    }
                },
                "orders": {order['name']: order for order in orders}
            }
            
            # 生成文件名
            filename = f"{export_type}导出.json"
            
            # 写入JSON文件
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("成功", f"{export_type}已导出到 {filename}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出JSON失败: {str(e)}")
            
    def save_data(self):
        """Save data to file (with cloud sync)"""
        from datetime import datetime
        data = {
            "orders": self.orders,
            "manufacturers": self.manufacturers,
            "bound_order_dir": self.bound_order_dir,
            "timestamp": datetime.now().isoformat(),  # 添加时间戳
            "version": "1.0"  # 数据版本
        }
        try:
            # 使用简化的云同步保存数据
            self.cloud_sync.save_and_sync(data, sync_to_cloud=True)
        except Exception as e:
            messagebox.showerror("错误", f"保存数据失败: {str(e)}")
            
    def save_data_local_only(self):
        """仅保存数据到本地文件，不进行云同步"""
        from datetime import datetime
        data = {
            "orders": self.orders,
            "manufacturers": self.manufacturers,
            "bound_order_dir": self.bound_order_dir,
            "timestamp": datetime.now().isoformat(),  # 添加时间戳
            "version": "1.0"  # 数据版本
        }
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(self.data_file), exist_ok=True)
            
            # 使用简化的云同步保存（仅本地）
            success = self.cloud_sync.save_and_sync(data, sync_to_cloud=False)
            if success:
                print(f"数据已保存到本地: {self.data_file}")
                self.unsaved_changes = False
                return True
            else:
                return False
                
        except Exception as e:
            error_msg = f"保存数据失败: {str(e)}\n文件路径: {self.data_file}"
            print(f"保存数据失败: {error_msg}")
            messagebox.showerror("错误", error_msg)
            return False
            
    def load_data(self):
        """从文件加载数据"""
        try:
            # 使用简化的云同步加载数据
            data = self.cloud_sync.load_and_sync(strategy="latest")
            
            if data:
                self.orders = data.get("orders", {})
                self.manufacturers = data.get("manufacturers", {})
                self.bound_order_dir = data.get("bound_order_dir", "")
                self.local_timestamp = data.get("timestamp", "")  # 保存本地时间戳
            else:
                # 如果本地文件不存在，创建初始数据
                from datetime import datetime
                initial = {
                    "orders": {}, 
                    "manufacturers": {}, 
                    "bound_order_dir": "",
                    "timestamp": datetime.now().isoformat(),
                    "version": "1.0"
                }
                with open(self.data_file, "w", encoding="utf-8") as f:
                    json.dump(initial, f, ensure_ascii=False, indent=2)
                self.orders = {}
                self.manufacturers = {}
                self.bound_order_dir = ""
                self.local_timestamp = initial["timestamp"]
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败: {str(e)}")
            self.local_timestamp = ""
            
    def load_data_local_only(self):
        """Load data from local file only, no cloud sync"""
        try:
            if os.path.exists(self.data_file):
                # 使用简化的云同步，仅本地加载
                data = self.cloud_sync.load_and_sync(strategy="local")
                
                if data:
                    self.orders = data.get("orders", {})
                    self.manufacturers = data.get("manufacturers", {})
                    self.bound_order_dir = data.get("bound_order_dir", "")
                    self.local_timestamp = data.get("timestamp", "")  # 保存本地时间戳
            else:
                # 如果本地文件不存在，创建初始数据
                from datetime import datetime
                initial = {
                    "orders": {}, 
                    "manufacturers": {}, 
                    "bound_order_dir": "",
                    "timestamp": datetime.now().isoformat(),
                    "version": "1.0"
                }
                with open(self.data_file, "w", encoding="utf-8") as f:
                    json.dump(initial, f, ensure_ascii=False, indent=2)
                self.orders = {}
                self.manufacturers = {}
                self.bound_order_dir = ""
                self.local_timestamp = initial["timestamp"]
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败: {str(e)}")
            self.local_timestamp = ""
    
    def on_closing(self):
        """程序关闭时的处理"""
        try:
            # 检查是否有未保存的更改
            current_data = {
                "orders": self.orders,
                "manufacturers": self.manufacturers,
                "bound_order_dir": self.bound_order_dir
            }
            
            # 读取当前本地文件数据
            local_data = None
            if os.path.exists(self.data_file):
                try:
                    with open(self.data_file, "r", encoding="utf-8") as f:
                        local_data = json.load(f)
                except:
                    pass
            
            # 比较数据是否有变化
            data_changed = (local_data != current_data)
            
            if data_changed:
                # 有未保存的更改，询问用户
                result = messagebox.askyesnocancel(
                    "保存更改", 
                    "数据有更改，是否保存并同步到云端？\n\n"
                    "是：保存并同步\n"
                    "否：不保存直接退出\n"
                    "取消：返回程序"
                )
                
                if result is True:  # 用户选择保存
                    self.save_data_with_exit_sync()
                    self.root.destroy()
                elif result is False:  # 用户选择不保存
                    self.root.destroy()
                # else: 用户选择取消，不执行任何操作
            else:
                # 没有更改，但如果有云同步配置，询问是否强制同步
                if self.cloud_sync.github_sync:
                    result = messagebox.askyesno(
                        "退出确认", 
                        "是否强制同步当前数据到云端？\n\n"
                        "是：强制同步后退出\n"
                        "否：直接退出"
                    )
                    if result:
                        self.save_data_with_exit_sync()
                
                self.root.destroy()
                
        except Exception as e:
            print(f"退出时出错: {e}")
            self.root.destroy()
    
    def save_data_with_exit_sync(self):
        """退出时的保存和同步"""
        try:
            from datetime import datetime
            data = {
                "orders": self.orders,
                "manufacturers": self.manufacturers,
                "bound_order_dir": self.bound_order_dir,
                "timestamp": datetime.now().isoformat(),  # 添加时间戳
                "version": "1.0"  # 数据版本
            }
            
            # 强制保存到本地和云端
            self.cloud_sync.save_and_sync(data, sync_to_cloud=True)
            print("退出时数据已保存并同步")
            
        except Exception as e:
            print(f"退出保存时出错: {e}")
            messagebox.showerror("错误", f"保存数据失败: {str(e)}")
    
    def setup_cloud_sync(self):
        """配置云同步"""
        # 创建配置窗口
        config_window = tk.Toplevel(self.root)
        config_window.title("配置GitHub同步")
        config_window.geometry("500x800")
        config_window.transient(self.root)
        config_window.grab_set()
        
        # 创建主框架
        main_frame = ttk.Frame(config_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题和说明
        title_label = ttk.Label(main_frame, text="GitHub同步配置", font=("Arial", 14, "bold"))
        title_label.pack(pady=5)
        
        # 详细说明
        help_text = (
            "使用说明：\n"
            "1. 确保您有GitHub账号\n"
            "2. 创建私有仓库 (Settings → Developer settings → Personal access tokens)\n"
            "3. 生成Token时务必勾选 'repo' 权限\n"
            "4. 仓库名格式：username/repo-name"
        )
        help_label = ttk.Label(main_frame, text=help_text, justify=tk.LEFT, foreground="blue")
        help_label.pack(pady=5)
        
        # GitHub Token
        ttk.Label(main_frame, text="GitHub个人访问令牌:").pack(pady=5)
        token_entry = ttk.Entry(main_frame, width=60, show="*")
        token_entry.pack(pady=5)
        
        # 仓库名
        ttk.Label(main_frame, text="仓库名 (格式: username/repo):").pack(pady=5)
        repo_entry = ttk.Entry(main_frame, width=60)
        repo_entry.pack(pady=5)
        
        # 文件路径
        ttk.Label(main_frame, text="文件路径 (默认: data.json):").pack(pady=5)
        path_entry = ttk.Entry(main_frame, width=60)
        path_entry.insert(0, "data.json")
        path_entry.pack(pady=5)
        
        # 同步选项框架
        sync_options_frame = ttk.LabelFrame(main_frame, text="同步选项", padding=10)
        sync_options_frame.pack(pady=10, fill=tk.X)
        
        # 自动同步选项
        auto_sync_var = tk.BooleanVar(value=False)  # 默认关闭自动同步
        auto_sync_check = ttk.Checkbutton(sync_options_frame, text="启用自动同步 (每5分钟)", variable=auto_sync_var)
        auto_sync_check.pack(anchor=tk.W, pady=2)
        
        # 同步间隔说明
        interval_label = ttk.Label(sync_options_frame, text="注意：启用后会每5分钟自动同步一次", foreground="gray")
        interval_label.pack(anchor=tk.W, pady=2)
        
        def test_connection():
            """测试连接"""
            token = token_entry.get().strip()
            repo = repo_entry.get().strip()
            
            if not token or not repo:
                messagebox.showwarning("警告", "请先填写GitHub Token和仓库名")
                return
            
            try:
                # 临时创建同步对象测试连接
                test_sync = GitHubSync(token, repo)
                result = test_sync.download_data()
                
                if result is not None:
                    messagebox.showinfo("成功", "连接成功！可以正常同步数据。")
                else:
                    # 尝试上传一个测试文件
                    test_data = {"test": "connection"}
                    if test_sync.upload_data(test_data):
                        messagebox.showinfo("成功", "连接成功！可以正常上传数据。")
                        # 删除测试文件
                        requests.delete(
                            f"https://api.github.com/repos/{repo}/contents/data.json",
                            headers=test_sync.headers,
                            json={"message": "删除测试文件", "sha": "test"}
                        )
                    else:
                        messagebox.showwarning("注意", "连接成功，但仓库为空。这是正常现象。")
                        
            except Exception as e:
                messagebox.showerror("错误", f"连接失败: {str(e)}\n\n"
                                         "请检查：\n"
                                         "1. GitHub Token是否正确\n"
                                         "2. 是否勾选了'repo'权限\n"
                                         "3. 仓库名格式是否正确")
        
        def save_config():
            token = token_entry.get().strip()
            repo = repo_entry.get().strip()
            path = path_entry.get().strip() or "data.json"
            
            if not token or not repo:
                messagebox.showwarning("警告", "请填写GitHub Token和仓库名")
                return
            
            try:
                # 设置云同步
                self.cloud_sync.setup_github(token, repo, path)
                self.cloud_sync.auto_sync = auto_sync_var.get()
                
                # 更新状态显示
                if auto_sync_var.get():
                    self.sync_status_label.config(text=f"已配置: {repo}", foreground="green")
                else:
                    self.sync_status_label.config(text=f"已配置: {repo} (手动)", foreground="blue")
                
                messagebox.showinfo("成功", "GitHub同步配置成功！\n\n"
                                         "提示：\n"
                                     "- 可以点击'立即同步'手动上传\n"
                                         "- 程序退出时会询问是否同步")
                config_window.destroy()
                
                # 配置完成后，重新检查云同步状态并继续流程
                self.update_sync_status_display()
                
                # 如果已有云同步配置，继续下载数据和配置厂家
                if self.cloud_sync.github_sync:
                    self.root.after(500, self.continue_after_cloud_sync_config)
                
            except Exception as e:
                messagebox.showerror("错误", f"配置失败: {str(e)}")
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="测试连接", command=test_connection).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="保存", command=save_config).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=config_window.destroy).pack(side=tk.LEFT, padx=5)
    
    def manual_sync(self):
        """Manual sync data (timestamp-based intelligent sync)"""
        # 检查同步权限
        if not self.check_sync_permission("手动同步"):
            return
            
        if not self.cloud_sync.github_sync:
            messagebox.showwarning("警告", "请先配置GitHub同步")
            return
        
        try:
            print("🔄 开始手动数据同步...")
            
            # 添加时间戳到当前数据
            from datetime import datetime
            current_time = datetime.now().isoformat()
            data = {
                "orders": self.orders,
                "manufacturers": self.manufacturers,
                "bound_order_dir": self.bound_order_dir,
                "timestamp": current_time,
                "version": "1.0"
            }
            
            print(f"📊 当前数据时间戳: {current_time}")
            print(f"📈 订单数量: {len(self.orders)}, 厂家数量: {len(self.manufacturers)}")
            
            # 使用简化的智能同步逻辑
            print("🔄 正在执行智能同步...")
            
            # 先保存当前数据到云端
            success = self.cloud_sync.save_and_sync(data, sync_to_cloud=True)
            
            if success:
                print("✅ 数据同步完成！已上传到云端")
                messagebox.showinfo("成功", "数据同步完成！已上传到云端")
            else:
                print("❌ 同步失败")
                messagebox.showerror("错误", "数据同步失败")
                
        except Exception as e:
            print(f"❌ 手动同步失败: {e}")
            messagebox.showerror("错误", f"手动同步失败: {str(e)}")
                
    def import_orders_from_excel(self):
        """从Excel文件导入订单"""
        # 检查导入权限 - 需要写权限
        if not self.check_import_permission():
            return
            
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror("错误", "请安装openpyxl库以使用Excel导入功能\n\n安装命令： pip install openpyxl")
            return
            
        # 选择Excel文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
            
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 解析Excel数据
            imported_orders = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
                if not row[0]:  # 如果订单号为空，跳过
                    continue
                    
                order_data = {
                    "name": str(row[0]) if row[0] else "",
                    "manufacturer": str(row[1]) if row[1] else "",
                    "total_area": float(row[2]) if row[2] else 0,
                    "unit_price": float(row[3]) if row[3] else 0,
                    "total_price": float(row[4]) if row[4] else 0,
                    "paid": bool(row[5]) if len(row) > 5 and row[5] else False,
                    "date": str(row[6]) if len(row) > 6 and row[6] else datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "path": "",
                    "rooms": {}
                }
                imported_orders.append(order_data)
            
            if imported_orders:
                self.process_imported_orders(imported_orders)
            else:
                messagebox.showinfo("信息", "未找到有效的订单数据")
                
        except Exception as e:
            messagebox.showerror("错误", f"导入Excel失败: {str(e)}")
            
    def import_orders_from_json(self):
        """从JSON文件导入订单"""
        # 检查导入权限 - 需要写权限
        if not self.check_import_permission():
            return
            
        # 选择JSON文件
        file_path = filedialog.askopenfilename(
            title="选择JSON文件",
            filetypes=[("JSON文件", "*.json")]
        )
        
        if not file_path:
            return
            
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                
            imported_orders = []
            
            # 如果是完整的数据备份格式
            if "orders" in data:
                imported_orders = list(data["orders"].values())
            # 如果是订单数组
            elif isinstance(data, list):
                imported_orders = data
            # 如果是单个订单对象
            elif isinstance(data, dict) and "name" in data:
                imported_orders = [data]
            else:
                messagebox.showerror("错误", "不支持的JSON格式")
                return
                
            if imported_orders:
                self.process_imported_orders(imported_orders)
            else:
                messagebox.showinfo("信息", "未找到有效的订单数据")
                
        except Exception as e:
            messagebox.showerror("错误", f"导入JSON失败: {str(e)}")
            
    def process_imported_orders(self, imported_orders):
        """处理导入的订单数据"""
        # 检测重复订单（深度对比）
        duplicate_orders = []
        new_orders = []
        
        for order in imported_orders:
            if order["name"] in self.orders:
                existing_order = self.orders[order["name"]]
                # 深度对比订单数据
                if self.compare_order_data(existing_order, order):
                    # 数据相同，跳过
                    continue
                else:
                    # 数据不同，添加到重复列表
                    duplicate_orders.append({
                        "new_order": order,
                        "existing_order": existing_order,
                        "differences": self.get_order_differences(existing_order, order)
                    })
            else:
                new_orders.append(order)
        
        # 如果有重复订单，显示确认对话框
        if duplicate_orders:
            choice = self.show_duplicate_confirmation_with_comparison(duplicate_orders)
            if choice == "cancel":
                return
            elif choice == "keep_existing":
                # 保留现有订单，移除重复的新订单
                pass
            elif choice == "use_new":
                # 使用新订单，添加到新订单列表
                for dup in duplicate_orders:
                    new_orders.append(dup["new_order"])
            elif isinstance(choice, dict):
                # 手动选择保留的订单
                for order_name, keep_new in choice.items():
                    if keep_new:
                        # 找到对应的新订单
                        for dup in duplicate_orders:
                            if dup["new_order"]["name"] == order_name:
                                new_orders.append(dup["new_order"])
                                break
        
        # 显示厂家配置对话框
        manufacturer = self.show_manufacturer_selection(len(new_orders))
        if not manufacturer:
            return
            
        # 应用厂家配置并添加订单
        total_imported = 0
        
        # 处理新订单（包括用户选择的替换订单）
        for order in new_orders:
            order["manufacturer"] = manufacturer
            if manufacturer in self.manufacturers:
                order["unit_price"] = self.manufacturers[manufacturer]["unit_price"]
                order["total_price"] = order["total_area"] * order["unit_price"]
            
            # 确保必要字段存在
            if "path" not in order:
                order["path"] = ""
            if "rooms" not in order:
                order["rooms"] = {}
            if "date" not in order or not order["date"]:
                order["date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
            self.orders[order["name"]] = order
            total_imported += 1
        
        # 保存数据并更新界面
        self.save_data()
        self.update_orders_list()
        self.update_dashboard()
        
        messagebox.showinfo("成功", f"成功导入 {total_imported} 个订单\n\n厂家配置: {manufacturer}")
        
    def compare_order_data(self, existing_order, new_order):
        """深度对比两个订单的数据是否相同"""
        # 对比基本信息
        basic_fields = ["total_area", "unit_price", "total_price", "paid", "manufacturer"]
        for field in basic_fields:
            if existing_order.get(field) != new_order.get(field):
                return False
        
        # 对比房间数据
        existing_rooms = existing_order.get("rooms", {})
        new_rooms = new_order.get("rooms", {})
        
        if set(existing_rooms.keys()) != set(new_rooms.keys()):
            return False
        
        # 对比每个房间的柜体数据
        for room_name in existing_rooms:
            existing_cabinets = existing_rooms[room_name].get("cabinets", {})
            new_cabinets = new_rooms[room_name].get("cabinets", {})
            
            if set(existing_cabinets.keys()) != set(new_cabinets.keys()):
                return False
            
            # 对比每个柜体的详细数据
            for cabinet_name in existing_cabinets:
                existing_cabinet = existing_cabinets[cabinet_name]
                new_cabinet = new_cabinets[cabinet_name]
                
                cabinet_fields = ["width", "height", "area"]
                for field in cabinet_fields:
                    if existing_cabinet.get(field) != new_cabinet.get(field):
                        return False
        
        return True
    
    def get_order_differences(self, existing_order, new_order):
        """获取两个订单之间的差异"""
        differences = []
        
        # 检查基本信息差异
        basic_fields = {
            "total_area": "总面积",
            "unit_price": "单价",
            "total_price": "总价",
            "paid": "结账状态",
            "manufacturer": "厂家"
        }
        
        for field, label in basic_fields.items():
            existing_val = existing_order.get(field)
            new_val = new_order.get(field)
            if existing_val != new_val:
                differences.append(f"{label}: 现有[{existing_val}] → 新的[{new_val}]")
        
        # 检查房间和柜体差异
        existing_rooms = existing_order.get("rooms", {})
        new_rooms = new_order.get("rooms", {})
        
        # 房间数量对比
        if len(existing_rooms) != len(new_rooms):
            differences.append(f"房间数量: 现有[{len(existing_rooms)}] → 新的[{len(new_rooms)}]")
        
        # 房间名称对比
        existing_room_names = set(existing_rooms.keys())
        new_room_names = set(new_rooms.keys())
        if existing_room_names != new_room_names:
            only_existing = existing_room_names - new_room_names
            only_new = new_room_names - existing_room_names
            if only_existing:
                differences.append(f"现有订单独有房间: {list(only_existing)}")
            if only_new:
                differences.append(f"新订单独有房间: {list(only_new)}")
        
        # 对比共同房间的柜体数据
        common_rooms = existing_room_names & new_room_names
        for room_name in common_rooms:
            existing_cabinets = existing_rooms[room_name].get("cabinets", {})
            new_cabinets = new_rooms[room_name].get("cabinets", {})
            
            # 柜体数量对比
            if len(existing_cabinets) != len(new_cabinets):
                differences.append(f"房间[{room_name}]柜体数量: 现有[{len(existing_cabinets)}] → 新的[{len(new_cabinets)}]")
            
            # 柜体名称对比
            existing_cabinet_names = set(existing_cabinets.keys())
            new_cabinet_names = set(new_cabinets.keys())
            if existing_cabinet_names != new_cabinet_names:
                only_existing = existing_cabinet_names - new_cabinet_names
                only_new = new_cabinet_names - existing_cabinet_names
                if only_existing:
                    differences.append(f"房间[{room_name}]现有订单独有柜体: {list(only_existing)}")
                if only_new:
                    differences.append(f"房间[{room_name}]新订单独有柜体: {list(only_new)}")
            
            # 对比共同柜体的详细数据
            common_cabinets = existing_cabinet_names & new_cabinet_names
            for cabinet_name in common_cabinets:
                existing_cabinet = existing_cabinets[cabinet_name]
                new_cabinet = new_cabinets[cabinet_name]
                
                cabinet_fields = {
                    "width": "宽度",
                    "height": "高度",
                    "area": "面积"
                }
                
                for field, label in cabinet_fields.items():
                    existing_val = existing_cabinet.get(field)
                    new_val = new_cabinet.get(field)
                    if existing_val != new_val:
                        differences.append(f"房间[{room_name}]柜体[{cabinet_name}]{label}: 现有[{existing_val}] → 新的[{new_val}]")
        
        return differences
        
    def show_duplicate_confirmation_with_comparison(self, duplicate_orders):
        """显示带数据对比的重复订单确认对话框"""
        confirm_window = tk.Toplevel(self.root)
        confirm_window.title("重复订单数据对比")
        confirm_window.geometry("800x600")
        confirm_window.resizable(True, True)
        
        # 居中显示
        confirm_window.transient(self.root)
        confirm_window.grab_set()
        
        # 标题
        title_label = ttk.Label(confirm_window, text="检测到重复订单且数据不同", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 说明
        info_label = ttk.Label(confirm_window, text="以下订单已存在且数据有差异，请选择处理方式：")
        info_label.pack(pady=5)
        
        # 创建笔记本控件用于显示每个重复订单
        notebook = ttk.Notebook(confirm_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        user_choices = {}  # 存储用户选择
        
        for i, dup_info in enumerate(duplicate_orders):
            existing_order = dup_info["existing_order"]
            new_order = dup_info["new_order"]
            differences = dup_info["differences"]
            order_name = new_order["name"]
            
            # 为每个重复订单创建一个选项卡
            tab_frame = ttk.Frame(notebook)
            notebook.add(tab_frame, text=f"订单{i+1}: {order_name[:15]}..." if len(order_name) > 15 else f"订单{i+1}: {order_name}")
            
            # 订单名称
            name_label = ttk.Label(tab_frame, text=f"订单名称：{order_name}", font=("Arial", 12, "bold"))
            name_label.pack(pady=5)
            
            # 差异信息显示
            diff_frame = ttk.LabelFrame(tab_frame, text="数据差异", padding=10)
            diff_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # 创建文本框显示差异
            diff_text = tk.Text(diff_frame, height=15, wrap=tk.WORD)
            diff_scrollbar = ttk.Scrollbar(diff_frame, orient=tk.VERTICAL, command=diff_text.yview)
            diff_text.configure(yscrollcommand=diff_scrollbar.set)
            
            diff_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            diff_text.pack(fill=tk.BOTH, expand=True)
            
            # 填充差异信息
            if differences:
                diff_text.insert(tk.END, "\n".join(differences))
            else:
                diff_text.insert(tk.END, "数据完全相同")
            
            diff_text.config(state=tk.DISABLED)
            
            # 选择按钮
            choice_frame = ttk.Frame(tab_frame)
            choice_frame.pack(fill=tk.X, padx=10, pady=10)
            
            choice_var = tk.StringVar(value="keep_existing")
            user_choices[order_name] = choice_var
            
            ttk.Radiobutton(choice_frame, text="保留现有订单", variable=choice_var, value="keep_existing").pack(side=tk.LEFT, padx=20)
            ttk.Radiobutton(choice_frame, text="使用新订单", variable=choice_var, value="use_new").pack(side=tk.LEFT, padx=20)
        
        # 全局操作按钮
        global_button_frame = ttk.Frame(confirm_window)
        global_button_frame.pack(fill=tk.X, padx=10, pady=5)
        
        def set_all_keep_existing():
            for choice_var in user_choices.values():
                choice_var.set("keep_existing")
                
        def set_all_use_new():
            for choice_var in user_choices.values():
                choice_var.set("use_new")
        
        ttk.Button(global_button_frame, text="全部保留现有", command=set_all_keep_existing).pack(side=tk.LEFT, padx=5)
        ttk.Button(global_button_frame, text="全部使用新的", command=set_all_use_new).pack(side=tk.LEFT, padx=5)
        
        # 最终确认按钮
        final_button_frame = ttk.Frame(confirm_window)
        final_button_frame.pack(pady=10)
        
        result = {"choice": None}
        
        def confirm_choices():
            # 收集用户选择
            choices = {}
            for order_name, choice_var in user_choices.items():
                choices[order_name] = choice_var.get() == "use_new"
            result["choice"] = choices
            confirm_window.destroy()
            
        def cancel_import():
            result["choice"] = "cancel"
            confirm_window.destroy()
        
        ttk.Button(final_button_frame, text="确认选择", command=confirm_choices).pack(side=tk.LEFT, padx=10)
        ttk.Button(final_button_frame, text="取消导入", command=cancel_import).pack(side=tk.LEFT, padx=10)
        
        # 等待用户选择
        confirm_window.wait_window()
        return result["choice"]
        
    def show_duplicate_confirmation(self, duplicate_orders):
        """显示重复订单确认对话框"""
        confirm_window = tk.Toplevel(self.root)
        confirm_window.title("重复订单确认")
        confirm_window.geometry("500x400")
        confirm_window.resizable(False, False)
        
        # 居中显示
        confirm_window.transient(self.root)
        confirm_window.grab_set()
        
        # 标题
        title_label = ttk.Label(confirm_window, text="检测到重复订单", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 说明
        info_label = ttk.Label(confirm_window, text="以下订单已存在，导入将覆盖现有数据：")
        info_label.pack(pady=5)
        
        # 重复订单列表
        list_frame = ttk.Frame(confirm_window)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        duplicates_listbox = tk.Listbox(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=duplicates_listbox.yview)
        duplicates_listbox.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        duplicates_listbox.pack(fill=tk.BOTH, expand=True)
        
        # 填充重复订单
        for order in duplicate_orders:
            duplicates_listbox.insert(tk.END, f"{order['name']} - {order.get('manufacturer', '未设置')}")
        
        # 按钮
        button_frame = ttk.Frame(confirm_window)
        button_frame.pack(pady=10)
        
        result = {"confirmed": False}
        
        def confirm_import():
            result["confirmed"] = True
            confirm_window.destroy()
            
        def cancel_import():
            result["confirmed"] = False
            confirm_window.destroy()
        
        ttk.Button(button_frame, text="确认覆盖", command=confirm_import).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="取消导入", command=cancel_import).pack(side=tk.LEFT, padx=10)
        
        # 等待用户选择
        confirm_window.wait_window()
        return result["confirmed"]
        
    def show_manufacturer_selection(self, order_count):
        """显示厂家选择对话框"""
        # 创建对话框
        select_window = tk.Toplevel(self.root)
        select_window.title("选择厂家")
        select_window.geometry("600x650")
        select_window.resizable(True, True)
        
        # 居中显示
        select_window.transient(self.root)
        select_window.grab_set()
        
        # 创建主容器
        main_frame = ttk.Frame(select_window, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="一键配置厂家", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        # 说明
        info_label = ttk.Label(main_frame, text=f"请选择要为 {order_count} 个订单配置的厂家")
        info_label.pack(pady=(0, 15))
        
        # 厂家列表框架
        list_label = ttk.Label(main_frame, text="可用厂家：", font=("Arial", 10, "bold"))
        list_label.pack(anchor="w", pady=(0, 5))
        
        # 列表框架
        list_container = ttk.Frame(main_frame)
        list_container.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # 列表框
        manufacturers_listbox = tk.Listbox(list_container, height=12, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=manufacturers_listbox.yview)
        manufacturers_listbox.configure(yscrollcommand=scrollbar.set)
        
        manufacturers_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 填充厂家列表
        if self.manufacturers:
            for name, data in self.manufacturers.items():
                manufacturers_listbox.insert(tk.END, f"{name} - {data['unit_price']}元/㎡")
        else:
            manufacturers_listbox.insert(tk.END, "没有厂家数据")
            manufacturers_listbox.insert(tk.END, "请先在厂家管理中添加厂家")
        
        # 分隔线
        separator = ttk.Separator(main_frame, orient="horizontal")
        separator.pack(fill=tk.X, pady=10)
        
        # 按钮框架
        button_container = ttk.Frame(main_frame)
        button_container.pack(fill=tk.X)
        
        # 用于存储结果
        result = {"manufacturer": None}
        
        def on_confirm():
            if not self.manufacturers:
                messagebox.showwarning("警告", "没有可用的厂家，请先在厂家管理中添加厂家")
                return
            
            selection = manufacturers_listbox.curselection()
            if not selection:
                messagebox.showwarning("警告", "请选择一个厂家")
                return
            
            manufacturer_text = manufacturers_listbox.get(selection[0])
            if "没有厂家数据" in manufacturer_text:
                messagebox.showwarning("警告", "请先在厂家管理中添加厂家")
                return
            
            manufacturer_name = manufacturer_text.split(" - ")[0]
            result["manufacturer"] = manufacturer_name
            select_window.destroy()
        
        def on_cancel():
            result["manufacturer"] = None
            select_window.destroy()
        
        # 创建按钮（使用更大的按钮和更明显的样式）
        confirm_button = tk.Button(
            button_container, 
            text="确定", 
            command=on_confirm,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            width=12,
            height=2
        )
        confirm_button.pack(side=tk.LEFT, padx=(0, 15))
        
        cancel_button = tk.Button(
            button_container, 
            text="取消", 
            command=on_cancel,
            font=("Arial", 12),
            bg="#f44336",
            fg="white",
            width=12,
            height=2
        )
        cancel_button.pack(side=tk.LEFT)
        
        # 等待用户选择
        select_window.wait_window()
        return result["manufacturer"]
        
    def export_partial_orders(self):
        """导出部分订单"""
        if not self.orders:
            messagebox.showwarning("警告", "没有订单数据可导出")
            return
            
        # 创建选择窗口
        select_window = tk.Toplevel(self.root)
        select_window.title("选择部分订单")
        select_window.geometry("800x650")
        
        # 标题
        title_label = ttk.Label(select_window, text="选择要导出的订单", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # 筛选选项
        filter_frame = ttk.LabelFrame(select_window, text="筛选和排序条件", padding=10)
        filter_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # 第一行：状态和厂家筛选
        row1_frame = ttk.Frame(filter_frame)
        row1_frame.pack(fill=tk.X, pady=5)
        
        # 状态筛选
        ttk.Label(row1_frame, text="状态:").pack(side=tk.LEFT)
        status_var = tk.StringVar(value="全部")
        status_combo = ttk.Combobox(row1_frame, textvariable=status_var, 
                                  values=["全部", "未结账", "已结账"], 
                                  state="readonly", width=15)
        status_combo.pack(side=tk.LEFT, padx=5)
        
        # 厂家筛选
        ttk.Label(row1_frame, text="厂家:").pack(side=tk.LEFT, padx=(20, 5))
        manufacturer_var = tk.StringVar(value="全部")
        manufacturer_values = ["全部"] + list(self.manufacturers.keys())
        manufacturer_combo = ttk.Combobox(row1_frame, textvariable=manufacturer_var,
                                        values=manufacturer_values, state="readonly", width=15)
        manufacturer_combo.pack(side=tk.LEFT, padx=5)
        
        # 第二行：时间排序
        row2_frame = ttk.Frame(filter_frame)
        row2_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(row2_frame, text="时间排序:").pack(side=tk.LEFT)
        time_sort_var = tk.StringVar(value="最新在前")
        time_sort_combo = ttk.Combobox(row2_frame, textvariable=time_sort_var,
                                      values=["最新在前", "最旧在前"],
                                      state="readonly", width=15)
        time_sort_combo.pack(side=tk.LEFT, padx=5)
        
        # 订单列表
        list_frame = ttk.LabelFrame(select_window, text="订单列表 (可多选)", padding=5)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # 创建表格
        columns = ("订单号", "厂家", "面积", "总价", "状态", "日期")
        orders_tree = ttk.Treeview(list_frame, columns=columns, show="headings", selectmode="extended")
        
        for col in columns:
            orders_tree.heading(col, text=col)
            orders_tree.column(col, width=100)
        
        # 添加滚动条
        tree_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=orders_tree.yview)
        orders_tree.configure(yscrollcommand=tree_scrollbar.set)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        orders_tree.pack(fill=tk.BOTH, expand=True)
        
        def update_order_list():
            """Update order list based on filter criteria (supports sorting)"""
            # 清空现有项目
            for item in orders_tree.get_children():
                orders_tree.delete(item)
            
            status_filter = status_var.get()
            manufacturer_filter = manufacturer_var.get()
            sort_order = time_sort_var.get()
            
            # 收集符合条件的订单
            filtered_orders = []
            for order in self.orders.values():
                # 状态筛选
                if status_filter == "未结账" and order["paid"]:
                    continue
                elif status_filter == "已结账" and not order["paid"]:
                    continue
                
                # 厂家筛选
                if manufacturer_filter != "全部" and order["manufacturer"] != manufacturer_filter:
                    continue
                
                filtered_orders.append(order)
            
            # 按时间排序
            try:
                if sort_order == "最新在前":
                    filtered_orders.sort(key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d %H:%M:%S"), reverse=True)
                else:
                    filtered_orders.sort(key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d %H:%M:%S"), reverse=False)
            except ValueError:
                # 如果日期格式有问题，按订单名排序
                filtered_orders.sort(key=lambda x: x["name"])
            
            # 添加到列表
            for order in filtered_orders:
                orders_tree.insert("", tk.END, values=(
                    order["name"],
                    order["manufacturer"],
                    f"{order['total_area']:.2f}",
                    f"{order['total_price']:.2f}",
                    "已结账" if order["paid"] else "未结账",
                    order["date"][:10]
                ))
        
        # 绑定筛选条件变化事件
        status_combo.bind("<<ComboboxSelected>>", lambda e: update_order_list())
        manufacturer_combo.bind("<<ComboboxSelected>>", lambda e: update_order_list())
        time_sort_combo.bind("<<ComboboxSelected>>", lambda e: update_order_list())
        
        # 初始填充列表
        update_order_list()
        
        # 操作按钮
        button_frame = ttk.Frame(select_window)
        button_frame.pack(pady=10)
        
        # 全选/取消全选按钮
        def select_all():
            for item in orders_tree.get_children():
                orders_tree.selection_add(item)
                
        def deselect_all():
            orders_tree.selection_remove(orders_tree.selection())
        
        ttk.Button(button_frame, text="全选", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消全选", command=deselect_all).pack(side=tk.LEFT, padx=5)
        
        # 导出按钮
        def do_export_excel():
            selections = orders_tree.selection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
            
            selected_orders = []
            for item in selections:
                values = orders_tree.item(item, "values")
                order_name = values[0]
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_excel(selected_orders, "部分订单")
                select_window.destroy()
                messagebox.showinfo("成功", f"已导出 {len(selected_orders)} 个订单")
        
        def do_export_pdf():
            selections = orders_tree.selection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
            
            selected_orders = []
            for item in selections:
                values = orders_tree.item(item, "values")
                order_name = values[0]
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_pdf(selected_orders, "部分订单")
                select_window.destroy()
                
        def do_export_json():
            selections = orders_tree.selection()
            if not selections:
                messagebox.showwarning("警告", "请至少选择一个订单")
                return
            
            selected_orders = []
            for item in selections:
                values = orders_tree.item(item, "values")
                order_name = values[0]
                if order_name in self.orders:
                    selected_orders.append(self.orders[order_name])
            
            if selected_orders:
                self.export_orders_to_json(selected_orders, "部分订单")
                select_window.destroy()
        
        ttk.Button(button_frame, text="导出为Excel", command=do_export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为PDF", command=do_export_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="导出为JSON", command=do_export_json).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=select_window.destroy).pack(side=tk.LEFT, padx=5)
                
    def export_period_summary(self, period_name, date_format):
        """导出所有周期性汇总数据"""
        from datetime import datetime
        
        # 按周期分组统计所有数据
        period_data = {}
        all_orders = []
        
        for order in self.orders.values():
            try:
                order_date = datetime.strptime(order['date'], "%Y-%m-%d %H:%M:%S")
                
                # 根据格式生成周期标识
                if date_format == "%Y-Q":
                    quarter = (order_date.month - 1) // 3 + 1
                    period_key = f"{order_date.year}-Q{quarter}"
                else:
                    period_key = order_date.strftime(date_format)
                    
                if period_key not in period_data:
                    period_data[period_key] = {
                        'count': 0,
                        'total_area': 0,
                        'total_price': 0,
                        'paid_count': 0,
                        'unpaid_count': 0
                    }
                    
                period_data[period_key]['count'] += 1
                period_data[period_key]['total_area'] += order['total_area']
                period_data[period_key]['total_price'] += order['total_price']
                
                if order['paid']:
                    period_data[period_key]['paid_count'] += 1
                else:
                    period_data[period_key]['unpaid_count'] += 1
                    
                all_orders.append(order)
                    
            except ValueError:
                continue
                
        if not period_data:
            messagebox.showinfo("提示", f"没有找到{period_name}数据")
            return
            
        # 使用新的导出方法
        self.export_period_summary_with_data(f"所有{period_name}", date_format, period_data, all_orders)

    def create_statistics_tab(self):
        pass
        
    def update_profit_calendar(self):
        """更新盈利日历显示"""
        try:
            from datetime import datetime, timedelta
            import calendar
            
            now = datetime.now()
            current_year = int(self.calendar_year_var.get()) if hasattr(self, 'calendar_year_var') else now.year
            current_month = int(self.calendar_month_var.get()) if hasattr(self, 'calendar_month_var') else now.month
            
            # 获取当月第一天和最后一天
            first_day = datetime(current_year, current_month, 1)
            if current_month == 12:
                last_day = datetime(current_year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = datetime(current_year, current_month + 1, 1) - timedelta(days=1)
            
            # 获取当月第一天是星期几 (0=Monday, 6=Sunday)
            first_weekday = first_day.weekday()
            # 转换为日历格式 (0=Sunday, 6=Saturday)
            first_weekday = (first_weekday + 1) % 7
            
            # 计算每日盈利数据
            daily_profits = {}
            if hasattr(self, 'orders') and self.orders:
                # 处理订单数据，支持字典和列表两种格式
                orders_data = self.orders.values() if isinstance(self.orders, dict) else self.orders
                for order in orders_data:
                    try:
                        # 尝试不同的日期格式
                        date_str = order.get('date', '')
                        if not date_str:
                            continue
                        
                        # 处理不同的日期格式
                        if ' ' in date_str:  # 包含时间的格式
                            order_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        else:  # 只有日期的格式
                            order_date = datetime.strptime(date_str, '%Y-%m-%d')
                        
                        if (order_date.year == current_year and 
                            order_date.month == current_month):
                            day = order_date.day
                            if day not in daily_profits:
                                daily_profits[day] = 0
                            # 尝试不同的金额字段名
                            amount = order.get('total_price', order.get('amount', 0))
                            daily_profits[day] += float(amount)
                    except (ValueError, KeyError, TypeError):
                        continue
            
            # 清空所有按钮
            for (week, day), btn in self.calendar_buttons.items():
                btn.config(text="", bg="#f0f0f0", state=tk.DISABLED)
            
            # 填充日历
            days_in_month = last_day.day
            for day in range(1, days_in_month + 1):
                # 计算在网格中的位置
                total_days = first_weekday + day - 1
                week = total_days // 7
                weekday = total_days % 7
                
                if week < 6 and (week, weekday) in self.calendar_buttons:
                    btn = self.calendar_buttons[(week, weekday)]
                    
                    # 根据盈利情况设置颜色和显示文本
                    profit = daily_profits.get(day, 0)
                    
                    # 构建显示文本：日期 + 金额（红色）
                    if profit > 0:
                        # 格式化金额显示
                        if profit >= 10000:
                            amount_text = f"¥{profit/10000:.1f}万"
                        elif profit >= 1000:
                            amount_text = f"¥{profit/1000:.1f}k"
                        else:
                            amount_text = f"¥{profit:.0f}"
                        
                        display_text = f"{day}\n{amount_text}"
                        
                        if profit >= 2000:  # 高盈利
                            btn.config(text=display_text, bg="#4CAF50", fg="#D32F2F", state=tk.NORMAL, 
                                      relief=tk.RAISED, bd=3, highlightbackground="#2E7D32")  # 深绿背景，红色字体
                        elif profit >= 1000:  # 中等盈利
                            btn.config(text=display_text, bg="#8BC34A", fg="#D32F2F", state=tk.NORMAL,
                                      relief=tk.RAISED, bd=3, highlightbackground="#689F38")  # 浅绿背景，红色字体
                        else:  # 低盈利
                            btn.config(text=display_text, bg="#FFC107", fg="#D32F2F", state=tk.NORMAL,
                                      relief=tk.RAISED, bd=3, highlightbackground="#F57C00")  # 黄色背景，红色字体
                    else:
                        btn.config(text=str(day), bg="#FAFAFA", fg="#424242", state=tk.NORMAL,
                                  relief=tk.RAISED, bd=2, highlightbackground="#E0E0E0")  # 无盈利，浅灰背景
                    
                    # 添加点击事件显示详细信息
                    btn.config(command=lambda d=day, p=profit: self.show_day_profit_detail(d, p))
                    
        except Exception as e:
            print(f"更新盈利日历失败: {e}")
    
    def show_day_profit_detail(self, day, profit):
        """Show daily profit details"""
        try:
            from datetime import datetime
            
            now = datetime.now()
            year = int(self.calendar_year_var.get()) if hasattr(self, 'calendar_year_var') else now.year
            month = int(self.calendar_month_var.get()) if hasattr(self, 'calendar_month_var') else now.month
            target_date = datetime(year, month, day)
            
            # 获取当日订单
            day_orders = []
            if hasattr(self, 'orders') and self.orders:
                # 处理订单数据，支持字典和列表两种格式
                orders_data = self.orders.values() if isinstance(self.orders, dict) else self.orders
                for order in orders_data:
                    try:
                        # 尝试不同的日期格式
                        date_str = order.get('date', '')
                        if not date_str:
                            continue
                        
                        # 处理不同的日期格式
                        if ' ' in date_str:  # 包含时间的格式
                            order_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        else:  # 只有日期的格式
                            order_date = datetime.strptime(date_str, '%Y-%m-%d')
                        
                        if order_date.date() == target_date.date():
                            day_orders.append(order)
                    except (ValueError, KeyError, TypeError):
                        continue
            
            # 创建详情窗口
            detail_window = tk.Toplevel(self.root)
            detail_window.title(f"{target_date.strftime('%Y年%m月%d日')} 盈利详情")
            detail_window.geometry("400x500")
            detail_window.resizable(False, False)
            
            # 居中显示
            detail_window.transient(self.root)
            detail_window.grab_set()
            
            # 标题
            title_label = ttk.Label(detail_window, 
                                  text=f"{target_date.strftime('%Y年%m月%d日')} 盈利详情",
                                  font=('Arial', 12, 'bold'))
            title_label.pack(pady=10)
            
            # 盈利汇总
            summary_frame = ttk.LabelFrame(detail_window, text="当日汇总")
            summary_frame.pack(fill=tk.X, padx=10, pady=5)
            
            ttk.Label(summary_frame, text=f"订单数量: {len(day_orders)}").pack(anchor=tk.W, padx=5, pady=2)
            ttk.Label(summary_frame, text=f"总金额: ¥{profit:.2f}").pack(anchor=tk.W, padx=5, pady=2)
            
            # 订单列表
            if day_orders:
                orders_frame = ttk.LabelFrame(detail_window, text="订单列表")
                orders_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
                
                # 创建滚动文本框
                text_frame = tk.Frame(orders_frame)
                text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                
                scrollbar = ttk.Scrollbar(text_frame)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                text_widget = tk.Text(text_frame, yscrollcommand=scrollbar.set, 
                                    wrap=tk.WORD, height=8)
                text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.config(command=text_widget.yview)
                
                # 填充订单信息
                for i, order in enumerate(day_orders, 1):
                     # 尝试不同的字段名
                     name = order.get('order_name', order.get('name', 'N/A'))
                     amount = order.get('total_price', order.get('amount', 0))
                     manufacturer = order.get('manufacturer', 'N/A')
                     area = order.get('total_area', order.get('area', 'N/A'))
                     
                     order_info = f"{i}. {name} - ¥{amount}\n"
                     order_info += f"   厂家: {manufacturer}\n"
                     order_info += f"   面积: {area}\n\n"
                     text_widget.insert(tk.END, order_info)
                
                text_widget.config(state=tk.DISABLED)
            
            # 关闭按钮
            ttk.Button(detail_window, text="关闭", 
                      command=detail_window.destroy).pack(pady=10)
                      
        except Exception as e:
            print(f"显示日盈利详情失败: {e}")
            messagebox.showerror("错误", f"显示详情失败: {e}")
    
    def generate_chart(self):
        pass
    
    def update_average_data(self):
        pass
    
    def export_chart_png(self):
        pass
    
    def export_chart_pdf(self):
        pass

if __name__ == "__main__":
    try:
        print("程序启动中...")
        import sys
        print(f"Python版本: {sys.version}")
        print(f"是否为编译后的exe: {hasattr(sys, 'frozen')}")
        
        print("创建主窗口...")
        root = tk.Tk()
        print("创建应用实例...")
        app = CustomOrderManagementApp(root)
        print("启动主循环...")
        root.mainloop()
        print("程序正常退出")
    except Exception as e:
        print(f"程序启动失败: {e}")
        import traceback
        traceback.print_exc()
        input("按回车键退出...")
    except KeyboardInterrupt:
        print("程序被用户中断")
    finally:
        print("程序结束")
